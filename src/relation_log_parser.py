"""Relation moshell log parser adapted from REL_PARSED.

This helper keeps the GUI integration lightweight by reusing the existing
parsing rules without importing the PyQt-based tooling directly.
"""

from __future__ import annotations

import os
import re
from collections import Counter
from typing import Callable, Iterable, Optional

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LogCallback = Optional[Callable[[str], None]]

_ILLEGAL_CHAR_PATTERN = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
_PROMPT_PATTERN = re.compile(r"^[A-Z0-9_\-]{4,180}>(.*?)$", re.IGNORECASE)
_RUN_SCRIPT_PATTERN = re.compile(r".*?run\s+.*?\$nodename\_(.*?).mos$", re.IGNORECASE)
_TRUNI_PATTERN = re.compile(r"^(trun|truni)\s+(.*?)$", re.IGNORECASE)
_TAG_PATTERN = re.compile(r'^!!!!.*?TAG\s+:"(.*?)".*?$', re.IGNORECASE)
_ALT_TAG_PATTERN = re.compile(r"^>>>\s+(\[.*?\]).*$", re.IGNORECASE)
_SET_PATTERN = re.compile(r"^(SET\s+.*?)$", re.IGNORECASE)
_CREATE_PATTERN = re.compile(r"^(CREATE.*?)$", re.IGNORECASE)
_DELETE_PATTERN = re.compile(r"^(DELETE.*?)$", re.IGNORECASE)
_TAG_ERROR_PATTERNS = [
    re.compile(r"^(ERROR:.*?:.*?)$", re.IGNORECASE),
    re.compile(r"(!!!!\s+(ERROR|Processing):.*?)$", re.IGNORECASE),
    re.compile(r"(!!!!\s+Processing.*?)$", re.IGNORECASE),
    re.compile(r"^(>>>.*?:.*?)$", re.IGNORECASE),
    re.compile(r"^(Total.*?MOs\s+attempted.*?)$", re.IGNORECASE),
]
_TOTAL_ATTEMPTED_PATTERN = re.compile(
    r"^Total.*?MOs\s+attempted,\s+([0-9]{1,5})\s+MOs\s+(.*?)$",
    re.IGNORECASE,
)


def _log(log_cb: LogCallback, message: str) -> None:
    if log_cb:
        log_cb(message)


def _category_checking1(tag_result: str) -> tuple[str, str]:
    if "Proxy ID" in tag_result:
        return "1 MOs created", "42FF00"
    if tag_result == "Executed" or re.search(
        r"([0-9]{1,90}\s+.*?\-\-\>.*?set\.)$",
        tag_result,
        re.IGNORECASE,
    ):
        return "1 MOs set", "42FF00"
    if tag_result == "Mo deleted":
        return "Mo deleted", "42FF00"
    if "MO already exists" in tag_result or tag_result == "MoNameAlreadyTaken":
        return "1 MOs already exists", "FF7575"
    if "Parent MO not found" in tag_result:
        return "Parent MO not found", "FF7575"
    if tag_result == "MoNotFound":
        return "MO not found", "FF7575"
    if (
        re.search(r"operation\-failed", tag_result, re.IGNORECASE)
        or re.search(r"unknown\-attribute", tag_result, re.IGNORECASE)
        or "failure" in tag_result.lower()
    ):
        return "ERROR operation-failed", "FF7575"
    return tag_result, "FF7575"


def _category_checking(tag_result: str) -> tuple[str, str]:
    match = _TOTAL_ATTEMPTED_PATTERN.search(tag_result or "")
    attempted = int(match.group(1)) if match else None
    if tag_result in {"1 MOs created", "1 MOs set", "Mo deleted"}:
        return tag_result, "42FF00"
    if match and attempted and attempted > 0:
        return tag_result, "42FF00"
    return tag_result, "FF7575"


def _format_report_value(value: str) -> str:
    match = _TOTAL_ATTEMPTED_PATTERN.match(str(value))
    if not match:
        return str(value)
    attempted = int(match.group(1))
    result = match.group(2)
    if attempted == 0:
        return str(value)
    return f"TOTAL X MOs {result}"


def _parse_single_log(log_path: str, selected_file: str) -> list[tuple]:
    filename = os.path.basename(log_path)
    node_log = (
        re.search(r"^(.*?).log$", filename, re.IGNORECASE).group(1)
        if re.search(r"^(.*?).log$", filename, re.IGNORECASE)
        else filename
    )
    local_log_data: list[tuple] = []

    with open(log_path, "rb") as f:
        raw_bytes = f.read(4)
    encoding = (
        "utf-16"
        if raw_bytes.startswith(b"\xff\xfe") or raw_bytes.startswith(b"\xfe\xff")
        else "utf-8"
    )

    with open(log_path, "r", encoding=encoding, errors="ignore") as file:
        type_script = "NULL"
        type_script_rnc = "NULL"
        line_execute = "NULL"
        tag_result = "NULL"
        truni_script = "NULL"
        tag_result_full = "NULL"
        number_tag = 0
        cmd_get = "NULL"
        att_list: list[str] = []

        for line in file:
            cmd_match = _PROMPT_PATTERN.search(line)
            cmd_get_1 = cmd_match.group(1).strip() if cmd_match else line_execute

            run_match = _RUN_SCRIPT_PATTERN.search(line)
            if run_match:
                type_script = run_match.group(1)

            truni_match = _TRUNI_PATTERN.search(cmd_get_1)
            if truni_match:
                truni_script = truni_match.group(1)

            if truni_script != "NULL":
                for key in ("trun", "truni"):
                    tsr_match = re.search(fr".*?{key}\s+(.*?).mo$", line, re.IGNORECASE)
                    if tsr_match:
                        type_script_rnc = tsr_match.group(1)

                for pattern in (_CREATE_PATTERN, _DELETE_PATTERN, _SET_PATTERN):
                    cmd = pattern.search(line)
                    if cmd:
                        line_execute = cmd.group(1)

                tag = _TAG_PATTERN.search(line)
                if tag:
                    tag_result = tag.group(1)
                    tag_result_full = line.rstrip()

                tag_alt = _ALT_TAG_PATTERN.search(line)
                if tag_alt:
                    tag_result = tag_alt.group(1)
                    tag_result_full = line.rstrip()

                if len(line.rstrip()) == 0 and line_execute != "NULL":
                    if len(tag_result_full.rstrip()) == 0:
                        tag_result = "Executed"
                    local_log_data.append(
                        (
                            selected_file,
                            node_log,
                            type_script_rnc,
                            line_execute.strip(),
                            tag_result,
                            tag_result_full,
                            [],
                        )
                    )
                    line_execute = "NULL"
                    tag_result = "NULL"
                    tag_result_full = ""

            if truni_script == "NULL":
                if _PROMPT_PATTERN.match(line):
                    if number_tag == 1:
                        if re.search(r"^(del|rdel|crn|set|lset)", cmd_get.strip(), re.IGNORECASE):
                            tag_report, _ = _category_checking1(tag_result)
                            local_log_data.append(
                                (
                                    selected_file,
                                    node_log,
                                    type_script,
                                    line_execute.strip(),
                                    tag_report.strip(),
                                    tag_result.strip(),
                                    att_list,
                                )
                            )
                            tag_result = "NULL"
                            number_tag = 0
                        else:
                            number_tag = 0
                    number_tag += 1
                    cmd_get = cmd_get_1

                for pattern in _TAG_ERROR_PATTERNS:
                    match = pattern.search(line)
                    if match:
                        tag_result = match.group(1)

                if line.strip().lower() == "end" and tag_result == "NULL":
                    tag_result = "Executed"

                if _PROMPT_PATTERN.match(line):
                    att_list = ["", line.strip()]
                    if line_execute == "NULL":
                        line_execute = cmd_get_1
                else:
                    att_list.append(line.strip())

            if "Checking ip contact...Not OK" in line:
                local_log_data.append((selected_file, node_log, "", "", "UNREMOTE", line.strip(), ""))
            if re.search(r"(?i)tbac\s*control\s*\-\s*unauthori[sz]ed\s*network\s*element", line):
                local_log_data.append((selected_file, node_log, "", "", "UNREMOTE", line.strip(), ""))

        if truni_script == "NULL" and number_tag == 1:
            if re.search(r"^(del|rdel|crn|set|lset)", cmd_get.strip(), re.IGNORECASE):
                tag_report, _ = _category_checking1(tag_result)
                local_log_data.append(
                    (
                        selected_file,
                        node_log,
                        type_script,
                        line_execute.strip(),
                        tag_report.strip(),
                        tag_result.strip(),
                        att_list,
                    )
                )

    return local_log_data


def _write_log_sheet(workbook: openpyxl.Workbook, log_data: list[tuple]) -> None:
    ws = workbook.active
    ws.title = "LOGS_MERGED"

    headers = ["CR", "Site", "Command", "Parameter", "Report", "TAG", "Script", "Full command log"]
    column_widths = [20, 20, 45, 25, 25, 20, 40, 70]

    for col_num, (header, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for idx, (file_log, site, type_script, line_execute, tag_result, line_full_log, att_list) in enumerate(
        log_data,
        start=2,
    ):
        clean_log = _ILLEGAL_CHAR_PATTERN.sub("", str(line_full_log))
        tag_remark, tag_color = _category_checking(str(tag_result))
        if line_execute == "NULL":
            line_execute = next(
                (
                    m.group(1)
                    for line in att_list
                    if isinstance(line, str) and (m := _PROMPT_PATTERN.match(line))
                ),
                "",
            )

        parameter_match = re.search(r"^SET\s+.*?\s+(.*?)\s+\=", str(line_execute), re.IGNORECASE)
        parameter = parameter_match.group(1) if parameter_match else ""
        if not parameter:
            parameter_match = re.search(r">\s+set\s+.*?\s+(.*?)\s+", str(line_execute), re.IGNORECASE)
            parameter = parameter_match.group(1) if parameter_match else ""

        ws.cell(row=idx, column=1, value=file_log).font = Font(size=9, bold=True)
        ws.cell(row=idx, column=2, value=site).font = Font(size=9, bold=True)
        ws.cell(row=idx, column=3, value=str(line_execute)).font = Font(size=9)
        ws.cell(row=idx, column=4, value=parameter).font = Font(size=9)
        ws.cell(row=idx, column=5, value=tag_remark).font = Font(size=9)
        ws.cell(row=idx, column=6, value=clean_log).font = Font(size=9)
        ws.cell(row=idx, column=7, value=type_script).font = Font(size=9)
        ws.cell(row=idx, column=8, value="\n".join(att_list) if isinstance(att_list, list) else str(att_list)).font = Font(size=9)
        ws.cell(row=idx, column=5).fill = PatternFill(start_color=tag_color, end_color=tag_color, fill_type="solid")


def _build_summary_rows(log_data: Iterable[tuple]) -> list[tuple[str, str, int, int, str]]:
    normalized_rows: list[tuple[str, str]] = []
    for row in log_data:
        cr = str(row[0])
        report = _format_report_value(str(row[4]))
        normalized_rows.append((cr, report))

    counts = Counter(normalized_rows)
    totals = Counter(cr for cr, _ in normalized_rows)

    summary_rows: list[tuple[str, str, int, int, str]] = []
    for cr, report in sorted(counts.keys()):
        count_result = counts[(cr, report)]
        total = totals[cr]
        percentage = "{:.1f}%".format((count_result / total) * 100) if total else "0.0%"
        summary_rows.append((cr, report, count_result, total, percentage))
    return summary_rows


def _write_summary_sheet(workbook: openpyxl.Workbook, summary_rows: list[tuple[str, str, int, int, str]]) -> None:
    pivot_sheet = workbook.create_sheet(title="Pivot Table")
    headers = ["", "CR", "Report", "Count Result", "Total", "Percentage"]
    widths = {2: 60, 3: 30, 4: 12, 5: 6, 6: 12}

    for idx, header in enumerate(headers, start=1):
        cell = pivot_sheet.cell(row=1, column=idx, value=header)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    for col_idx, width in widths.items():
        pivot_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, (cr, report, count_result, total, percentage) in enumerate(summary_rows, start=2):
        values = [row_idx - 2, cr, report, count_result, total, percentage]
        for col_idx, value in enumerate(values, start=1):
            cell = pivot_sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 2:
                cell.font = Font(bold=True)
            if col_idx == 3:
                _, tag_color = _category_checking(str(value))
                cell.fill = PatternFill(start_color=tag_color, end_color=tag_color, fill_type="solid")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )


def build_relation_log_excel(
    log_files: Iterable[str],
    output_dir: str,
    node_name: str,
    log_cb: LogCallback = None,
) -> str:
    """Parse downloaded relation moshell logs into an Excel workbook."""
    existing_logs = [path for path in log_files if path.lower().endswith(".log") and os.path.exists(path)]
    if not existing_logs:
        raise ValueError("No downloaded .log files were available to parse.")

    os.makedirs(output_dir, exist_ok=True)
    selected_file = node_name or "RELATION"
    log_data: list[tuple] = []

    for idx, log_path in enumerate(sorted(existing_logs), start=1):
        _log(log_cb, f"Parsing relation log {idx}/{len(existing_logs)}: {os.path.basename(log_path)}")
        try:
            log_data.extend(_parse_single_log(log_path, selected_file))
        except Exception as exc:
            _log(log_cb, f"Skipping {os.path.basename(log_path)}: {exc}")

    workbook = openpyxl.Workbook()
    _write_log_sheet(workbook, log_data)
    _write_summary_sheet(workbook, _build_summary_rows(log_data))

    safe_node = re.sub(r"[^A-Za-z0-9_\-]+", "_", selected_file).strip("_") or "RELATION"
    output_path = os.path.join(output_dir, f"RELATION_{safe_node}_PARSED.xlsx")
    workbook.save(output_path)
    return output_path
