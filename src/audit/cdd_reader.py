"""
cdd_reader.py — turn a wide CDD Excel sheet into a flat list of AuditItem,
driven entirely by ``audit_map.json`` so sheets/columns/parameters can be
added by editing JSON (no code change).

Each map profile describes one (sheet → MO) mapping:

    {
      "name": "LTE Cell Parameters",
      "tech": "lte_nr",           # which CDD file (lte_nr | gsm)
      "category": "cell",         # node | cell | relation
      "sheet": "CDD",
      "header_row": 3,
      "node_key_column": "eNodeBName",   # filter rows to the audited node
      "mo_fdn": "ENodeBFunction=1,EUtranCellFDD={CellName}",
      "columns": [
        {"cdd": "TAC", "attr": "tac"},
        {"cdd": "DataRadioBearer.dlMaxRetxThreshold",
         "attr": "dlMaxRetxThreshold",
         "mo": "ENodeBFunction=1,DataRadioBearer=1"}   # per-column MO override
      ]
    }

``mo_fdn`` / ``mo`` are templates: ``{ColumnHeader}`` is replaced with that
row's value. ``node_key_column`` matches when its value starts with (or
equals) the audited node name — so only that node's rows are read.
"""
from __future__ import annotations

import json
import os
import re
from typing import Dict, List

from openpyxl import load_workbook

from .audit_core import AuditItem


def load_map(explicit_path: str = "") -> dict:
    """Load audit_map.json — explicit path, else next to the exe, else the
    bundled copy beside this package."""
    candidates = []
    if explicit_path:
        candidates.append(explicit_path)
    try:
        from app_path import get_app_dir
        candidates.append(os.path.join(get_app_dir(), "audit_map.json"))
    except Exception:
        pass
    candidates.append(os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "audit_map.json"))
    for p in candidates:
        if p and os.path.isfile(p):
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
    raise FileNotFoundError("audit_map.json not found in: " + ", ".join(candidates))


def resolve_map_path() -> str:
    """Path of the audit_map.json the app actually uses (for the in-app editor):
    the exe-dir copy if present, else the bundled/source copy. Prefers a
    writable exe-dir location when running frozen."""
    candidates = []
    try:
        from app_path import get_app_dir
        candidates.append(os.path.join(get_app_dir(), "audit_map.json"))
    except Exception:
        pass
    pkg_copy = os.path.normpath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "audit_map.json"))
    candidates.append(pkg_copy)
    for p in candidates:
        if p and os.path.isfile(p):
            return p
    return candidates[0] if candidates else pkg_copy


_PLACEHOLDER = re.compile(r"\{([^}]+)\}")


def _fill_template(template: str, row: Dict[str, str]) -> str:
    """Replace {Header} with the row's value for that header."""
    def repl(m):
        return str(row.get(m.group(1), "")).strip()
    return _PLACEHOLDER.sub(repl, template)


def _norm(s) -> str:
    return str(s).strip().lower() if s is not None else ""


def _hnorm(s) -> str:
    """Whitespace-normalized header: nbsp → space, collapse runs, strip."""
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip()


def read_audit_items(cdd_paths: Dict[str, str], node_name: str,
                     audit_map: dict, log=lambda m: None) -> List[AuditItem]:
    """Read all profiles whose tech has a CDD file provided → AuditItems for
    the audited ``node_name``.

    Each workbook is opened ONCE and every sheet is materialized ONCE into a
    cache shared across profiles — many GSM profiles hit the same big sheet, so
    re-reading per profile made audits take minutes."""
    node_l = _norm(node_name)
    items: List[AuditItem] = []
    wbcache: Dict[str, object] = {}       # path -> workbook
    sheet_cache: Dict[tuple, tuple] = {}  # (path, sheet, header_row) -> (col_idx, rows)
    try:
        for prof in audit_map.get("profiles", []):
            tech = prof.get("tech", "")
            path = cdd_paths.get(tech)
            if not path or not os.path.isfile(path):
                continue
            try:
                items += _read_profile(path, node_l, prof, log,
                                       sheet_cache, wbcache)
            except Exception as exc:
                log(f"[audit] profile '{prof.get('name')}' failed: {exc}")
    finally:
        for wb in wbcache.values():
            try:
                wb.close()
            except Exception:
                pass
    return items


def _get_sheet(path, sheet, header_row, sheet_cache, wbcache):
    """Return (col_idx, data_rows) for a sheet, reading it at most once."""
    key = (path, sheet, header_row)
    if key in sheet_cache:
        return sheet_cache[key]
    wb = wbcache.get(path)
    if wb is None:
        wb = load_workbook(path, read_only=True, data_only=True)
        wbcache[path] = wb
    if sheet not in wb.sheetnames:
        sheet_cache[key] = (None, None)
        return sheet_cache[key]
    rows = list(wb[sheet].iter_rows(min_row=header_row, values_only=True))
    if not rows:
        sheet_cache[key] = ({}, [])
        return sheet_cache[key]
    col_idx: Dict[str, int] = {}
    for i, h in enumerate(rows[0]):
        name = str(h).strip() if h is not None else ""
        if name and name not in col_idx:
            col_idx[name] = i
        # Also index a whitespace-normalized alias (collapse double spaces,
        # nbsp → space) so a CDD header like "No of  UL TRX" still matches the
        # map's "No of UL TRX".
        norm = _hnorm(name)
        if norm and norm not in col_idx:
            col_idx[norm] = i
    sheet_cache[key] = (col_idx, rows[1:])
    return sheet_cache[key]


def _build_cell_map(expand, sheet_cache, wbcache, log) -> Dict[tuple, List[str]]:
    """Map (eNodeBName, Freq Band) → [CellName,...] from a pairing sheet, so a
    relation row keyed by band (e.g. EUTRANCell='L18') can be expanded to every
    real cell of that band on the node."""
    sh = expand["sheet"]
    hr = int(expand.get("header_row", 1))
    idx, data = _get_sheet(expand.get("_path"), sh, hr, sheet_cache, wbcache)
    if idx is None:
        log(f"[audit] cell_expand sheet '{sh}' not found — relation not expanded")
        return {}
    try:
        ei, bi, ci = (idx[expand["enb_column"]], idx[expand["band_column"]],
                      idx[expand["cell_column"]])
    except KeyError as k:
        log(f"[audit] cell_expand column {k} missing in '{sh}' — not expanded")
        return {}
    cmap: Dict[tuple, List[str]] = {}
    for row in data:
        if max(ei, bi, ci) >= len(row):
            continue
        enb, band, cell = row[ei], row[bi], row[ci]
        if not enb or not band or not cell:
            continue
        cmap.setdefault((_norm(enb), _norm(band)), []).append(str(cell).strip())
    return cmap


def _read_profile(path: str, node_l: str, prof: dict, log,
                  sheet_cache, wbcache) -> List[AuditItem]:
    sheet = prof["sheet"]
    header_row = int(prof.get("header_row", 1))
    node_key = prof["node_key_column"]
    # How the CDD node-key column relates to the audited node name:
    #  * default: the CDD value IS (or is a prefix of) the full node name
    #    (eNodeBName/gNBName style).
    #  * "prefix_of_node": the CDD value is the site/PLA ID (e.g. MIN2782) — a
    #    prefix of the node name (MIN2782_ROCA…B03), as in the LLD workbook.
    node_key_match = prof.get("node_key_match", "")
    default_mo = prof.get("mo_fdn", "")
    columns = prof["columns"]
    category = prof.get("category", "")
    tech = prof.get("tech", "")
    # BSC/cmedit-sourced params (GeranCell) can't be set via moshell, so tag
    # them here to keep them out of the generated .mos script.
    from_cmedit = prof.get("source") == "cmedit"
    expand = prof.get("cell_expand")   # optional band→cell expansion

    col_idx, data = _get_sheet(path, sheet, header_row, sheet_cache, wbcache)
    if col_idx is None:
        log(f"[audit] sheet '{sheet}' not in {os.path.basename(path)} — skipped")
        return []
    if node_key not in col_idx:
        log(f"[audit] node_key '{node_key}' not in sheet '{sheet}' — skipped")
        return []

    if expand:
        expand = dict(expand)
        expand["_path"] = path        # cell_expand reads a sheet in the same file
    cell_map = _build_cell_map(expand, sheet_cache, wbcache, log) if expand else {}
    band_col = expand.get("row_band_column") if expand else None

    # Optional: pick the cell MO class from the Freq Band (TDD vs FDD). The
    # chosen class is exposed to templates as {_cellmo}.
    cmm = prof.get("cell_mo_map")
    cmm_col = cmm.get("band_column") if cmm else None
    cmm_map = {k.lower(): v for k, v in (cmm.get("map", {}) if cmm else {}).items()}
    cmm_default = cmm.get("default", "") if cmm else ""

    items: List[AuditItem] = []
    nk = col_idx[node_key]
    for row in data:
        if nk >= len(row):
            continue
        keyval = _norm(row[nk])
        if not keyval:
            continue
        # Row belongs to the audited node?
        if node_key_match == "prefix_of_node":
            # CDD key is the site/PLA ID; match when the node name starts with
            # it at a segment boundary (MIN2782 → MIN2782_…, not MIN27820…).
            if not (keyval == node_l or node_l.startswith(keyval + "_")):
                continue
        elif not (keyval == node_l
                  or keyval.startswith(node_l + "_")
                  or keyval.startswith(node_l + "-")):
            # Boundary-aware: the audited Site ID / node must be followed by a
            # separator, so "MIN278" matches "MIN278_BULUAN…" but NOT
            # "MIN2780_…" / "MIN2781_…" (which is a different site).
            continue
        rowdict = {name: (row[i] if i < len(row) else "")
                   for name, i in col_idx.items()}
        if cmm:
            bval = _norm(row[col_idx[cmm_col]]) if cmm_col in col_idx else ""
            rowdict["_cellmo"] = cmm_map.get(bval, cmm_default)

        # Resolve which cell(s) this row applies to. With cell_expand the row's
        # cell field holds a BAND (e.g. 'L18') → expand to every real cell of
        # that band on this node; otherwise a single pass with no injected cell.
        if expand and band_col and band_col in col_idx:
            band = _norm(row[col_idx[band_col]])
            cells = cell_map.get((keyval, band), [])
            if not cells:
                log(f"[audit] no {band_col}='{rowdict.get(band_col)}' cells for "
                    f"'{row[nk]}' — relation row skipped")
                continue
        else:
            cells = [None]

        for cell in cells:
            rd = dict(rowdict)
            if cell is not None:
                rd["_cell"] = cell
            for col in columns:
                cdd_col = col["cdd"]
                ci = col_idx.get(cdd_col)
                if ci is None:
                    ci = col_idx.get(_hnorm(cdd_col))
                if ci is None:
                    continue
                raw = row[ci] if ci < len(row) else None
                expected = "" if raw is None else str(raw).strip()
                if expected == "":
                    continue
                mo_tmpl = col.get("mo", default_mo)
                mo_local = _fill_template(mo_tmpl, rd)
                node_val = str(row[nk]).strip()
                key_val = cell if cell is not None else node_val
                # A ``split`` column derives several REAL params from one CDD
                # value (e.g. MIMO "32T32R" → noOfTxAntennas=32, noOfRxAntennas=32
                # on the SectorCarrier via via_ref), so each becomes its own MO
                # attribute — auditable and directly settable in the script.
                splits = col.get("split")
                if splits:
                    for sp in splits:
                        m = re.search(sp["regex"], expected)
                        if not m:
                            continue
                        items.append(AuditItem(
                            category=category, tech=tech, mo_local=mo_local,
                            parameter=sp["attr"], expected=m.group(1),
                            key=key_val, source=f"{sheet}!{cdd_col}",
                            norm=sp.get("norm", col.get("norm", "")),
                            via_ref=col.get("via_ref", ""),
                            attr_alt=sp.get("attr_alt", ""),
                            from_cmedit=from_cmedit,
                            node=node_val))
                    continue
                items.append(AuditItem(
                    category=category, tech=tech, mo_local=mo_local,
                    parameter=col["attr"], expected=expected,
                    key=key_val,
                    source=f"{sheet}!{cdd_col}",
                    norm=col.get("norm", ""),
                    via_ref=col.get("via_ref", ""),
                    attr_format=col.get("attr_format", ""),
                    attr_fallback=col.get("attr_fallback", ""),
                    attr_alt=col.get("attr_alt", ""),
                    from_cmedit=from_cmedit,
                    node=node_val))
    return items
