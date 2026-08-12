"""
lld_audit.py — audit a baseband node against the LLD (Tx/CPRI) workbook.

Key topology fact (verified on real dumps): a site's LLD row is keyed by
**PLA ID** (e.g. MIN3250) and its ``BBID`` column (BB1/BB2/BB3) selects the
physical **node** — the node name ends with the matching ``B0<k>`` suffix
(BB1→…B01, BB2→…B02, BB3→…B03). On *every* node the baseband unit is the same
local FRU ``BB-1``; there is no BB-2/BB-3 FRU on one node. So to audit node
``…B0<k>`` we look only at BBID ``BB<k>`` / column ``BB-<k>``.

Two audits, both producing ``AuditResult`` rows straight into the shared Excel
report:

* **Baseband type** (sheet ``Tx connectivity``, column ``BB-<k>``) vs the node
  FRU ``BB-1`` ``productData.productName`` — e.g. CDD ``RP6655`` == node
  ``RAN Processor 6655`` (compared on the product number via ``bbtype``).
* **RiLink / CPRI port allocation** (sheet ``CPRI connectivity``, rows with
  ``BBID==BB<k>``) vs the node ``RiLink`` MOs, matched by the baseband
  ``RiPort`` (riPortRef1, always on ``BB-1``):
    1. RiLink present on the planned baseband port.
    2. riPortRef2 points at the planned ``Radio DATA Port``.
    3. radio type ≈ the radio FRU on riPortRef2 (heuristic — the node stores no
       radio ``productName``, only the FRU name like ``AAS_B41_RRU1``).
"""
from __future__ import annotations

import collections
import difflib
import re
from typing import Dict, List

from openpyxl import load_workbook

from .audit_core import LldResult, normalize_bbtype


def _norm(s) -> str:
    return str(s).strip().lower() if s is not None else ""


def _node_bb_index(node_name: str) -> int:
    """Baseband index from the node-name suffix: '…B01'→1, '…B02'→2. 0 if none."""
    m = re.search(r"B0*(\d+)\s*$", node_name or "")
    return int(m.group(1)) if m else 0


def _bbid_index(bbid: str) -> int:
    m = re.search(r"(\d+)", str(bbid or ""))
    return int(m.group(1)) if m else 0


def _ref_parts(ref: str):
    """('…FieldReplaceableUnit=BB-1,RiPort=H') → ('BB-1', 'H')."""
    fru = re.search(r"FieldReplaceableUnit=([^,]+)", ref or "")
    port = re.search(r"RiPort=([^,]+)", ref or "")
    return (fru.group(1).strip() if fru else "",
            port.group(1).strip() if port else "")


def _product_name(a: Dict[str, str]) -> str:
    """productName for an FRU — from a flattened ``productData.productName`` or a
    standalone ``productName`` (cmdump), else parsed out of the raw ``productData``
    struct string the modump emits, e.g.
    ``{productionDate=..., productName=RAN Processor 6655, ...}``."""
    v = a.get("productData.productName") or a.get("productName")
    if v:
        return v
    pd = a.get("productData")
    if pd:
        m = re.search(r"productName\s*=\s*([^,}]+)", str(pd))
        if m:
            return m.group(1).strip()
    return ""


def _fru_product_name(records: Dict[str, Dict[str, str]], node: str) -> str:
    """The node's baseband FRU (BB-1) productName, if present."""
    node_l = _norm(node)
    for ldn, a in records.items():
        if ldn.split(",")[-1] != "FieldReplaceableUnit=BB-1":
            continue
        me = re.search(r"ManagedElement=([^,]+)", ldn)
        if me and _norm(me.group(1)) != node_l:
            continue
        return _product_name(a)
    return ""


def _index_node_rilinks(records: Dict[str, Dict[str, str]], node: str):
    """This node's RiLinks keyed by baseband RiPort (upper). The baseband FRU is
    always BB-1, so the port alone identifies the link. Value = (riLinkId,
    radio_fru, radio_port)."""
    node_l = _norm(node)
    idx: Dict[str, tuple] = {}
    for ldn, a in records.items():
        if ldn.split(",")[-1].split("=", 1)[0] != "RiLink":
            continue
        me = re.search(r"ManagedElement=([^,]+)", ldn)
        if me and _norm(me.group(1)) != node_l:
            continue
        bb_fru, bb_port = _ref_parts(a.get("riPortRef1", ""))
        radio_fru, radio_port = _ref_parts(a.get("riPortRef2", ""))
        if not bb_port:
            continue
        idx[bb_port.upper()] = (a.get("riLinkId", "") or ldn.split("=")[-1],
                                radio_fru, radio_port)
    return idx


_BAND_RE = re.compile(r"B\d+[A-Z0-9]*", re.IGNORECASE)


def _canon_band(s: str) -> str:
    """Canonical band key from a radio type / FRU name — the ``B<digits>``
    tokens joined, so ``Radio 4499 B0AB28`` and node FRU ``B0B28_RRU1_L`` both
    reduce to ``B0B28`` (and ``AIR 3229 B41B78AF`` / ``AAS_B41B78_RRU1`` →
    ``B41B78``). Distinguishes B41 / B78 / B41B78 cleanly."""
    return "".join(re.findall(r"B\d+", str(s or "").upper()))


def _rru_index(radio_fru: str) -> str:
    """The RRU/sector number in a radio FRU name — the digits right after
    ``RRU``: ``AAS_B41_RRU2`` → ``2``, ``B0B28_RRU123_1`` → ``123``. Empty when
    the name has no ``RRU`` token."""
    m = re.search(r"RRU(\d+)", radio_fru or "", re.IGNORECASE)
    return m.group(1) if m else ""


def _radio_type_matches(cdd_type: str, radio_fru: str) -> bool:
    """Best-effort radio-type check: compare CDD radio type vs the radio FRU
    name on riPortRef2 via shared band tokens (B41 / B1B3 / B0AB28≈B0B28)."""
    a = cdd_type.upper().replace(" ", "")
    b = radio_fru.upper().replace(" ", "")
    if not a or not b:
        return False
    cdd_bands = _BAND_RE.findall(a)
    fru_bands = _BAND_RE.findall(b)
    for cb in cdd_bands:
        for fb in fru_bands:
            if cb == fb or difflib.SequenceMatcher(None, cb, fb).ratio() >= 0.8:
                return True
    return any(cb in b for cb in cdd_bands) or any(fb in a for fb in fru_bands)


def _sheet_rows(wb, sheet: str, header_row: int):
    if sheet not in wb.sheetnames:
        return None, None
    rows = list(wb[sheet].iter_rows(min_row=header_row, values_only=True))
    if not rows:
        return {}, []
    col = {}
    for i, h in enumerate(rows[0]):
        name = str(h).strip() if h is not None else ""
        if name and name not in col:
            col[name] = i
    return col, rows[1:]


def _row_matches_node(row, col, node_l) -> bool:
    i = col.get("PLA ID", -1)
    pla = _norm(row[i]) if 0 <= i < len(row) else ""
    return bool(pla) and (pla == node_l or node_l.startswith(pla + "_"))


def audit_lld(lld_path: str, node_name: str,
              records: Dict[str, Dict[str, str]],
              tx_sheet: str = "Tx connectivity", tx_header_row: int = 1,
              cpri_sheet: str = "CPRI connectivity", cpri_header_row: int = 2,
              log=lambda m: None) -> List[LldResult]:
    """Audit ``node_name`` against the LLD workbook. Returns LldResult rows for
    the baseband-type and RiLink/CPRI checks (empty if the node isn't in the
    LLD or has no ``B0<k>`` suffix)."""
    k = _node_bb_index(node_name)
    if not k:
        log(f"[audit/lld] {node_name}: no B0<k> suffix — LLD audit skipped")
        return []
    node_l = _norm(node_name)
    bbid = f"BB{k}"
    out: List[LldResult] = []

    wb = load_workbook(lld_path, read_only=True, data_only=True)
    try:
        # ── Baseband type ──────────────────────────────────────────
        col, data = _sheet_rows(wb, tx_sheet, tx_header_row)
        bb_col = f"BB-{k}"
        if col is None:
            log(f"[audit/lld] sheet '{tx_sheet}' missing — baseband skipped")
        elif bb_col not in col:
            log(f"[audit/lld] column '{bb_col}' missing in '{tx_sheet}'")
        else:
            expected = ""
            for row in data:
                if _row_matches_node(row, col, node_l):
                    v = row[col[bb_col]] if col[bb_col] < len(row) else None
                    expected = "" if v is None else str(v).strip()
                    break
            if expected:
                actual = _fru_product_name(records, node_name)
                hw_ok = bool(actual) and (
                    normalize_bbtype(actual) == normalize_bbtype(expected))
                status = ("NotFound" if not actual
                          else ("Match" if hw_ok else "Mismatch"))
                out.append(LldResult(
                    node=node_name, bbid=bbid, ref_cell="",
                    hw_type_lld=expected, hw_type_node=actual,
                    status=status, source=f"{tx_sheet}!{bb_col}",
                    hw_ok=(status != "Mismatch")))

        # ── RiLink / CPRI port allocation ──────────────────────────
        col, data = _sheet_rows(wb, cpri_sheet, cpri_header_row)
        if col is None:
            log(f"[audit/lld] sheet '{cpri_sheet}' missing — RiLink skipped")
        else:
            need = ["BBID", "BB RI Port", "Radio DATA Port"]
            missing = [c for c in need if c not in col]
            if missing:
                log(f"[audit/lld] columns missing in '{cpri_sheet}': {missing}")
            else:
                out += _audit_rilink_rows(col, data, node_name, node_l, k, bbid,
                                          records, cpri_sheet, log)
    finally:
        wb.close()
    return out


def _audit_rilink_rows(col, data, node_name, node_l, k, bbid, records, sheet, log):
    def cell(row, name):
        i = col.get(name, -1)
        return row[i] if 0 <= i < len(row) else None

    planned = []
    for row in data:
        if not _row_matches_node(row, col, node_l):
            continue
        if _bbid_index(cell(row, "BBID")) != k:      # only this node's baseband
            continue
        planned.append({
            "port": str(cell(row, "BB RI Port") or "").strip(),
            "data_port": str(cell(row, "Radio DATA Port") or "").strip(),
            "radio_type": str(cell(row, "Radio Type") or "").strip(),
            "sector": str(cell(row, "Sector") or "").strip(),
            "radio_no": str(cell(row, "Radio Number") or "").strip(),
        })

    idx = _index_node_rilinks(records, node_name)
    log(f"[audit/lld] {node_name}: BBID={bbid} → {len(planned)} planned link(s), "
        f"{len(idx)} RiLink MO(s) on node")

    # Pool of node links, keyed by port; consumed as we pair them.
    pool = {port: (fru, dport) for port, (_rid, fru, dport) in idx.items()}

    def _pair(p):
        """Pick this planned row's node link: same BB port first, else the first
        unpaired node link whose radio band matches (AAS wired on a different
        port than planned). Returns (node_port, fru, dport) or None."""
        up = p["port"].upper()
        if up in pool:
            fru, dport = pool.pop(up)
            return up, fru, dport
        for np in sorted(pool):
            fru, dport = pool[np]
            if p["radio_type"] and _radio_type_matches(p["radio_type"], fru):
                del pool[np]
                return np, fru, dport
        return None

    out: List[LldResult] = []
    for p in planned:
        ref_cell = (f"S{p['sector']}/R{p['radio_no']}"
                    if p["sector"] or p["radio_no"] else "")
        hit = _pair(p)
        if hit is None:                       # planned but no matching node link
            out.append(LldResult(
                node=node_name, bbid=bbid, ref_cell=ref_cell,
                bb_port_lld=p["port"], bb_port_node="",
                hw_type_lld=p["radio_type"], hw_type_node="",
                data_port_lld=p["data_port"], data_port_node="",
                status="NotFound", source=f"{sheet}!BB RI Port"))
            continue
        node_port, radio_fru, radio_port = hit
        # All three pairs are compared; any difference → Mismatch, and the
        # differing pair is highlighted. A BB port that was matched by radio
        # (e.g. AAS planned P, wired H) counts as a BB-port mismatch.
        bb_ok = _norm(p["port"]) == _norm(node_port)
        # HW check = radio band/type matches AND the radio FRU's RRU index lines
        # up with the planned Sector (double-check): a shared radio named
        # ``B0B28_RRU123_1`` on Sector 1 fails because ``123`` ≠ ``1`` — the RRU
        # should be per-sector (RRU1/RRU2/RRU3).
        band_ok = (not p["radio_type"]
                   or _radio_type_matches(p["radio_type"], radio_fru))
        rru = _rru_index(radio_fru)
        sector_ok = (not p["sector"]) or (not rru) or (rru == p["sector"])
        hw_ok = band_ok and sector_ok
        data_ok = (not p["data_port"]
                   or _norm(p["data_port"]) == _norm(radio_port))
        status = "Match" if (bb_ok and hw_ok and data_ok) else "Mismatch"
        out.append(LldResult(
            node=node_name, bbid=bbid, ref_cell=ref_cell,
            bb_port_lld=p["port"], bb_port_node=node_port,
            hw_type_lld=p["radio_type"], hw_type_node=radio_fru,
            data_port_lld=p["data_port"], data_port_node=radio_port,
            status=status, source=f"{sheet}!BB RI Port",
            bb_ok=bb_ok, hw_ok=hw_ok, data_ok=data_ok))

    # Node links left unpaired = present on the baseband but not in the LLD.
    for np in sorted(pool):
        fru, dport = pool[np]
        out.append(LldResult(
            node=node_name, bbid=bbid, ref_cell="",
            bb_port_lld="", bb_port_node=np,
            hw_type_lld="", hw_type_node=fru,
            data_port_lld="", data_port_node=dport,
            status="Unplanned", source="node RiLink"))

    # RRU count per radio type: the LLD plans N rows of a band; the node must
    # have N RRUs of that band. Counted per (BBID, band) TOTAL — robust to how
    # the AAS bands are split/combined (B41/B78 vs B41B78) and to shared-sector
    # rows (Sector "123"). One summary row per band.
    lld_cnt: Dict[str, int] = collections.Counter()
    lld_type: Dict[str, str] = {}
    for p in planned:
        cb = _canon_band(p["radio_type"])
        if not cb:
            continue
        lld_cnt[cb] += 1
        lld_type.setdefault(cb, p["radio_type"])
    node_cnt: Dict[str, int] = collections.Counter()
    node_frus: Dict[str, list] = collections.defaultdict(list)
    for _p, (_rid, fru, _dp) in idx.items():
        cb = _canon_band(fru)
        if cb:
            node_cnt[cb] += 1
            node_frus[cb].append(fru)
    for cb in sorted(set(lld_cnt) | set(node_cnt)):
        exp, act = lld_cnt.get(cb, 0), node_cnt.get(cb, 0)
        status = ("Match" if exp == act
                  else "Unplanned" if exp == 0
                  else "NotFound" if act == 0
                  else "Mismatch")
        out.append(LldResult(
            node=node_name, bbid=bbid, ref_cell=f"count {cb}",
            hw_type_lld=(f"{exp} x {lld_type.get(cb, cb)}" if exp else ""),
            hw_type_node=(f"{act} RRU ({', '.join(sorted(node_frus[cb]))})"
                          if act else ""),
            status=status, source=f"{sheet}!RRU count",
            hw_ok=(status == "Match")))
    return out
