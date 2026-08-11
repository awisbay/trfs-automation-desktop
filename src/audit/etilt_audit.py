"""
etilt_audit.py — electrical antenna tilt (ElecTilt) audit.

There is no direct cell → RET reference in the dump, so cells are linked to
their ``RetSubUnit`` by the RET's ``userLabel``, which follows the cell-name
convention: ``<site><band-letters>-<sector>[_port]`` e.g. ``GFATIMLY-1_R1``
(bands L+Y, sector 1). A cell such as ``GFATIML-171`` (site GFATIM, band L,
sector = last digit 1) matches any RET whose userLabel starts with the site,
whose band-letter group contains the cell's band letter, and whose sector
matches.

  * primary source — ``RetSubUnit.electricalAntennaTilt`` (0.1°, so ÷10 → the
    CDD's degrees: node 20 == CDD 2);
  * fallback (NR / AIR beamforming, no physical RET) —
    ``CommonBeamforming.digitalTilt`` under the cell's NRSectorCarrier.

Multiple RET branches for one cell normally carry the same tilt; if they differ
the row is a Mismatch showing each value.
"""
from __future__ import annotations

import re
from typing import Dict, List, Optional

import openpyxl

from .audit_core import AuditResult


def _norm(v):
    return "" if v is None else str(v).strip()


def _parse_cell(cell: str):
    """``GFATIML-171`` → (site 'GFATIM', band 'L', sector '1'). Sector is the
    last digit of the whole name; band is the last letter before the first
    dash."""
    pre = cell.split("-")[0]
    if len(pre) < 2 or not pre[-1].isalpha():
        return None
    band, site = pre[-1].upper(), pre[:-1].upper()
    md = re.search(r"(\d)\D*$", cell)
    if not md:
        return None
    return site, band, md.group(1)


def _parse_ret(ul: str):
    """``GFATIMLY-1_R1`` → (alpha-prefix 'GFATIMLY', sector '1')."""
    m = re.match(r"^([A-Za-z]+)-(\d+)", ul or "")
    return (m.group(1).upper(), m.group(2)) if m else None


def _tilt_deg(raw) -> str:
    """RET/beamforming tilt (0.1°) → degrees string: 20 → '2', 25 → '2.5'."""
    try:
        f = int(raw) / 10.0
    except (ValueError, TypeError):
        return _norm(raw)
    return str(int(f)) if f == int(f) else str(f)


def _eq(a: str, b: str) -> bool:
    try:
        return abs(float(a) - float(b)) < 1e-6
    except (ValueError, TypeError):
        return _norm(a).lower() == _norm(b).lower()


def read_etilt_targets(path: str, sheet: str, header_row: int,
                       node_key_col: str, node: str,
                       log=lambda m: None) -> List[tuple]:
    """(cellName, node, elecTilt) rows for ``node`` from a CDD sheet."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        log(f"[audit/etilt] open {sheet}: {exc}")
        return []
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet]
    rows = ws.iter_rows(min_row=header_row, values_only=True)
    header = [re.sub(r"\s+", " ", str(c).replace("\xa0", " ")).strip().lower()
              if c is not None else "" for c in next(rows)]

    def col(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None
    ci_node = col(node_key_col.lower())
    ci_cell = col("cellname")
    ci_tilt = col("electilt", "etilt")
    if ci_node is None or ci_cell is None or ci_tilt is None:
        wb.close()
        return []
    node_l = node.strip().lower()
    out = []
    for r in rows:
        nv = _norm(r[ci_node]) if ci_node < len(r) else ""
        if nv.lower() != node_l:
            continue
        cell = _norm(r[ci_cell]) if ci_cell < len(r) else ""
        tilt = _norm(r[ci_tilt]) if ci_tilt < len(r) else ""
        if cell and tilt != "":
            out.append((cell, nv, tilt))
    wb.close()
    return out


def audit_etilt(targets: List[tuple],
                records: Dict[str, Dict[str, str]],
                log=lambda m: None) -> List[AuditResult]:
    """Audit each (cell, node, cddTilt) target against the node's RET /
    beamforming tilt."""
    if not targets:
        return []

    def _below_me(ldn):
        return re.sub(r"^.*?ManagedElement=[^,]+,", "", ldn)

    # RET index: (alphaPrefix, sector) → list of (retMoBelowME, tiltRaw).
    import collections
    ret_by = collections.defaultdict(list)
    for ldn, a in records.items():
        if ldn.split(",")[-1].split("=", 1)[0] != "RetSubUnit":
            continue
        p = _parse_ret(a.get("userLabel") or "")
        t = a.get("electricalAntennaTilt")
        if p and t not in (None, ""):
            ret_by[p].append((_below_me(ldn), str(t)))
    # digitalTilt fallback (AIR/beamforming, no physical RET):
    #   LTE cell → sectorCarrierRef → SectorCarrier.digitalTilt
    #   NR  cell → nRSectorCarrierRef → CommonBeamforming.digitalTilt
    sc_digital: Dict[str, tuple] = {}       # SectorCarrier below-ME → (mo, tilt)
    for ldn, a in records.items():
        if ldn.split(",")[-1].split("=", 1)[0] == "SectorCarrier" \
                and a.get("digitalTilt") not in (None, ""):
            sc_digital[_below_me(ldn)] = (_below_me(ldn), a.get("digitalTilt"))
    cb_by_nrsc: Dict[str, tuple] = {}       # NRSectorCarrier leaf → (mo, tilt)
    for ldn, a in records.items():
        if "CommonBeamforming=" in ldn and a.get("digitalTilt") not in (None, ""):
            m = re.search(r"NRSectorCarrier=([^,]+)", ldn)
            if m:
                cb_by_nrsc[m.group(1)] = (_below_me(ldn), a.get("digitalTilt"))
    # cell → carrier handle for the fallback: ('sc', SectorCarrier below-ME) or
    # ('nrsc', NRSectorCarrier leaf).
    cell_carrier: Dict[str, tuple] = {}
    for ldn, a in records.items():
        m = re.search(r"(?:^|,)EUtranCell(?:FDD|TDD)=([^,]+)$", ldn)
        if m and a.get("sectorCarrierRef"):
            cell_carrier[m.group(1)] = ("sc", _below_me(a["sectorCarrierRef"]))
            continue
        m = re.search(r"(?:^|,)NRCell(?:CU|DU)=([^,]+)$", ldn)
        if m and a.get("nRSectorCarrierRef"):
            r = re.search(r"NRSectorCarrier=([^,]+)", a["nRSectorCarrierRef"])
            if r:
                cell_carrier[m.group(1)] = ("nrsc", r.group(1))

    out: List[AuditResult] = []
    for cell, node, cdd_tilt in targets:
        pc = _parse_cell(cell)
        if not pc:
            continue
        site, band, sector = pc
        # (mo, tiltRaw, param) for every matching RET, else the NR fallback.
        matched = []
        for (ap, rs), lst in ret_by.items():
            if rs == sector and ap.startswith(site) and band in ap[len(site):]:
                matched += [(mo, t, "electricalAntennaTilt") for mo, t in lst]
        src = "RetSubUnit.electricalAntennaTilt"
        if not matched:
            # No physical RET → digital beamforming tilt on the cell's carrier.
            kind, key = cell_carrier.get(cell, (None, None))
            cb = None
            if kind == "sc":
                cb = sc_digital.get(key)
                src = "SectorCarrier.digitalTilt"
            elif kind == "nrsc":
                cb = cb_by_nrsc.get(key)
                src = "CommonBeamforming.digitalTilt"
            if cb is None:
                out.append(AuditResult(
                    "etilt", node, "RetSubUnit", "electricalAntennaTilt",
                    cdd_tilt, "", "NotFound",
                    "no RET userLabel / digitalTilt match", node, ref_cell=cell))
                continue
            matched = [(cb[0], str(cb[1]), "digitalTilt")]
        # One row per matching RET MO — each is directly settable (value ×10).
        seen = set()
        for mo, t_raw, param in matched:
            if mo in seen:
                continue
            seen.add(mo)
            status = "Match" if _eq(_tilt_deg(t_raw), cdd_tilt) else "Mismatch"
            out.append(AuditResult(
                "etilt", node, mo, param, cdd_tilt, _tilt_deg(t_raw),
                status, src, node, ref_cell=cell, norm="tilt10"))
    return out
