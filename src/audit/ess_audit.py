"""
ess_audit.py — ESS (E-UTRAN/NR Spectrum Sharing) pairing audit.

The CDD's ``ESS`` sheet pairs an LTE cell with an NR cell and gives the expected
``essScLocalId`` / ``essScPairId``. For each pair (of the audited node) this
checks, against the node dump:

  1. the LTE cell and NR cell both EXIST on the node;
  2. ``essScLocalId`` / ``essScPairId`` are present AND EQUAL on the LTE
     ``SectorCarrier`` and the NR ``NRSectorCarrier`` (and equal the sheet);
  3. ``essEnabled = true`` on the LTE→NR ``GUtranCellRelation`` and the
     NR→LTE ``EUtranCellRelation``.

Pattern (calibrated on a real dump), e.g. LTE cell GFATIML-171 ↔ NR GFATIMP-501:
  SectorCarrier=B28_S1    essScLocalId=171  essScPairId=5010000000171
  NRSectorCarrier=N28_S1  essScLocalId=171  essScPairId=5010000000171
  EUtranCellFDD=GFATIML-171,...,GUtranCellRelation=5152-<gNBID>-501  essEnabled=true
  NRCellCU=GFATIMP-501,EUtranCellRelation=GFATIML-171               essEnabled=true
``essScPairId`` = NR LocalCellID followed by the LTE LocalCellID zero-padded to
10 digits (501 + 0000000171 → 5010000000171).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, List

import openpyxl


@dataclass
class EssResult:
    node: str
    lte_cell: str
    nr_cell: str
    lte_exists: bool
    nr_exists: bool
    ess_local: str          # expected essScLocalId (from CDD)
    sc_local: str           # SectorCarrier essScLocalId (node)
    nrsc_local: str         # NRSectorCarrier essScLocalId (node)
    ess_pair: str           # expected essScPairId (from CDD)
    sc_pair: str            # SectorCarrier essScPairId (node)
    nrsc_pair: str          # NRSectorCarrier essScPairId (node)
    ess_lte: str            # essEnabled on the LTE GUtranCellRelation
    ess_nr: str             # essEnabled on the NR EUtranCellRelation
    status: str             # Match / Mismatch


def _norm(v):
    return "" if v is None else str(v).strip()


def read_ess_pairs(path: str, node: str, log=lambda m: None) -> List[dict]:
    """Read the CDD ``ESS`` sheet → pair dicts for the audited node.

    Header (row 1): eNodeBName, ENodeBID, CellName(LTE), LocalCellID(LTE),
    gNB Name, gNB ID, CellName(NR), LocalCellID(NR), NRSectorCarrier,
    NRSectorCarrier.essScLocalId, NRSectorCarrier.essScPairId, ..."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "ESS" not in wb.sheetnames:
        log("[audit/ess] no 'ESS' sheet in the LTE CDD — ESS audit skipped.")
        wb.close()
        return []
    ws = wb["ESS"]
    rows = ws.iter_rows(values_only=True)
    header = [_norm(c) for c in next(rows)]

    def col(*names):
        for i, h in enumerate(header):
            hl = h.replace("\xa0", " ").strip().lower()
            if hl in names:
                return i
        return None

    # CellName appears twice (LTE then NR); take them positionally.
    cellname_idxs = [i for i, h in enumerate(header)
                     if h.replace("\xa0", " ").strip().lower() == "cellname"]
    localid_idxs = [i for i, h in enumerate(header)
                    if h.replace("\xa0", " ").strip().lower() == "localcellid"]
    idx = {
        "enb": col("enodebname"),
        "lte_cell": cellname_idxs[0] if len(cellname_idxs) > 0 else None,
        "lte_local": localid_idxs[0] if len(localid_idxs) > 0 else None,
        "gnb": col("gnb name"),
        "gnb_id": col("gnb id"),
        "nr_cell": cellname_idxs[1] if len(cellname_idxs) > 1 else None,
        "nr_local": localid_idxs[1] if len(localid_idxs) > 1 else None,
        "ess_local": col("nrsectorcarrier.esssclocalid"),
        "ess_pair": col("nrsectorcarrier.essscpairid"),
    }
    node_l = node.strip().lower()
    pairs = []
    for r in rows:
        enb = _norm(r[idx["enb"]]) if idx["enb"] is not None else ""
        gnb = _norm(r[idx["gnb"]]) if idx["gnb"] is not None else ""
        if node_l not in (enb.lower(), gnb.lower()):
            continue
        pairs.append({
            "node": enb or gnb,
            "lte_cell": _norm(r[idx["lte_cell"]]),
            "lte_local": _norm(r[idx["lte_local"]]),
            "gnb_id": _norm(r[idx["gnb_id"]]) if idx["gnb_id"] is not None else "",
            "nr_cell": _norm(r[idx["nr_cell"]]),
            "nr_local": _norm(r[idx["nr_local"]]),
            "ess_local": _norm(r[idx["ess_local"]]),
            "ess_pair": _norm(r[idx["ess_pair"]]),
        })
    wb.close()
    return pairs


_LTE_CELL = re.compile(r"(?:^|,)EUtranCell(?:FDD|TDD)=([^,]+)$")
_NR_CELL = re.compile(r"(?:^|,)NRCell(?:CU|DU)=([^,]+)$")
_SHARING_GROUP = re.compile(r"SpectrumSharingFunction=\d+,SharingGroup=\d+$")


def count_sharing_groups(records: Dict[str, Dict[str, str]], node: str) -> int:
    """Count ``SpectrumSharingFunction=1,SharingGroup=N`` MOs on ``node`` — this
    should equal the number of ESS pairs (one sharing group per ESS cell)."""
    node_l = node.strip().lower()
    n = 0
    for ldn in records:
        if _SHARING_GROUP.search(ldn):
            m = re.search(r"ManagedElement=([^,]+)", ldn)
            if m and m.group(1).lower() == node_l:
                n += 1
    return n


def audit_ess(pairs: List[dict], records: Dict[str, Dict[str, str]],
              log=lambda m: None) -> List[EssResult]:
    """Audit the ESS pairs against the node dump records."""
    if not pairs:
        return []
    # Existing cells on the node.
    lte_cells, nr_cells = set(), set()
    # essScLocalId → essScPairId, per carrier type (non-zero only).
    sc_pair_by_local: Dict[str, str] = {}
    nrsc_pair_by_local: Dict[str, str] = {}
    # essEnabled lookups.
    ess_lte: Dict[tuple, str] = {}   # (lteCell, gNBID, nrLocalId) → essEnabled
    ess_nr: Dict[tuple, str] = {}    # (nrCell, lteCell)           → essEnabled

    for ldn, a in records.items():
        m = _LTE_CELL.search(ldn)
        if m:
            lte_cells.add(m.group(1))
        m = _NR_CELL.search(ldn)
        if m:
            nr_cells.add(m.group(1))
        leaf = ldn.split(",")[-1].split("=", 1)[0]
        if leaf == "SectorCarrier":
            sl = _norm(a.get("essScLocalId"))
            if sl and sl != "0":
                sc_pair_by_local.setdefault(sl, _norm(a.get("essScPairId")))
        elif leaf == "NRSectorCarrier":
            sl = _norm(a.get("essScLocalId"))
            if sl and sl != "0":
                nrsc_pair_by_local.setdefault(sl, _norm(a.get("essScPairId")))
        ess = a.get("essEnabled")
        if ess is not None:
            lm = re.search(r"EUtranCell(?:FDD|TDD)=([^,]+)", ldn)
            # GUtranCellRelation id = "<x>-<gNBID padded>-<nrLocalId>"; match the
            # specific gNB so a neighbour's relation ending in the same localid
            # can't clobber this pair's essEnabled.
            gr = re.search(r"GUtranCellRelation=\d+-0*(\d+)-(\d+)$", ldn)
            if lm and gr:
                ess_lte[(lm.group(1), gr.group(1), gr.group(2))] = \
                    _norm(ess).lower()
            nm = re.search(r"NRCell(?:CU|DU)=([^,]+)", ldn)
            er = re.search(r"EUtranCellRelation=([^,]+)$", ldn)
            if nm and er:
                ess_nr[(nm.group(1), er.group(1))] = _norm(ess).lower()

    out: List[EssResult] = []
    for p in pairs:
        lte_cell, nr_cell = p["lte_cell"], p["nr_cell"]
        ess_local, ess_pair = p["ess_local"], p["ess_pair"]
        lte_exists = lte_cell in lte_cells
        nr_exists = nr_cell in nr_cells
        sc_pair = sc_pair_by_local.get(ess_local, "")
        nrsc_pair = nrsc_pair_by_local.get(ess_local, "")
        sc_local = ess_local if ess_local in sc_pair_by_local else ""
        nrsc_local = ess_local if ess_local in nrsc_pair_by_local else ""
        e_lte = ess_lte.get((lte_cell, p["gnb_id"], p["nr_local"]), "")
        e_nr = ess_nr.get((nr_cell, lte_cell), "")

        ok = (lte_exists and nr_exists
              and sc_local == ess_local and sc_pair == ess_pair
              and nrsc_local == ess_local and nrsc_pair == ess_pair
              and e_lte == "true" and e_nr == "true")
        out.append(EssResult(
            node=p["node"], lte_cell=lte_cell, nr_cell=nr_cell,
            lte_exists=lte_exists, nr_exists=nr_exists,
            ess_local=ess_local, sc_local=sc_local, nrsc_local=nrsc_local,
            ess_pair=ess_pair, sc_pair=sc_pair, nrsc_pair=nrsc_pair,
            ess_lte=e_lte or "(none)", ess_nr=e_nr or "(none)",
            status="Match" if ok else "Mismatch"))
    return out
