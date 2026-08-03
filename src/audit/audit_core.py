"""
audit_core.py — compare CDD expected values against parsed node dump records
and write an Excel report.

``normalize`` + the Match/Mismatch/NotFound verdict are ported from
enp-generator ``services/audit_engine.py`` / ``services/node_audit.py``.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Value normalization (from enp-generator audit_engine.normalize) ──
def normalize(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else str(f)
    except (ValueError, TypeError):
        return s.lower()


def normalize_segments(v) -> str:
    """Normalize a hyphen-separated value (e.g. CGI ``MCC-MNC-LAC-CI``) by
    stripping leading zeros from each numeric segment, so ``515-02-00087-60031``
    equals ``515-02-87-60031``. Non-numeric segments fall back to ``normalize``.
    """
    if v is None:
        return ""
    parts = str(v).strip().split("-")
    out = []
    for p in parts:
        p = p.strip()
        out.append(str(int(p)) if p.isdigit() else normalize(p))
    return "-".join(out)


_LATLON_B = re.compile(r"^([NSEW])\s*0*(\d+)[-\s]0*(\d+)[-\s]0*(\d+(?:\.\d+)?)",
                       re.IGNORECASE)                         # N07-40-08.15
_LATLON_A = re.compile(r"0*(\d+)\s*[°\-]\s*0*(\d+)\s*['’\-]\s*"
                       r"0*(\d+(?:\.\d+)?)\s*[\"”\-]?\s*([NSEW])",
                       re.IGNORECASE)                         # 7°40'8.15"N


def normalize_latlong(v) -> str:
    """Normalize a DMS coordinate to signed decimal degrees so the CDD form
    (e.g. 7°40'8.15\"N) and the node form (e.g. N07-40-08.15) compare equal."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() == "null":
        return ""
    m = _LATLON_B.match(s)
    if m:
        hemi, d, mn, sec = m.group(1), m.group(2), m.group(3), m.group(4)
    else:
        m = _LATLON_A.search(s)
        if not m:
            return s.lower()
        d, mn, sec, hemi = m.group(1), m.group(2), m.group(3), m.group(4)
    dec = int(d) + int(mn) / 60.0 + float(sec) / 3600.0
    if hemi.upper() in ("S", "W"):
        dec = -dec
    return f"{dec:.5f}"


def normalize_list(v) -> str:
    """Normalize a multi-value field so different renderings compare equal:
    CDD ``0 1 2 3`` or ``1&37`` vs node ``[0, 1, 2, 3]`` / ``[1, 37]``. Splits on
    brackets/comma/ampersand/space, numeric-normalizes each token, rejoins."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() == "null":
        return ""
    parts = [p for p in re.split(r"[\[\],&\s]+", s) if p != ""]
    out = []
    for p in parts:
        try:
            f = float(p)
            out.append(str(int(f)) if f.is_integer() else str(f))
        except (ValueError, TypeError):
            out.append(p.lower())
    return ",".join(out)


def normalize_geo(v) -> str:
    """Normalize LTE/NR coordinates: the dump stores integer **micro-degrees**
    (e.g. 6999000 = 6.999°), the CDD stores decimal degrees (6.999). Convert
    both to decimal degrees so they compare equal."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() == "null":
        return ""
    if re.fullmatch(r"-?\d+", s):          # integer micro-degrees (from dump)
        return f"{int(s) / 1e6:.5f}"
    try:
        return f"{float(s):.5f}"           # decimal degrees (from CDD)
    except (ValueError, TypeError):
        return s.lower()


def normalize_bbtype(v) -> str:
    """Normalize a baseband/RAN-processor type so the CDD short form matches the
    node's ``productName``: CDD ``RP6655`` / ``BB6621`` vs node ``RAN Processor
    6655`` / ``Baseband 6621`` — compare on the trailing product number."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() == "null":
        return ""
    m = re.search(r"(\d{4})", s)
    return m.group(1) if m else s.lower()


# name -> comparator; selected per-column via the CDD map's optional "norm".
_NORMALIZERS = {"segments": normalize_segments, "latlong": normalize_latlong,
                "list": normalize_list, "geo": normalize_geo,
                "bbtype": normalize_bbtype}


# Boolean equivalence for the compare: many GSM/LTE attrs are ``1``/``0`` in the
# CDD but ``ACTIVE``/``INACTIVE`` (or ON/OFF, …) on the node — the same value.
_BOOL_ON = {"1", "on", "active", "true", "enabled", "yes", "unlocked"}
_BOOL_OFF = {"0", "off", "inactive", "false", "disabled", "no", "locked"}


def _bool_canon(x):
    s = str(x).strip().lower()
    if s in _BOOL_ON:
        return "1"
    if s in _BOOL_OFF:
        return "0"
    return None


def _bool_equal(a, b) -> bool:
    """True when both sides are boolean-like and denote the same state, so CDD
    ``1`` == node ``ACTIVE`` (and ``0`` == ``INACTIVE``/``OFF``/…)."""
    ca, cb = _bool_canon(a), _bool_canon(b)
    return ca is not None and ca == cb


@dataclass
class AuditItem:
    """One expected (MO, parameter) coming from the CDD."""
    category: str          # node | cell | relation
    tech: str              # lte_nr | gsm
    mo_local: str          # FDN below ManagedElement, e.g.
                           #   ENodeBFunction=1,EUtranCellFDD=BOALANG-L1
    parameter: str
    expected: str
    key: str               # human key (cell / node name) for the report
    source: str            # "<sheet>!<column>" for traceability
    norm: str = ""         # optional comparator hint (e.g. "segments" for CGI)
    via_ref: str = ""      # follow this ref attr on the MO to the target MO
                           #   (e.g. "sectorCarrierRef") before reading parameter
    attr_format: str = ""  # build actual from >1 attrs, e.g.
                           #   "{noOfTxAntennas}T{noOfRxAntennas}R" → "4T4R"
    node: str = ""         # node this row belongs to (eNodeBName/gNBName…)


@dataclass
class AuditResult:
    category: str
    key: str
    mo: str
    parameter: str
    expected: str
    actual: str
    status: str            # Match | Mismatch | NotFound | MO_NotFound
    source: str
    node: str = ""
    ref_cell: str = ""     # the cell a via_ref param belongs to (MO points to
                           #   the real MO, e.g. SectorCarrier)
    norm: str = ""         # comparator hint carried from the CDD item, so a
                           #   generated set can re-format the value to the
                           #   node's convention (list/latlong/geo/…)


@dataclass
class LldResult:
    """One LLD (physical baseband / CPRI) row — one planned link (or the
    baseband unit), with LLD-vs-Node values side by side. Kept separate from the
    logical CDD ``AuditResult`` so it gets its own sheet with paired
    (LLD | Node) columns rather than the generic MO/Parameter ones.

    A link is paired to a node RiLink by BB port first, then by radio (band) —
    so an AAS radio planned on port P but wired on port H still lines up on one
    row, exposing the port mapping instead of reporting a phantom hole."""
    node: str
    bbid: str                  # "BB1" (from the node's B0<k> suffix)
    ref_cell: str = ""         # Sector/Radio, e.g. "S2/R1"
    bb_port_lld: str = ""      # BB RiPort — planned vs actual
    bb_port_node: str = ""
    hw_type_lld: str = ""      # baseband type or radio type — planned vs actual
    hw_type_node: str = ""
    data_port_lld: str = ""    # Radio DATA port — planned vs actual
    data_port_node: str = ""
    status: str = ""           # Match | Mismatch | NotFound | Extra
    source: str = ""
    # per-metric verdicts (False → that LLD|Node pair differs → highlighted).
    bb_ok: bool = True
    hw_ok: bool = True
    data_ok: bool = True


def _index_records(records: Dict[str, Dict[str, str]]):
    """Build matching indexes for CDD ``mo_local`` → dump/cmedit records.

    Returns ``(g_exact, g_leaf, by_node)`` where:
      * g_exact — global: FDN below ManagedElement → attrs.
      * g_leaf  — global: last MO segment ``Class=id`` → attrs (BSC GSM,
                  node-name-independent).
      * by_node — per-node {node: {"exact": {...}, "leaf": {...}}}, so that
                  when several nodes' dumps are merged, a node MO like
                  ``ENodeBFunction=1`` doesn't collide across nodes.
    All keys lower-cased."""
    g_exact: Dict[str, Dict[str, str]] = {}
    g_leaf: Dict[str, Dict[str, str]] = {}
    by_node: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = {}
    for ldn, attrs in records.items():
        local = re.sub(r"^.*?ManagedElement=[^,]+,?", "", ldn).lower()
        last = ldn.split(",")[-1].strip().lower()
        g_exact[local] = attrs
        if "=" in last:
            g_leaf[last] = attrs
        m = re.search(r"ManagedElement=([^,]+)", ldn)
        if m:
            nd = by_node.setdefault(m.group(1).lower(),
                                    {"exact": {}, "leaf": {}})
            nd["exact"][local] = attrs
            if "=" in last:
                nd["leaf"][last] = attrs
    return g_exact, g_leaf, by_node


def _find_mo(mo: str, node: str, g_exact, g_leaf, by_node):
    """Look up an MO's attrs, preferring the owning node's index (avoids
    cross-node collisions) and falling back to the global index (GSM/BSC)."""
    mo_l = mo.lower()
    single = "," not in mo
    nd = by_node.get((node or "").lower())
    if nd is not None:
        hit = nd["exact"].get(mo_l)
        if hit is None and single:
            hit = nd["leaf"].get(mo_l)
        if hit is not None:
            return hit
    hit = g_exact.get(mo_l)
    if hit is None and single:
        hit = g_leaf.get(mo_l)
    return hit


_CELL_RE = re.compile(
    r"(?:EUtranCellFDD|EUtranCellTDD|NRCellDU|NRCellCU|GeranCell)=([^,]+)")


def _cell_of(mo_local: str) -> str:
    """Extract the cell identity from an MO FDN, for the Reference Cell column."""
    m = _CELL_RE.search(mo_local)
    return m.group(1) if m else ""


def _resolve_ref(attrs: Dict[str, str], ref_attr: str, g_exact, g_leaf, by_node):
    """Follow a reference attribute (e.g. ``sectorCarrierRef`` on EUtranCellFDD,
    whose value is a full FDN to a SectorCarrier MO) to the referenced MO. The
    ref FDN carries its own ManagedElement, so we resolve within that node to
    stay collision-free. Returns ``(below_ME_fdn, target_attrs)`` — the FDN so
    the report/script can point at the REAL MO, not the referencing cell."""
    low = {k.lower(): v for k, v in attrs.items()}
    ref = attrs.get(ref_attr) or low.get(ref_attr.lower())
    if not ref:
        return "", None
    ref = str(ref)
    node_m = re.search(r"ManagedElement=([^,]+)", ref)
    below = re.sub(r"^.*?ManagedElement=[^,]+,", "", ref).strip()
    node = node_m.group(1) if node_m else ""
    return below, _find_mo(below.lower(), node, g_exact, g_leaf, by_node)


def compare(items: List[AuditItem],
            records: Dict[str, Dict[str, str]]) -> List[AuditResult]:
    """Compare each CDD AuditItem against the node dump / cmedit records."""
    g_exact, g_leaf, by_node = _index_records(records)
    results: List[AuditResult] = []
    for it in items:
        ref_cell = _cell_of(it.mo_local)
        eff_mo = it.mo_local          # the MO the report/script should target
        attrs = _find_mo(it.mo_local, it.node, g_exact, g_leaf, by_node)
        if attrs is None:
            results.append(AuditResult(
                it.category, it.key, it.mo_local, it.parameter,
                it.expected, "", "MO_NotFound", it.source, it.node, ref_cell,
                it.norm))
            continue
        # Follow a reference (e.g. sectorCarrierRef → SectorCarrier MO) when the
        # audited attribute lives on the referenced MO. The MO reported becomes
        # the REAL MO (SectorCarrier), and the cell moves to the ref_cell column.
        lookup = attrs
        if it.via_ref:
            rmo, target = _resolve_ref(attrs, it.via_ref, g_exact, g_leaf, by_node)
            if target is None:
                results.append(AuditResult(
                    it.category, it.key, it.mo_local, it.parameter,
                    it.expected, "", "MO_NotFound", it.source, it.node, ref_cell,
                    it.norm))
                continue
            lookup = target
            eff_mo = rmo or it.mo_local
        low = {k.lower(): v for k, v in lookup.items()}
        # Composite value (display-only): rebuild the actual into the CDD's own
        # format, e.g. "{noOfTxAntennas}T{noOfRxAntennas}R" → "4T4R". Prefer the
        # ``split`` mechanism (real component attrs) for anything you also want
        # to emit as a set line.
        if it.attr_format:
            missing = []

            def _repl(m, _low=low):
                name = m.group(1)
                v = _low.get(name.lower())
                if v is None or str(v) == "":
                    missing.append(name)
                    return ""
                return str(v)

            actual_str = re.sub(r"\{([^}]+)\}", _repl, it.attr_format)
            if missing:
                status, actual_str = "NotFound", ""
            else:
                norm = _NORMALIZERS.get(it.norm, normalize)
                status = ("Match" if (norm(actual_str) == norm(it.expected)
                                      or _bool_equal(it.expected, actual_str))
                          else "Mismatch")
            results.append(AuditResult(
                it.category, it.key, eff_mo, it.parameter,
                it.expected, actual_str, status, it.source, it.node, ref_cell,
                it.norm))
            continue
        # Attribute lookup is case-insensitive (dump uses canonical casing).
        actual = lookup.get(it.parameter)
        if actual is None:
            actual = low.get(it.parameter.lower())
        if actual is None:
            status = "NotFound"
            actual_str = ""
        else:
            actual_str = str(actual)
            norm = _NORMALIZERS.get(it.norm, normalize)
            status = ("Match" if (norm(actual_str) == norm(it.expected)
                                  or _bool_equal(it.expected, actual_str))
                      else "Mismatch")
        results.append(AuditResult(
            it.category, it.key, eff_mo, it.parameter,
            it.expected, actual_str, status, it.source, it.node, ref_cell,
            it.norm))
    return results


_CELL_MO_INV = re.compile(
    r"(?:^|,)(EUtranCellFDD|EUtranCellTDD|NRCellDU|NRCellCU|GeranCell)=([^,]+)",
    re.IGNORECASE)


def cell_inventory_check(items: List[AuditItem],
                         records: Dict[str, Dict[str, str]]) -> List[AuditResult]:
    """Compare the SET of cells the CDD defines against the cells actually on
    the node(s), per cell MO class: the count and the exact names. Produces a
    'Cell count' row plus, when they differ, a 'Missing on node' / 'Extra on
    node' row so the operator sees which cell names don't line up."""
    import collections
    cdd = collections.defaultdict(set)      # class_lower -> {id_lower}
    actual = collections.defaultdict(set)
    disp: Dict[str, str] = {}               # id_lower -> original casing

    for it in items:
        if it.category != "cell":
            continue
        m = _CELL_MO_INV.search("," + it.mo_local)
        if not m:
            continue
        cls, cid = m.group(1).lower(), m.group(2)
        cdd[cls].add(cid.lower())
        disp[cid.lower()] = cid
    for ldn in records:
        m = _CELL_MO_INV.search(ldn)
        if not m:
            continue
        cls, cid = m.group(1).lower(), m.group(2)
        actual[cls].add(cid.lower())
        disp.setdefault(cid.lower(), cid)

    _CANON = {"eutrancellfdd": "EUtranCellFDD", "eutrancelltdd": "EUtranCellTDD",
              "nrcelldu": "NRCellDU", "nrcellcu": "NRCellCU",
              "gerancell": "GeranCell"}
    out: List[AuditResult] = []
    for cls in sorted(cdd):                  # only classes the CDD defines
        exp, act = cdd[cls], actual.get(cls, set())
        name = _CANON.get(cls, cls)
        missing = sorted(disp.get(x, x) for x in (exp - act))
        extra = sorted(disp.get(x, x) for x in (act - exp))
        ok = (len(exp) == len(act) and not missing and not extra)
        out.append(AuditResult(
            "cell-count", name, name, "cell count",
            str(len(exp)), str(len(act)),
            "Match" if ok else "Mismatch", "CDD vs node cells"))
        if missing:
            out.append(AuditResult(
                "cell-count", name, name, "missing on node",
                ", ".join(missing), "(not found)", "Mismatch",
                "in CDD, absent on node"))
        if extra:
            out.append(AuditResult(
                "cell-count", name, name, "extra on node",
                "(not in CDD)", ", ".join(extra), "Mismatch",
                "on node, absent from CDD"))
    return out


def _col_attrs(col: dict) -> List[str]:
    out = []
    if col.get("attr"):
        out.append(col["attr"].lower())
    for sp in col.get("split") or []:
        if sp.get("attr"):
            out.append(sp["attr"].lower())
    return out


def audited_attr_sets(audit_map: dict):
    """Derive the audited attribute universe for the *reverse* pass (node params
    with no CDD row). Returns ``(gsm_attrs, lte_class_attrs, ref_attrs)``:

      * gsm_attrs      — every GSM attribute (child-MO attrs are merged onto the
                         GeranCell record, so match by name alone).
      * lte_class_attrs— {(leaf MO class, attr)} for LTE/NR, so a short attr name
                         only matches on the right MO class.
      * ref_attrs      — via_ref/split attrs (on a referenced MO, e.g.
                         SectorCarrier); matched by name (they're distinctive).
    """
    gsm_attrs = set()
    lte_class_attrs = set()
    ref_attrs = set()
    for p in audit_map.get("profiles", []):
        cols = p.get("columns", [])
        if p.get("tech") == "gsm":
            for c in cols:
                gsm_attrs.update(_col_attrs(c))
            continue
        classes = set()
        leaf = p.get("mo_fdn", "").split(",")[-1].split("=")[0].strip()
        if leaf and "{" not in leaf:
            classes.add(leaf.lower())
        cmm = p.get("cell_mo_map") or {}
        for v in (cmm.get("map") or {}).values():
            classes.add(str(v).lower())
        if cmm.get("default"):
            classes.add(str(cmm["default"]).lower())
        for c in cols:
            cls = classes
            per_mo = c.get("mo")
            if per_mo:
                l = per_mo.split(",")[-1].split("=")[0].strip()
                if l and "{" not in l:
                    cls = {l.lower()}
            for a in _col_attrs(c):
                if c.get("via_ref"):
                    ref_attrs.add(a)
                else:
                    for cc in cls:
                        lte_class_attrs.add((cc, a))
    return gsm_attrs, lte_class_attrs, ref_attrs


def build_reverse_rows(node: str, records: Dict[str, Dict[str, str]],
                       gsm_attrs, lte_class_attrs, ref_attrs,
                       cell_rx=None) -> List[AuditResult]:
    """For a node with NO CDD rows, surface its ACTUAL audited parameters —
    CDD (expected) left blank, status ``CDD_missing``. LTE/NR MOs are matched by
    ``ManagedElement=<node>``; GSM GeranCells (BSC-level, no node in the FDN) by
    the site's cell-id regex."""
    out: List[AuditResult] = []
    nl = node.lower()
    for ldn, attrs in records.items():
        me = re.search(r"ManagedElement=([^,]+)", ldn)
        leaf_cls = ldn.split(",")[-1].split("=")[0].strip().lower()
        if me:
            if me.group(1).lower() != nl:
                continue
            local = re.sub(r"^.*?ManagedElement=[^,]+,?", "", ldn)
            is_gsm = False
        else:
            m = re.search(r"GeranCell=([^,]+)", ldn)
            if not (cell_rx is not None and m and cell_rx.match(m.group(1))):
                continue
            local = ldn
            is_gsm = True
        cell = _cell_of(local)
        for attr, val in attrs.items():
            if attr.startswith("__") or val is None or str(val).strip() == "":
                continue
            al = attr.lower()
            # ref_attrs (MIMO/tilt/power on a referenced MO) only count on a
            # *SectorCarrier — not on lookalike attrs of other MOs (e.g. an
            # ExternalEUtranCellFDD that also carries noOfTxAntennas).
            hit = (al in gsm_attrs) if is_gsm else (
                (leaf_cls, al) in lte_class_attrs
                or (al in ref_attrs and leaf_cls.endswith("sectorcarrier")))
            if hit:
                out.append(AuditResult(
                    "gsm" if is_gsm else "cell", cell or node, local, attr,
                    "", str(val), "CDD_missing", "node (no CDD)", node, cell))
    return out


# ── Excel report ────────────────────────────────────────────────────
_FILL_MISMATCH = PatternFill("solid", fgColor="FFC7CE")   # red
_FILL_NOTFOUND = PatternFill("solid", fgColor="FFEB9C")   # yellow
_FILL_MATCH = PatternFill("solid", fgColor="C6EFCE")      # green
_FILL_EXTRA = PatternFill("solid", fgColor="BDD7EE")      # blue (unplanned)
_FILL_CDDMISSING = PatternFill("solid", fgColor="E2CFF3")  # purple (node-only)
_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_STATUS_FILL = {"Mismatch": _FILL_MISMATCH, "NotFound": _FILL_NOTFOUND,
                "MO_NotFound": _FILL_NOTFOUND, "Match": _FILL_MATCH,
                "Extra": _FILL_EXTRA, "CDD_missing": _FILL_CDDMISSING}


_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _write_summary(summ, results, meta, lld_results=None):
    counts = {"Match": 0, "Mismatch": 0, "NotFound": 0, "MO_NotFound": 0,
              "CDD_missing": 0}
    for r in results:
        counts[r.status] = counts.get(r.status, 0) + 1
    total = len(results)
    # CDD-missing rows are node-only (no CDD to grade against) — kept out of the
    # graded total / compliance, shown as a separate informational KPI.
    graded = total - counts["CDD_missing"]

    # ── Title banner ─────────────────────────────────────────────
    summ.merge_cells("A1:C1")
    t = summ["A1"]
    t.value = "CDD Audit Report"
    t.font = Font(bold=True, size=18, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="2F5597")
    t.alignment = _CENTER
    summ.row_dimensions[1].height = 32

    # ── Meta block ───────────────────────────────────────────────
    r = 3
    for k, v in meta.items():
        lc = summ.cell(r, 1, k)
        summ.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        vc = summ.cell(r, 2, v)
        lc.font = Font(bold=True, color="1F3864")
        lc.fill = PatternFill("solid", fgColor="E7EFF9")
        for c in (lc, vc):
            c.border = _BORDER
            c.alignment = _LEFT
        r += 1

    # ── Results section ──────────────────────────────────────────
    r += 1
    summ.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    h = summ.cell(r, 1, "Results")
    h.font = Font(bold=True, size=13, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor="4472C4")
    h.alignment = _CENTER
    r += 1
    for c, txt in enumerate(("Metric", "Count", "Share"), 1):
        cc = summ.cell(r, c, txt)
        cc.font = Font(bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor="8EA9DB")
        cc.alignment = _CENTER
        cc.border = _BORDER
    r += 1

    def kpi(label, count, fill, font_color, pct=True):
        nonlocal r
        share = f"{count / graded * 100:.1f}%" if (graded and pct) else ""
        a = summ.cell(r, 1, label)
        b = summ.cell(r, 2, count)
        c = summ.cell(r, 3, share)
        for cell in (a, b, c):
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(bold=True, color=font_color)
            cell.border = _BORDER
            cell.alignment = _CENTER
        a.alignment = _LEFT
        r += 1

    kpi("Total checks", graded, "D9E1F2", "1F3864", pct=False)
    kpi("✔  Match", counts["Match"], "C6EFCE", "006100")
    kpi("✖  Mismatch", counts["Mismatch"], "FFC7CE", "9C0006")
    kpi("⚠  Parameter not found", counts["NotFound"], "FFEB9C", "9C6500")
    kpi("⚠  MO not found", counts["MO_NotFound"], "F8CBAD", "833C00")
    if counts["CDD_missing"]:
        kpi("● CDD missing (node-only)", counts["CDD_missing"],
            "E2CFF3", "5B2A86", pct=False)

    # ── Compliance headline ──────────────────────────────────────
    r += 1
    comp = counts["Match"] / graded * 100 if graded else 0.0
    lc = summ.cell(r, 1, "Compliance")
    lc.font = Font(bold=True, size=12, color="1F3864")
    lc.alignment = _LEFT
    summ.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    vc = summ.cell(r, 2, f"{comp:.1f}%")
    good, warn = comp >= 95, comp >= 80
    fill = "C6EFCE" if good else ("FFEB9C" if warn else "FFC7CE")
    fcol = "006100" if good else ("9C6500" if warn else "9C0006")
    vc.fill = PatternFill("solid", fgColor=fill)
    vc.font = Font(bold=True, size=12, color=fcol)
    vc.alignment = _CENTER
    for cell in (lc, vc):
        cell.border = _BORDER

    # ── Per-node breakdown (batch/cluster audits) ────────────────
    nodes = {}
    for res in results:
        n = res.node or res.key or "-"
        nd = nodes.setdefault(n, {"Match": 0, "Mismatch": 0,
                                  "NotFound": 0, "MO_NotFound": 0})
        nd[res.status] = nd.get(res.status, 0) + 1
    if len(nodes) > 1:
        r += 2
        summ.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        h = summ.cell(r, 1, "Per-node breakdown")
        h.font = Font(bold=True, size=13, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="4472C4")
        h.alignment = _CENTER
        r += 1
        for c, txt in enumerate(("Node", "Match", "Mismatch",
                                 "Not found", "MO not found", "Compliance"), 1):
            cc = summ.cell(r, c, txt)
            cc.font = Font(bold=True, color="FFFFFF")
            cc.fill = PatternFill("solid", fgColor="8EA9DB")
            cc.alignment = _CENTER
            cc.border = _BORDER
        r += 1
        for n in sorted(nodes):
            nd = nodes[n]
            tot = sum(nd.values())
            pc = nd["Match"] / tot * 100 if tot else 0.0
            g, w = pc >= 95, pc >= 80
            pfill = "C6EFCE" if g else ("FFEB9C" if w else "FFC7CE")
            pfcol = "006100" if g else ("9C6500" if w else "9C0006")
            vals = [n, nd["Match"], nd["Mismatch"], nd["NotFound"],
                    nd["MO_NotFound"], f"{pc:.1f}%"]
            for c, v in enumerate(vals, 1):
                cc = summ.cell(r, c, v)
                cc.border = _BORDER
                cc.alignment = _LEFT if c == 1 else _CENTER
            summ.cell(r, 2).fill = PatternFill("solid", fgColor="C6EFCE")
            summ.cell(r, 3).fill = PatternFill("solid", fgColor="FFC7CE")
            pcell = summ.cell(r, 6)
            pcell.fill = PatternFill("solid", fgColor=pfill)
            pcell.font = Font(bold=True, color=pfcol)
            r += 1
        for col, wdt in (("D", 14), ("E", 16), ("F", 14)):
            summ.column_dimensions[col].width = wdt

    # ── LLD (physical baseband / CPRI) block ─────────────────────
    if lld_results:
        lc = {"Match": 0, "Mismatch": 0, "NotFound": 0, "Extra": 0}
        for x in lld_results:
            lc[x.status] = lc.get(x.status, 0) + 1
        graded = lc["Match"] + lc["Mismatch"] + lc["NotFound"]
        r += 2
        summ.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        h = summ.cell(r, 1, "LLD — baseband & CPRI (see 'LLD' sheet)")
        h.font = Font(bold=True, size=13, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="4472C4")
        h.alignment = _CENTER
        r += 1
        rows = [
            ("Checks (graded)", graded, "D9E1F2", "1F3864", False),
            ("✔  Match", lc["Match"], "C6EFCE", "006100", True),
            ("✖  Mismatch", lc["Mismatch"], "FFC7CE", "9C0006", True),
            ("⚠  Not found", lc["NotFound"], "FFEB9C", "9C6500", True),
            ("➕ Unplanned (node extra)", lc["Extra"], "BDD7EE", "1F3864", False),
        ]
        for label, count, fill, fcol, pct in rows:
            share = f"{count / graded * 100:.1f}%" if (graded and pct) else ""
            a = summ.cell(r, 1, label)
            b = summ.cell(r, 2, count)
            c = summ.cell(r, 3, share)
            for cell in (a, b, c):
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(bold=True, color=fcol)
                cell.border = _BORDER
                cell.alignment = _CENTER
            a.alignment = _LEFT
            r += 1
        comp = lc["Match"] / graded * 100 if graded else 0.0
        g, w = comp >= 95, comp >= 80
        lc2 = summ.cell(r, 1, "LLD compliance")
        lc2.font = Font(bold=True, size=12, color="1F3864")
        lc2.alignment = _LEFT
        summ.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        vc = summ.cell(r, 2, f"{comp:.1f}%")
        vc.fill = PatternFill("solid", fgColor=(
            "C6EFCE" if g else ("FFEB9C" if w else "FFC7CE")))
        vc.font = Font(bold=True, size=12, color=(
            "006100" if g else ("9C6500" if w else "9C0006")))
        vc.alignment = _CENTER
        for cell in (lc2, vc):
            cell.border = _BORDER

    summ.column_dimensions["A"].width = 40
    summ.column_dimensions["B"].width = 24
    summ.column_dimensions["C"].width = 22


def _write_lld_sheet(ws, lld_results: List["LldResult"]) -> None:
    """Physical baseband / CPRI checks on their own sheet, with a two-row header
    grouping each metric into paired LLD | Node columns:

        Node │ BBID │ Sector/Radio │ BB RI Port │  HW Type  │ Radio DATA Port │ Status
                                     LLD │ Node   LLD │ Node    LLD │ Node

    One row per planned link (or the baseband unit). Status colour matches the
    Detail sheet (green/red/yellow) plus blue for unplanned (Extra) node links.
    """
    # ── header (rows 1-2) ────────────────────────────────────────
    #  single (2-row-merged) columns, then 3 paired groups, then Status.
    ws.merge_cells("A1:A2"); ws["A1"] = "Node"
    ws.merge_cells("B1:B2"); ws["B1"] = "BBID"
    ws.merge_cells("C1:C2"); ws["C1"] = "Sector/Radio"
    ws.merge_cells("D1:E1"); ws["D1"] = "BB RI Port"
    ws.merge_cells("F1:G1"); ws["F1"] = "HW Type"
    ws.merge_cells("H1:I1"); ws["H1"] = "Radio DATA Port"
    ws.merge_cells("J1:J2"); ws["J1"] = "Status"
    subs = {4: "LLD", 5: "Node", 6: "LLD", 7: "Node", 8: "LLD", 9: "Node"}
    for c, txt in subs.items():
        ws.cell(2, c, txt)
    for row in (1, 2):
        for c in range(1, 11):
            cell = ws.cell(row, c)
            cell.fill = _FILL_HEADER
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER

    status_col = 10

    def _sort_key(r):
        # baseband row (no port) first per node, then by planned/actual port
        port = r.bb_port_lld or r.bb_port_node
        return (r.node, r.bbid, port == "", port)

    for r in sorted(lld_results, key=_sort_key):
        ws.append([r.node, r.bbid, r.ref_cell,
                   r.bb_port_lld, r.bb_port_node,
                   r.hw_type_lld, r.hw_type_node,
                   r.data_port_lld, r.data_port_node, r.status])
        row = ws.max_row
        fill = _STATUS_FILL.get(r.status)
        if fill:
            ws.cell(row=row, column=status_col).fill = fill
        # Highlight the specific LLD|Node pair(s) that differ (yellow).
        for ok, cols in ((r.bb_ok, (4, 5)), (r.hw_ok, (6, 7)),
                         (r.data_ok, (8, 9))):
            if not ok:
                for c in cols:
                    ws.cell(row, c).fill = _FILL_NOTFOUND
        for c in range(1, 11):
            ws.cell(row, c).border = _BORDER

    ws.freeze_panes = "A3"
    widths = {"A": 34, "B": 6, "C": 12, "D": 8, "E": 8, "F": 20,
              "G": 22, "H": 10, "I": 10, "J": 10}
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt


def write_excel(results: List[AuditResult], out_path: str, meta: dict,
                lld_results: Optional[List["LldResult"]] = None) -> str:
    wb = Workbook()

    # Summary sheet
    summ = wb.active
    summ.title = "Summary"
    _write_summary(summ, results, meta, lld_results)

    # Detail sheet
    ws = wb.create_sheet("Detail")
    headers = ["Category", "Node", "Reference Cell",
               "MO (below ManagedElement)", "Parameter",
               "CDD (expected)", "Node (actual)", "Status", "Source"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = _FILL_HEADER
        c.font = _HEADER_FONT
    status_col = 8   # 1-based index of the Status column
    for r in results:
        ws.append([r.category, (r.node or r.key), r.ref_cell, r.mo, r.parameter,
                   r.expected, r.actual, r.status, r.source])
        row = ws.max_row
        fill = _STATUS_FILL.get(r.status)
        if fill:
            ws.cell(row=row, column=status_col).fill = fill

    ws.freeze_panes = "A2"
    # Auto-size the Detail sheet only; Summary has hand-tuned widths + merges.
    for col in ws.columns:
        width = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)), 60))
        ws.column_dimensions[letter].width = max(width + 2, 10)

    # LLD sheet (physical baseband / CPRI) — only when there are LLD checks.
    if lld_results:
        _write_lld_sheet(wb.create_sheet("LLD"), lld_results)

    wb.save(out_path)
    return out_path


# ── moshell SetParameter script generation ──────────────────────────
# Top-level MO areas in the order a moshell alignment script conventionally
# lists them; anything else is appended after, alphabetically.
_MO_GROUP_ORDER = [
    "Transport", "SystemFunctions", "NodeSupport", "Equipment",
    "ManagedElement", "ENodeBFunction", "GNBDUFunction", "GNBCUCPFunction",
    "GNBCUUPFunction", "NRNetwork", "GeranCell",
]


def _mo_group(mo: str) -> str:
    """Top-level MO class of an FDN, e.g. 'ENodeBFunction=1,EUtranCellFDD=X'
    → 'ENodeBFunction'."""
    first = mo.split(",", 1)[0]
    return first.split("=", 1)[0].strip() or "Other"


def _banner(name: str) -> str:
    bar = "-" * (len(name) + 2)
    return f"# {bar}\n# {name}\n# {bar}"


def generate_moshell_scripts(results: List[AuditResult], out_dir: str,
                             site: str, audit_xlsx: str,
                             generated_by: str = "",
                             statuses=("Mismatch",)) -> List[str]:
    """Write one moshell ``set`` script per node from the audit results whose
    status is in ``statuses`` (default: Mismatch only). Each ``set`` line uses
    the CDD (expected) value. Returns the list of files written."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = _stamp()

    # node → group → list of (mo, parameter, expected); de-dup identical
    # (node, mo, parameter) so a param audited from repeated CDD rows yields
    # a single set line.
    by_node: Dict[str, Dict[str, list]] = {}
    seen = set()
    for r in results:
        if r.status not in statuses or r.expected == "":
            continue
        node = r.node or r.key or site
        sig = (node, r.mo, r.parameter)
        if sig in seen:
            continue
        seen.add(sig)
        by_node.setdefault(node, {}).setdefault(_mo_group(r.mo), []).append(
            (r.mo, r.parameter, r.expected))

    written: List[str] = []
    for node, groups in by_node.items():
        lines = [
            "# " + "-" * 60,
            f"# Generate by: {generated_by or 'NodeCraft'}",
            f"# Datetime: {stamp}",
            f"# Audit File: {os.path.basename(audit_xlsx)}",
            "# Sheet Name: Detail",
            "# " + "-" * 60,
            "",
            f"l mkdir ~/LOGS/{site}",
            '$timeCheck = `date "+%y%m%d_%H%M%S"`',
            f"l+ ~/LOGS/{site}/{node}_SetParameter_$timeCheck.log",
            "",
            "lt all",
            "gs+",
            "alt",
            "",
        ]
        ordered = [g for g in _MO_GROUP_ORDER if g in groups]
        ordered += sorted(g for g in groups if g not in _MO_GROUP_ORDER)
        for g in ordered:
            lines.append("")
            lines.append(_banner(g))
            # Sort by parameter (then MO) so all lines of one parameter are
            # contiguous — easy to review or delete a whole parameter at once.
            for mo, param, val in sorted(groups[g], key=lambda x: (x[1], x[0])):
                lines.append(f"set {mo}$ {param} {val}")
        # Close the log opened with l+ at the top.
        lines.append("")
        lines.append("l-")
        path = os.path.join(out_dir, f"{node}_SetParameter_{stamp}.mos")
        with open(path, "w", encoding="utf-8", newline="\n") as fh:
            fh.write("\n".join(lines) + "\n")
        written.append(path)
    return written


# ── ENM cmedit / cmbulk SetParameter script generation ──────────────
# Ported from C:\dev\enp-generator (services/mo_script_generator.py) — the
# CMEdit CLI and CM Bulk CLI ``set`` formats. These are written to SEPARATE
# files from the moshell/.mos script (different command syntax) and are NOT
# meant to be run by the in-app Run Scripts button — only the .mos is.


def _collect_set_rows(results: List[AuditResult], statuses, site: str):
    """node → ordered list of (mo, parameter, value, norm, actual) for the rows
    to align, de-duplicated on (node, mo, parameter). Shared by every generator
    so the three formats stay in lock-step. ``norm``/``actual`` let the value be
    re-formatted to the node's convention when a set line is emitted."""
    by_node: Dict[str, list] = {}
    seen = set()
    for r in results:
        if r.status not in statuses or r.expected == "":
            continue
        node = r.node or r.key or site
        sig = (node, r.mo, r.parameter)
        if sig in seen:
            continue
        seen.add(sig)
        by_node.setdefault(node, []).append(
            (r.mo, r.parameter, r.expected, r.norm, r.actual))
    return by_node


# Boolean/enum domains: a CDD ``1``/``0`` maps to the ON/OFF member of the same
# domain the node reports for that attribute (detected from the actual value).
_BOOL_DOMAINS = [("ON", "OFF"), ("ACTIVE", "INACTIVE"), ("ENABLED", "DISABLED"),
                 ("TRUE", "FALSE"), ("YES", "NO"), ("UNLOCKED", "LOCKED")]


def _dms_cdd_to_node(v: str) -> str:
    """CDD DMS coordinate (``8°3'35.1\"N``) → node form (``N08-03-35.1``):
    hemisphere first, degrees/minutes/seconds hyphen-joined and zero-padded."""
    m = _LATLON_A.search(str(v))
    if not m:
        return str(v)
    d, mn, sec, hemi = m.group(1), m.group(2), m.group(3), m.group(4).upper()
    if "." in sec:
        ip, fp = sec.split(".", 1)
        secf = f"{int(ip):02d}.{fp}"
    else:
        secf = f"{int(sec):02d}"
    return f"{hemi}{int(d):02d}-{int(mn):02d}-{secf}"


def _format_set_value(value, norm: str = "", actual: str = "") -> str:
    """Format a CDD value into the node's own convention for a ``set`` — value
    from the CDD, *format* from the node:

      * ``list``    ``0 1 2 3`` / ``4&12&17``  → ``[0, 1, 2, 3]`` / ``[4, 12, 17]``
      * ``latlong`` ``8°3'35.1\"N``            → ``N08-03-35.1``
      * ``geo``     ``6.999`` (decimal deg)    → ``6999000`` (µdeg integer)
      * boolean     ``1`` / ``0``              → ``ACTIVE`` / ``INACTIVE`` … per
                                                 the domain the node reports.
    """
    s = "" if value is None else str(value).strip()
    if s == "":
        return s
    if norm == "list":
        parts = [p for p in re.split(r"[\[\],&\s]+", s) if p]
        # Only bracketise a genuine multi-value list; a lone token (e.g. a
        # ChannelGroup dchNo of "OFF", or a single value) is left as-is.
        return "[" + ", ".join(parts) + "]" if len(parts) > 1 else s
    if norm == "latlong":
        return _dms_cdd_to_node(s)
    if norm == "geo":
        try:
            return str(int(round(float(s) * 1e6)))
        except (ValueError, TypeError):
            return s
    if s in ("0", "1"):
        a = str(actual or "").strip().upper()
        for on, off in _BOOL_DOMAINS:
            if a in (on, off):
                return on if s == "1" else off
    if s.lower() in ("true", "false"):
        return s.lower()
    return s


def _enm_prefix(fdn_prefix: str = "") -> str:
    """Normalize the ENM SubNetwork prefix.

    The main form stores the short subnetwork value (for example ``T7``), while
    ENM set commands need the full ONRM-rooted FDN prefix. Already-expanded
    prefixes are preserved for callers/tests that pass a full FDN prefix.
    """
    prefix = (fdn_prefix or "").strip().strip(",")
    if not prefix:
        return ""
    if "=" in prefix:
        return prefix.rstrip(",")
    return f"SubNetwork=ONRM_ROOT_MO_R,SubNetwork={prefix}"


def _mo_below_managed_element(mo: str) -> str:
    """Return the FDN path below ManagedElement, even if ``mo`` is rooted."""
    parts = [p.strip() for p in (mo or "").strip().strip(",").split(",")
             if p.strip()]
    for i in range(len(parts) - 1, -1, -1):
        name = parts[i].split("=", 1)[0].strip().lower()
        if name == "managedelement":
            return ",".join(parts[i + 1:])
    return ",".join(parts)


def _enm_fdn(node: str, mo: str, fdn_prefix: str = "") -> str:
    """Full ENM FDN for a ``cmedit`` target:
    ``[<prefix>,]MeContext=<node>,ManagedElement=<node>[,<mo>]``. ``mo`` is
    normally below ManagedElement, but rooted values from parsed references are
    tolerated and normalized before the final FDN is assembled."""
    rel_mo = _mo_below_managed_element(mo)
    base = f"MeContext={node},ManagedElement={node}"
    if rel_mo:
        base = f"{base},{rel_mo}"
    prefix = _enm_prefix(fdn_prefix)
    return f"{prefix},{base}" if prefix else base


# GSM profiles whose ``cmedit_mo`` is NOT a child MO under GeranCell: the base
# cell (attrs live directly on GeranCell) and Trx (on the RadioNode, not the BSC).
_GSM_NON_CHILD = {"gerancell", "trx"}


def build_gsm_child_map(audit_map: dict) -> Dict[str, str]:
    """parameter (lower) → GSM child-MO class, from the audit map's GSM
    profiles. GSM child-MO attributes are merged onto the GeranCell record for
    the compare, which loses the child-MO identity — this recovers it so the
    cmedit/cmbulk FDN can append the child segment (``…,IdleModeAndPaging=1``)."""
    m: Dict[str, str] = {}
    for p in audit_map.get("profiles", []):
        if p.get("tech") != "gsm":
            continue
        child = (p.get("cmedit_mo") or "").strip()
        if not child or child.lower() in _GSM_NON_CHILD:
            continue
        for c in p.get("columns", []):
            a = (c.get("attr") or "").strip()
            if a:
                m.setdefault(a.lower(), child)
    return m


def _build_fdn(node: str, mo: str, param: str = "", fdn_prefix: str = "",
               gsm_fdn_prefix: str = "",
               gsm_child_map: Optional[Dict[str, str]] = None) -> str:
    """Pick the right FDN shape for a target MO. GSM ``GeranCell`` MOs live on
    the BSC (not the audited RadioNode) with a fixed BscFunction/GeranCellM
    lineage, so they use the configurable ``gsm_fdn_prefix`` template — and, when
    the parameter belongs to a child MO, the child segment is appended
    (``…,GeranCell=<id>,<Child>=1``). Every other MO uses the node-rooted
    ``MeContext=<node>,ManagedElement=<node>`` FDN built from the dump path."""
    top = mo.split("=", 1)[0].strip()
    if gsm_fdn_prefix and top == "GeranCell":
        # Append the child segment only for a bare ``GeranCell=<id>``; when the
        # mo already carries a child (e.g. ChannelGroup=0 from a per-index CDD
        # column) trust that real instance instead of forcing ``=1``.
        leaf = mo
        if "," not in mo:
            child = (gsm_child_map or {}).get((param or "").lower())
            if child:
                leaf = f"{mo},{child}=1"
        return f"{gsm_fdn_prefix.rstrip(',')},{leaf}"
    return _enm_fdn(node, mo, fdn_prefix)


def _enm_value(v) -> str:
    """Format a value for ENM set: booleans lower-cased, everything else passed
    through (structs already rendered as ``{a=..,b=..}`` are kept verbatim)."""
    s = "" if v is None else str(v).strip()
    if s.lower() in ("true", "false"):
        return s.lower()
    return s


def _enm_header(fmt: str, generated_by: str, stamp: str, audit_xlsx: str):
    return [
        "# " + "-" * 60,
        f"# Generate by: {generated_by or 'NodeCraft'}",
        f"# Format: {fmt}",
        f"# Datetime: {stamp}",
        f"# Audit File: {os.path.basename(audit_xlsx)}",
        "# Sheet Name: Detail",
        "# NOTE: review before applying - not run by the in-app Run button.",
        "# NOTE: singleton GSM child MOs use index =1 (ChannelGroup uses its"
        " real index).",
        "# " + "-" * 60,
        "",
    ]


def generate_cmedit_scripts(results: List[AuditResult], out_dir: str,
                            site: str, audit_xlsx: str, generated_by: str = "",
                            statuses=("Mismatch",),
                            fdn_prefix: str = "",
                            gsm_fdn_prefix: str = "",
                            gsm_child_map: Optional[Dict[str, str]] = None
                            ) -> List[str]:
    """One CMEdit CLI file per node: ``cmedit set <FDN> <param>=<value>``
    (one command per parameter). Returns the files written."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = _stamp()
    written: List[str] = []
    for node, rows in _collect_set_rows(results, statuses, site).items():
        lines = _enm_header("CMEDIT", generated_by, stamp, audit_xlsx)
        for mo, param, val, norm, actual in sorted(rows, key=lambda x: (x[1], x[0])):
            fdn = _build_fdn(node, mo, param, fdn_prefix, gsm_fdn_prefix,
                             gsm_child_map)
            lines.append(
                f"cmedit set {fdn} {param}={_format_set_value(val, norm, actual)}")
        path = os.path.join(out_dir, f"{node}_SetParameter_{stamp}_cmedit.txt")
        with open(path, "w", encoding="utf-8", newline="\n") as fh:
            fh.write("\n".join(lines) + "\n")
        written.append(path)
    return written


def build_cmedit_commands(results: List[AuditResult], site: str,
                          statuses=("Mismatch",), fdn_prefix: str = "",
                          gsm_fdn_prefix: str = "",
                          gsm_child_map: Optional[Dict[str, str]] = None
                          ) -> List[tuple]:
    """The exact ``cmedit set <FDN> <param>=<value>`` commands for the rows to
    align — same FDN/value logic as the CMEdit file export, but returned as
    ``(node, command)`` tuples so the app can apply them live over SSH."""
    cmds: List[tuple] = []
    for node, rows in _collect_set_rows(results, statuses, site).items():
        for mo, param, val, norm, actual in sorted(rows, key=lambda x: (x[1], x[0])):
            fdn = _build_fdn(node, mo, param, fdn_prefix, gsm_fdn_prefix,
                             gsm_child_map)
            cmds.append(
                (node, f"cmedit set {fdn} {param}={_format_set_value(val, norm, actual)}"))
    return cmds


def generate_cmbulk_scripts(results: List[AuditResult], out_dir: str,
                            site: str, audit_xlsx: str, generated_by: str = "",
                            statuses=("Mismatch",),
                            fdn_prefix: str = "",
                            gsm_fdn_prefix: str = "",
                            gsm_child_map: Optional[Dict[str, str]] = None
                            ) -> List[str]:
    """One CM Bulk CLI file per node, parameters grouped per FDN into one block::

        set
        FDN : <FDN>
        param1 : value1
        param2 : value2

    Grouping is by the full FDN, so GSM child MOs (each with its own
    ``…,<Child>=1`` FDN) get their own block rather than being folded under the
    GeranCell."""
    os.makedirs(out_dir, exist_ok=True)
    stamp = _stamp()
    written: List[str] = []
    for node, rows in _collect_set_rows(results, statuses, site).items():
        # group by the computed FDN, preserving first-seen order
        fdn_params: Dict[str, list] = {}
        for mo, param, val, norm, actual in rows:
            fdn = _build_fdn(node, mo, param, fdn_prefix, gsm_fdn_prefix,
                             gsm_child_map)
            fdn_params.setdefault(fdn, []).append(
                (param, _format_set_value(val, norm, actual)))
        lines = _enm_header("CMBULK", generated_by, stamp, audit_xlsx)
        for fdn, params in fdn_params.items():
            lines.append("set")
            lines.append(f"FDN : {fdn}")
            for param, v in sorted(params, key=lambda x: x[0]):
                if v.startswith("{") and v.endswith("}"):
                    lines.append(f"{param}={v}")
                else:
                    lines.append(f"{param} : {v}")
            lines.append("")
        path = os.path.join(out_dir, f"{node}_SetParameter_{stamp}_cmbulk.txt")
        with open(path, "w", encoding="utf-8", newline="\n") as fh:
            fh.write("\n".join(lines) + "\n")
        written.append(path)
    return written


def _stamp() -> str:
    import datetime as _dt
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
