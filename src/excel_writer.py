"""
Copy Excel template and embed screenshot images into the appropriate sheets.
"""
import os
import shutil
import logging
from PIL import Image as PILImage
PILImage.MAX_IMAGE_PIXELS = None  # Allow large screenshots
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from typing import Dict, List, Optional

from config_loader import AppConfig, get_full_path

logger = logging.getLogger(__name__)

# Maps category names to Excel sheet names
CATEGORY_TO_SHEET = {
    "VSWR": "VSWR",
    "CELL": "CELL",
    "NOC_Logs": "NOC_Logs",
    "ALARM": "ALARM",
    "PIM": "PIM",
    "RET": "RET",
}

# ENM items get embedded in the same sheet as their parent category
ENM_PARENT_CATEGORY = {
    "CELL_ENM": "CELL",
    "NOC_Logs_ENM": "NOC_Logs",
    "ALARM_ENM": "ALARM",
}


def create_band_excel(
    config: AppConfig,
    band: str,
) -> str:
    """
    Copy the template Excel file for a specific band.

    Args:
        config: Application configuration
        band: Band key (e.g., "L900", "NR700")

    Returns:
        Path to the new Excel file
    """
    template_path = get_full_path(config, config.paths.template)
    output_dir = get_full_path(config, config.paths.output_dir)

    filename = f"{config.site.shortcode}_{band}_POST_BFT_SRS_Rev1.xlsx"
    output_path = os.path.join(output_dir, filename)

    # If file is locked (open in Excel), add timestamp suffix
    try:
        shutil.copy2(template_path, output_path)
    except PermissionError:
        from datetime import datetime
        ts = datetime.now().strftime("%H%M%S")
        filename = f"{config.site.shortcode}_{band}_POST_BFT_SRS_Rev1_{ts}.xlsx"
        output_path = os.path.join(output_dir, filename)
        shutil.copy2(template_path, output_path)
        logger.warning(f"Original file was locked, using: {filename}")

    logger.info(f"Created Excel file: {output_path}")
    return output_path


def insert_screenshot(
    excel_path: str,
    sheet_name: str,
    image_path: str,
    cell: str = "A2",
    image_width: Optional[int] = None,
    image_height: Optional[int] = None,
) -> None:
    """
    Insert a screenshot image into a specific sheet in the Excel file.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        image_path: Path to the screenshot PNG file
        cell: Cell position to anchor the image (default A2, below header)
        image_width: Optional image width override in pixels
        image_height: Optional image height override in pixels
    """
    wb = load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in {excel_path}, creating it")
        wb.create_sheet(sheet_name)

    ws = wb[sheet_name]

    img = XlImage(image_path)

    # Scale image to fit reasonably in Excel
    if image_width:
        img.width = image_width
    if image_height:
        img.height = image_height

    ws.add_image(img, cell)
    wb.save(excel_path)
    logger.info(f"Inserted screenshot into {sheet_name} at {cell}: {image_path}")


def insert_screenshots_for_band(
    excel_path: str,
    screenshots: Dict[str, List[str]],
) -> None:
    """
    Insert all screenshots for a band into the appropriate sheets.

    SSH (moshell) screenshots are placed first, then ENM screenshots are
    appended below them on the same sheet so the two never overlap.

    Args:
        excel_path: Path to the Excel file
        screenshots: Dict mapping category to list of screenshot paths
    """
    wb = load_workbook(excel_path)

    # Sort categories so SSH (parent) entries are processed before their
    # matching _ENM entries on the same sheet. Unknown categories are skipped.
    def sort_key(cat: str):
        is_enm = 1 if cat.endswith("_ENM") else 0
        return (is_enm, cat)

    # Per-sheet cursor: tracks the next free row on each sheet across
    # categories. openpyxl images are floating and do NOT advance
    # ws.max_row, so we must track this ourselves.
    next_row_by_sheet: Dict[str, int] = {}

    for category in sorted(screenshots.keys(), key=sort_key):
        image_paths = screenshots[category]

        if category in ENM_PARENT_CATEGORY:
            sheet_name = ENM_PARENT_CATEGORY[category]
        elif category in CATEGORY_TO_SHEET:
            sheet_name = CATEGORY_TO_SHEET[category]
        else:
            logger.warning(f"Unknown category '{category}', skipping")
            continue

        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found, creating it")
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # First time touching this sheet — start just below the template header
        if sheet_name not in next_row_by_sheet:
            next_row_by_sheet[sheet_name] = max(ws.max_row + 2, 2)

        start_row = next_row_by_sheet[sheet_name]

        for img_path in image_paths:
            if not os.path.exists(img_path):
                logger.warning(f"Screenshot not found: {img_path}")
                continue

            img = XlImage(img_path)

            # Label row above the image
            label_row = start_row
            image_row = start_row + 1
            if category.endswith("_ENM"):
                ws[f"A{label_row}"] = f"[ENM] {category.replace('_ENM', '')}"
            else:
                ws[f"A{label_row}"] = category

            ws.add_image(img, f"A{image_row}")

            # Excel default row height is ~20px. Add a generous buffer so
            # consecutive images never collide.
            px_height = img.height or 600
            rows_for_image = max(8, int(px_height / 18) + 3)
            start_row = image_row + rows_for_image + 2

        next_row_by_sheet[sheet_name] = start_row

    wb.save(excel_path)
    logger.info(f"All screenshots inserted into {excel_path}")
