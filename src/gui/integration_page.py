"""
Integration workflow — step selection checklist + multi-column progress for LTE/NR, LTE/NR #2, and GSM nodes.

Two phases:
  /integration      — Step selector (checkboxes, Select/Deselect All, Run button)
  /integration_run  — Progress page (2-3 node columns, high-level log, timer)
"""
import asyncio
import collections
import logging
import os
import queue
import re
import sys
import threading
import time
from datetime import datetime
from typing import Optional

import flet as ft

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

from baseline_log_parser import build_baseline_log_excel, parse_baseline_summary
from relation_log_parser import build_relation_log_excel

from gui.theme import (
    ACCENT,
    ACCENT_WARM,
    BG_BOTTOM,
    BG_TOP,
    BORDER,
    DANGER,
    INFO,
    PANEL,
    PANEL_RAISED,
    SUCCESS,
    TEXT,
    TEXT_MUTED,
    background_gradient,
    badge,
    panel,
    secondary_button_style,
)

logger = logging.getLogger(__name__)

# ── Step definitions ─────────────────────────────────────────────
# Each tuple: (key, display_label, applies_to, log_suffix)
#   applies_to: "both" | "lte_nr" | "gsm"
#   log_suffix: short name used in the log filename
# All steps that appear in the summary table (including remark-only rows).
#
# applies_to values:
#   "both"        — runs on every node type
#   "lte_nr"      — runs on every LTE/NR node (primary + secondary)
#   "lte_primary" — runs on the primary LTE/NR node ONLY (skips lte2 + gsm).
#                   Used for one-shot per-site setup that doesn't need to
#                   be repeated for the second BB / GSM node — e.g. the
#                   External Alarm template install.
#   "gsm"         — runs on GSM node only
INTEGRATION_STEPS = [
    ("sw_package_check",  "SW Package Check",              "both",    "SW_PACKAGE"),
    ("bb_power_on",       "BB Power On Onsite",            "both",    "BB_POWER_ON"),
    ("bb_transport",      "BB Transport Connectivity",     "both",    "BB_TRANSPORT"),
    ("create_arne",       "Add Node in ENM",              "both",    "ARNE"),
    ("enrollment",       "Perform Site Enrollment in ENM", "both",    "ENROLLMENT"),
    ("uri_setting",       "URI Reconfig",                 "both",    "URI"),
    ("sw_level_check",    "SW Level Check",                "both",    "SW_LEVEL"),
    ("enrollment_sync",   "Node Synchronized in ENM",    "both",    "ENROLLMENT"),
    ("install_lkf",       "Install LKF (License File)",    "both",   "LKF"),
    ("relation",          "Load Neighbour Relation Scripts", "lte_nr", "RELATION"),
    ("baseline",          "Load Baseline Scripts",         "lte_nr",  "BASELINE"),
    ("ret_scripts",       "Load RET Scripts",              "both",    "RET"),
    ("pm_measurement",    "Validate Performance Counter",  "both",    "PM"),
    ("external_alarm",    "Pre-define External Alarm",     "lte_primary", "EXTERNAL_ALARM"),
    ("verify_mme",        "Verify Core Connectivity (MME)", "lte_nr", "MME"),
    ("sgw_check",         "Verify SGw Reachability",      "both",    "SGW_CHECK"),
    ("gsm_cell_define",  "GSM Cells and MO Defined in BSC", "gsm",   "GSM_CELL_DEFINE"),
    ("bsc_neighbours",    "BSC Neighbours Defined",        "gsm",     "BSC_NEIGHBOURS"),
    ("network_audit",     "Network Configuration Audit (NAT)", "both", "NAT"),
    ("sync_check",        "Synchronization",               "both",    "SYNC"),
    ("backup_cv",         "Configuration Backup and Upload to ENM", "both", "BACKUP_CV"),
    ("take_cm_dump",      "Take CM Dump",                  "both",    "CM_DUMP"),
    ("take_dump",         "Take Dump",                     "both",    "DUMP"),
]

# Steps that only appear in the summary table (not as checkboxes, not run).
# Their status is derived from alias steps or hardcoded.
REMARK_STEPS = {
    "sw_package_check",   # pre-verified manual check → always Yes
    "bb_power_on",        # pre-verified manual check → always Yes
    "enrollment_sync",
    "bb_transport",
    "network_audit",
}

# Summary-only rows that should stay out of the live progress columns and
# always display as N/A in the final summary.
SUMMARY_NA_STEPS = {
    "network_audit",
}


SUMMARY_LABELS = {
    "sw_package_check":  "SW Package Check",
    "bb_power_on":       "BB Power On Onsite",
    "bb_transport":      "BB Transport Connectivity",
    "create_arne":       "Add Node in ENM",
    "enrollment":        "Perform Site Enrollment in ENM",
    "uri_setting":       "URI Reconfig",
    "sw_level_check":    "SW Level Check",
    "enrollment_sync":   "Node Synchronized in ENM",
    "install_lkf":       "Install LKF (License File)",
    "relation":          "Load Neighbour Relation Scripts",
    "baseline":          "Load Baseline Scripts",
    "ret_scripts":       "Load RET Scripts",
    "external_alarm":    "Pre-define External Alarm",
    "verify_mme":        "Verify Core Connectivity (MME)",
    "sgw_check":         "Verify SGw Reachability",
    "gsm_cell_define":   "GSM Cells and MO Defined in BSC",
    "bsc_neighbours":    "BSC Neighbours Defined",
    "network_audit":     "Network Configuration Audit (NAT)",
    "sync_check":        "Synchronization",
    "pm_measurement":    "Validate Performance Counter",
    "backup_cv":         "Configuration Backup and Upload to ENM",
    "take_cm_dump":      "Take CM Dump",
    "take_dump":         "Take Dump",
}

def _step_applies_to_node(
    applies_to: str,
    ntag: str,
    gsm_on_primary: bool = False,
) -> bool:
    """Whether a step with the given ``applies_to`` scope is relevant
    to the node identified by ``ntag`` (``"lte"`` / ``"lte2"`` / ``"gsm"``).

    Module-level helper so the checklist builder, the progress builder,
    the run loop, and the summary table all use one consistent rule.

    ``applies_to`` values:
      ``both``        → every node
      ``lte_nr``      → ``lte`` or ``lte2``
      ``lte_primary`` → ``lte`` only (per-site setup that runs once)
      ``gsm``         → ``gsm`` node; with ``gsm_on_primary=True`` it
                        also matches ``lte`` (single-RAT-radio site
                        where the operator typed a BSC name but no
                        separate GSM node — GSM is co-located on the
                        primary LTE/NR node).
    """
    if applies_to == "both":
        return True
    if applies_to == "lte_nr":
        return ntag in ("lte", "lte2")
    if applies_to == "lte_primary":
        return ntag == "lte"
    if applies_to == "gsm":
        if ntag == "gsm":
            return True
        if gsm_on_primary and ntag == "lte":
            return True
        return False
    return False


_RESULT_SYMBOLS = {
    "done":    "✅",
    "warn":    "⚠️",
    "error":   "❌",
    "skip":    "⏩",
    "running": "🔄",
}


# ═════════════════════════════════════════════════════════════════
#  Phase 1 — Step Selection Checklist
# ═════════════════════════════════════════════════════════════════
class IntegrationPage:
    """Checklist page where user picks which integration steps to run."""

    def __init__(self, page: ft.Page):
        self.page = page
        self.form = getattr(page, "integration_form", {})

        self.lte_name = self.form.get("node_name", "LTE/NR Node")
        self.lte_ip = self.form.get("node_ip", "")
        self.lte_subnet = self.form.get("subnetwork", "")
        self.lte2_name = self.form.get("node2_name", "")
        self.lte2_ip = self.form.get("node2_ip", "")
        self.lte2_subnet = self.form.get("node2_subnetwork", "")
        self.gsm_name = self.form.get("gsm_node_name", "")
        self.gsm_ip = self.form.get("gsm_node_ip", "")
        self.gsm_subnet = self.form.get("gsm_subnetwork", "")
        self.has_lte2 = bool(self.lte2_name)
        self.has_gsm = bool(self.gsm_name)
        # Co-located mode: a single physical node hosts BOTH LTE/NR
        # AND GSM (multi-RAT, no separate GSM DN). Operator signals
        # this by entering a BSC name but leaving the GSM node name +
        # IP blank. All GSM-scope steps then run on the primary LTE
        # node instead of getting their own column.
        self.gsm_on_primary = (
            not self.has_gsm
            and bool(str(self.form.get("bsc_name", "") or "").strip())
        )
        self.shortcode = self.form.get("shortcode", "UNKNOWN")

        # Checkboxes are keyed step → node_tag → Checkbox so each node
        # can independently include/exclude steps. Old shape (step → cb)
        # is preserved as a flat union in ``self.flat_checkboxes`` for
        # the global Select All / counter logic.
        self.checkboxes: dict[str, dict[str, ft.Checkbox]] = {}

        # Determine which node columns to show, with short headers.
        # A single LTE → "LTE"; two LTEs → "LTE 1", "LTE 2"; plus GSM.
        self.node_columns: list[tuple[str, str]] = []
        if self.has_lte2:
            self.node_columns.append(("lte",  "LTE 1"))
            self.node_columns.append(("lte2", "LTE 2"))
        else:
            self.node_columns.append(("lte", "LTE"))
        if self.has_gsm:
            self.node_columns.append(("gsm", "GSM"))

    def _step_applies_to_node(self, applies_to: str, ntag: str) -> bool:
        # Thin wrapper over the module-level helper — passes the
        # co-located-GSM flag so GSM-scope steps fold into the LTE
        # column when there's no separate GSM node.
        return _step_applies_to_node(
            applies_to, ntag,
            gsm_on_primary=getattr(self, "gsm_on_primary", False),
        )

    def _step_disabled_reason(self, key: str) -> Optional[str]:
        """Return a short reason why a step's checkboxes should be
        force-unchecked + disabled at checklist build time, or None
        if the step is selectable.

        Some steps depend on a browsed file from the form page. When
        that file is empty, there is nothing to install/load, so the
        checklist hides the choice from the operator entirely:
        checkbox starts unchecked, disabled (can't be ticked), and a
        small italic hint explains why.

        NOTE: ``install_lkf`` is intentionally NOT disabled when no zip
        is provided — LKF now always runs (``lkfinstall.py <node>``
        directly, the license may already be imported in ENM). Only
        ``relation`` still requires its file.
        """
        if key == "relation":
            if not str(self.form.get("relation_file", "")).strip():
                return "no relation file selected"
        return None

    def build(self) -> ft.View:
        # Tag the OS window with the SHORTCODE so multiple open
        # NodeCraft windows are distinguishable from the taskbar.
        try:
            self.page.title = f"NodeCraft — {self.shortcode} (Setup)"
            self.page.update()
        except Exception:
            pass

        # ── Build the multi-column checklist grid ───────────────
        # Layout: one row per step, one checkbox column per node.
        # Cells where the step doesn't apply to the node show "—".
        CHECKBOX_COL_WIDTH = 70

        def _column_color(ntag: str):
            if ntag == "gsm":
                return ACCENT_WARM
            if ntag == "lte2":
                return ACCENT
            return INFO

        # Header row: "Step" label + one column per node
        header_cells = [
            ft.Container(
                content=ft.Text(
                    "Step", size=12, color=TEXT_MUTED,
                    weight=ft.FontWeight.BOLD,
                ),
                expand=True,
            ),
        ]
        for ntag, nlabel in self.node_columns:
            header_cells.append(
                ft.Container(
                    content=ft.Text(
                        nlabel, size=12, color=_column_color(ntag),
                        weight=ft.FontWeight.BOLD,
                    ),
                    width=CHECKBOX_COL_WIDTH,
                    alignment=ft.Alignment(0, 0),
                )
            )
        header_row = ft.Container(
            padding=ft.Padding.symmetric(horizontal=12, vertical=8),
            content=ft.Row(
                header_cells, spacing=10,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        )

        check_rows: list = [header_row, ft.Divider(height=1, color=BORDER)]

        for key, label, applies_to, _ in INTEGRATION_STEPS:
            # Skip remark-only rows: they appear in the summary table
            # but have no runnable logic, so no checkbox needed.
            if key in REMARK_STEPS or key in SUMMARY_NA_STEPS:
                continue

            # Skip steps that apply to no visible column
            if not any(
                self._step_applies_to_node(applies_to, nt)
                for nt, _ in self.node_columns
            ):
                continue

            disabled_reason = self._step_disabled_reason(key)

            # Step label cell — when the step is force-disabled because
            # of a missing file, the label is dimmed and a small italic
            # hint is shown underneath so the operator understands why
            # the checkboxes don't respond.
            if disabled_reason:
                label_widget = ft.Column(
                    [
                        ft.Text(
                            label, size=14, color=TEXT_MUTED,
                            weight=ft.FontWeight.W_500,
                        ),
                        ft.Text(
                            f"({disabled_reason})", size=11,
                            color=TEXT_MUTED, italic=True,
                        ),
                    ],
                    spacing=2,
                )
            else:
                label_widget = ft.Text(
                    label, size=14, color=TEXT,
                    weight=ft.FontWeight.W_500,
                )

            cells = [
                ft.Container(content=label_widget, expand=True),
            ]
            self.checkboxes[key] = {}

            for ntag, _ in self.node_columns:
                if self._step_applies_to_node(applies_to, ntag):
                    cb = ft.Checkbox(
                        # Force-uncheck + disable when a required file
                        # is missing; otherwise default-checked.
                        value=False if disabled_reason else True,
                        disabled=bool(disabled_reason),
                        label=None,
                        active_color=ACCENT,
                        check_color="#06242A",
                        on_change=self._on_checkbox_change,
                    )
                    self.checkboxes[key][ntag] = cb
                    cells.append(
                        ft.Container(
                            content=cb,
                            width=CHECKBOX_COL_WIDTH,
                            alignment=ft.Alignment(0, 0),
                        )
                    )
                else:
                    cells.append(
                        ft.Container(
                            content=ft.Text(
                                "—", size=14, color=TEXT_MUTED,
                            ),
                            width=CHECKBOX_COL_WIDTH,
                            alignment=ft.Alignment(0, 0),
                        )
                    )

            row = ft.Container(
                padding=ft.Padding.symmetric(horizontal=12, vertical=4),
                border_radius=10,
                content=ft.Row(
                    cells, spacing=10,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
            )
            check_rows.append(row)

        # ── Counter text ────────────────────────────────────────
        total_cbs = sum(len(d) for d in self.checkboxes.values())
        self.counter_text = ft.Text(
            f"{total_cbs}/{total_cbs} steps selected",
            size=12, color=TEXT_MUTED,
        )

        # ── Select All / Deselect All buttons ───────────────────
        select_all_btn = ft.TextButton(
            "Select All",
            icon=ft.Icons.CHECK_BOX,
            style=ft.ButtonStyle(color=ACCENT),
            on_click=self._select_all,
        )
        deselect_all_btn = ft.TextButton(
            "Deselect All",
            icon=ft.Icons.CHECK_BOX_OUTLINE_BLANK,
            style=ft.ButtonStyle(color=TEXT_MUTED),
            on_click=self._deselect_all,
        )

        # ── Node info summary ───────────────────────────────────
        node_info_parts = [
            ft.Icon(ft.Icons.CELL_TOWER, size=16, color=INFO),
            ft.Text(f"{self.lte_name}", size=13, color=TEXT,
                    weight=ft.FontWeight.W_500),
        ]
        if self.lte_ip:
            node_info_parts.append(
                ft.Text(f"({self.lte_ip})", size=12, color=TEXT_MUTED))
        if self.has_lte2:
            node_info_parts.extend([
                ft.Container(width=16),
                ft.Icon(ft.Icons.CELL_TOWER, size=16, color=ACCENT),
                ft.Text(f"{self.lte2_name}", size=13, color=TEXT,
                        weight=ft.FontWeight.W_500),
            ])
            if self.lte2_ip:
                node_info_parts.append(
                    ft.Text(f"({self.lte2_ip})", size=12, color=TEXT_MUTED))
        if self.has_gsm:
            node_info_parts.extend([
                ft.Container(width=16),
                ft.Icon(ft.Icons.SETTINGS_INPUT_ANTENNA, size=16,
                        color=ACCENT_WARM),
                ft.Text(f"{self.gsm_name}", size=13, color=TEXT,
                        weight=ft.FontWeight.W_500),
            ])
            if self.gsm_ip:
                node_info_parts.append(
                    ft.Text(f"({self.gsm_ip})", size=12, color=TEXT_MUTED))

        node_info_row = ft.Row(node_info_parts, spacing=6,
                               vertical_alignment=ft.CrossAxisAlignment.CENTER)

        # ── Run button ──────────────────────────────────────────
        self.run_button = ft.ElevatedButton(
            "Run Integration",
            icon=ft.Icons.PLAY_ARROW_ROUNDED,
            style=ft.ButtonStyle(
                bgcolor=ACCENT, color="#06242A",
                padding=ft.Padding.symmetric(horizontal=28, vertical=16),
                shape=ft.RoundedRectangleBorder(radius=14),
                text_style=ft.TextStyle(
                    size=16, weight=ft.FontWeight.BOLD),
            ),
            on_click=self._on_run,
        )
        back_button = ft.ElevatedButton(
            "Back",
            icon=ft.Icons.ARROW_BACK,
            style=secondary_button_style(),
            on_click=self._go_back,
        )

        # ── Layout ──────────────────────────────────────────────
        checklist_panel = panel(
            ft.Column(
                [
                    ft.Row(
                        [
                            ft.Text("Select Steps to Run", size=18,
                                    weight=ft.FontWeight.BOLD, color=TEXT),
                            ft.Container(expand=True),
                            self.counter_text,
                        ],
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                    ft.Row(
                        [select_all_btn, deselect_all_btn],
                        spacing=4,
                    ),
                    ft.Divider(height=1, color=BORDER),
                    *check_rows,
                ],
                spacing=6,
                scroll=ft.ScrollMode.AUTO,
            ),
            bgcolor=PANEL,
            padding=24,
        )

        body = ft.Container(
            expand=True,
            gradient=background_gradient(),
            padding=ft.Padding.symmetric(horizontal=28, vertical=20),
            content=ft.Column(
                [
                    ft.Text("Integration Setup", size=24,
                            weight=ft.FontWeight.BOLD, color=TEXT),
                    ft.Text(f"Site ID: {self.shortcode}", size=13,
                            color=TEXT_MUTED),
                    node_info_row,
                    ft.Container(height=8),
                    ft.Container(
                        content=checklist_panel,
                        expand=True,
                    ),
                    ft.Container(height=8),
                    ft.Row(
                        [back_button, ft.Container(expand=True),
                         self.run_button],
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                ],
                spacing=8,
                expand=True,
            ),
        )

        return ft.View(
            route="/integration", padding=0, spacing=0,
            bgcolor=BG_TOP, controls=[body],
        )

    # ── Checkbox helpers ────────────────────────────────────────
    def _iter_checkboxes(self):
        for per_node in self.checkboxes.values():
            for cb in per_node.values():
                yield cb

    def _update_counter(self):
        total = 0
        selected = 0
        for cb in self._iter_checkboxes():
            total += 1
            if cb.value:
                selected += 1
        self.counter_text.value = f"{selected}/{total} steps selected"
        self.run_button.disabled = (selected == 0)
        try:
            self.page.update()
        except Exception:
            pass

    def _on_checkbox_change(self, e):
        self._update_counter()

    def _select_all(self, e):
        # Skip disabled checkboxes: they're disabled because their
        # required file isn't selected, so toggling them on would be
        # misleading — the runtime guard would still skip the step.
        for cb in self._iter_checkboxes():
            if not cb.disabled:
                cb.value = True
        self._update_counter()

    def _deselect_all(self, e):
        for cb in self._iter_checkboxes():
            if not cb.disabled:
                cb.value = False
        self._update_counter()

    # ── Navigation ──────────────────────────────────────────────
    def _on_preflight(self, e):
        """Read-only pre-run validation: config script paths on the gateway +
        BSC broker reachability. Runs off-thread so the UI stays responsive."""
        host = self.form.get("host", "")
        user = self.form.get("username", "")
        pwd = self.form.get("password", "")
        if not (host and user and pwd):
            self._alert_simple("Pre-flight Check",
                               "Enter SSH host / username / password on the "
                               "form first.")
            return
        self.preflight_button.disabled = True
        self.preflight_button.text = "Checking…"
        self.page.update()
        threading.Thread(target=self._preflight_worker, daemon=True).start()

    def _preflight_worker(self):
        import config_validator
        from integration_runner import IntegrationSSH, _CFG
        _log = lambda m: print(f"[preflight] {m}")
        results, err = [], ""
        ssh = None
        try:
            port = int(self.form.get("port", 5023) or 5023)
        except (ValueError, TypeError):
            port = 5023
        try:
            ssh = IntegrationSSH(
                host=self.form.get("host", ""), port=port,
                username=self.form.get("username", ""),
                password=self.form.get("password", ""),
                log_callback=_log)
            ssh.connect(timeout=30)
            results = config_validator.run_preflight(
                ssh, _CFG, self.form.get("bsc_name", ""), _log)
        except Exception as exc:
            err = f"{type(exc).__name__}: {exc}"
        finally:
            if ssh is not None:
                try:
                    ssh.disconnect()
                except Exception:
                    pass
        self.preflight_button.disabled = False
        self.preflight_button.text = "Pre-flight Check"
        try:
            if err:
                self._alert_simple("Pre-flight Check failed", err)
            else:
                self._show_preflight_dialog(results)
        except Exception:
            pass
        try:
            self.page.update()
        except Exception:
            pass

    def _show_dialog_safe(self, dlg) -> None:
        try:
            if hasattr(self.page, "show_dialog"):
                self.page.show_dialog(dlg)
                return
        except Exception:
            pass
        try:
            self.page.overlay.append(dlg)
            dlg.open = True
            self.page.update()
        except Exception:
            pass

    def _close_dialog_safe(self, dlg) -> None:
        try:
            if hasattr(self.page, "pop_dialog"):
                self.page.pop_dialog()
                return
        except Exception:
            pass
        try:
            dlg.open = False
            if dlg in (self.page.overlay or []):
                self.page.overlay.remove(dlg)
            self.page.update()
        except Exception:
            pass

    def _alert_simple(self, title: str, message: str) -> None:
        dlg = ft.AlertDialog(
            modal=True, title=ft.Text(title),
            content=ft.Text(message),
            actions=[ft.TextButton(
                "OK", on_click=lambda e: self._close_dialog_safe(dlg))],
        )
        self._show_dialog_safe(dlg)

    def _show_preflight_dialog(self, results: list) -> None:
        colour = {"pass": SUCCESS, "warn": ACCENT_WARM,
                  "fail": DANGER, "skip": TEXT_MUTED}
        icon = {"pass": ft.Icons.CHECK_CIRCLE, "warn": ft.Icons.WARNING_AMBER,
                "fail": ft.Icons.CANCEL, "skip": ft.Icons.REMOVE_CIRCLE_OUTLINE}
        fails = sum(1 for r in results if r.status == "fail")
        warns = sum(1 for r in results if r.status == "warn")
        rows = []
        for r in results:
            c = colour.get(r.status, TEXT_MUTED)
            rows.append(ft.Row([
                ft.Icon(icon.get(r.status, ft.Icons.HELP), color=c, size=18),
                ft.Text(r.name, width=170, color=TEXT,
                        weight=ft.FontWeight.W_600),
                ft.Text(r.detail, expand=True, color=TEXT_MUTED, size=12,
                        selectable=True),
            ], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER))
        head = (f"{fails} blocker(s), {warns} warning(s)" if fails or warns
                else "All checks passed ✓")
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Pre-flight Check — {head}",
                          color=(DANGER if fails else
                                 ACCENT_WARM if warns else SUCCESS),
                          weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=760,
                content=ft.Column(rows, spacing=6, scroll=ft.ScrollMode.AUTO,
                                  tight=True)),
            actions=[ft.TextButton(
                "Close", on_click=lambda e: self._close_dialog_safe(dlg))],
        )
        self._show_dialog_safe(dlg)

    def _collect_selection(self) -> dict:
        """{node_tag: set(step_key)} the operator has ticked right now."""
        selected_per_node: dict[str, set[str]] = {}
        for step_key, per_node in self.checkboxes.items():
            for ntag, cb in per_node.items():
                if cb.value:
                    selected_per_node.setdefault(ntag, set()).add(step_key)
        return selected_per_node

    def _on_run(self, e):
        selected_per_node = self._collect_selection()
        if not any(selected_per_node.values()):
            return

        # Flat union — kept for any code that doesn't yet know about
        # per-node selection (counters, "any node selected this step").
        flat_union: set[str] = set()
        for s in selected_per_node.values():
            flat_union |= s

        # Hand off to the progress page via page-level attributes.
        self.page.integration_selected_steps_per_node = selected_per_node
        self.page.integration_selected_steps = flat_union
        self.page.go("/integration_run")

    def _go_back(self, e):
        self.page.go("/form")

    # ── Multi-site queue ─────────────────────────────────────────
    def _get_queue(self):
        from integration_queue import IntegrationQueue
        from app_path import get_app_dir
        if getattr(self, "_queue", None) is None:
            self._queue = IntegrationQueue(get_app_dir())
        return self._queue

    def _queue_btn_label(self) -> str:
        try:
            n = len([j for j in self._get_queue().jobs
                     if j.status in ("pending", "running")])
        except Exception:
            n = 0
        return f"Queue ({n})"

    def _add_to_queue(self, e):
        selected = self._collect_selection()
        if not any(selected.values()):
            self._alert_simple("Add to Queue",
                               "Tick at least one step first.")
            return
        form = dict(getattr(self.page, "integration_form", {}) or {})
        shortcode = form.get("shortcode", "") or "UNKNOWN"
        self._get_queue().add(
            shortcode, form,
            {k: sorted(v) for k, v in selected.items()})
        self.queue_view_btn.text = self._queue_btn_label()
        self._alert_simple("Added to Queue",
                           f"{shortcode} queued. Open Queue to run all sites "
                           f"back-to-back.")

    def _open_queue(self, e):
        q = self._get_queue()
        _icon = {"pending": (ft.Icons.SCHEDULE, TEXT_MUTED),
                 "running": (ft.Icons.AUTORENEW, ACCENT),
                 "done": (ft.Icons.CHECK_CIRCLE, SUCCESS),
                 "failed": (ft.Icons.CANCEL, DANGER),
                 "skipped": (ft.Icons.REMOVE_CIRCLE_OUTLINE, TEXT_MUTED)}
        list_col = ft.Column([], spacing=6, scroll=ft.ScrollMode.AUTO,
                             height=340)

        def _rebuild():
            list_col.controls = []
            for job in q.jobs:
                ic, col = _icon.get(job.status, (ft.Icons.HELP, TEXT_MUTED))
                list_col.controls.append(ft.Row([
                    ft.Icon(ic, color=col, size=18),
                    ft.Text(job.label(), expand=True, color=TEXT, size=13),
                    ft.Text(job.status, color=col, size=11),
                    ft.IconButton(ft.Icons.ARROW_UPWARD, icon_size=16,
                                  tooltip="Move up",
                                  on_click=lambda e, i=job.id: (
                                      q.move(i, -1), _rebuild(),
                                      self.page.update())),
                    ft.IconButton(ft.Icons.ARROW_DOWNWARD, icon_size=16,
                                  tooltip="Move down",
                                  on_click=lambda e, i=job.id: (
                                      q.move(i, 1), _rebuild(),
                                      self.page.update())),
                    ft.IconButton(ft.Icons.DELETE_OUTLINE, icon_size=16,
                                  tooltip="Remove",
                                  on_click=lambda e, i=job.id: (
                                      q.remove(i), _rebuild(),
                                      self.page.update())),
                ], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER))
            if not q.jobs:
                list_col.controls.append(
                    ft.Text("Queue is empty — use 'Add to Queue' on a site.",
                            color=TEXT_MUTED, size=12))

        _rebuild()
        c = q.counts()
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Integration Queue — {c['pending']} pending, "
                          f"{c['done']} done, {c['failed']} failed",
                          weight=ft.FontWeight.BOLD),
            content=ft.Container(width=780, content=list_col),
            actions=[
                ft.TextButton("Clear finished",
                              on_click=lambda e: (q.clear_finished(), _rebuild(),
                                                  self.page.update())),
                ft.TextButton("Close",
                              on_click=lambda e: self._close_dialog_safe(dlg)),
                ft.ElevatedButton(
                    "Run Queue", icon=ft.Icons.PLAY_ARROW_ROUNDED,
                    disabled=(c["pending"] == 0),
                    on_click=lambda e: (self._close_dialog_safe(dlg),
                                        self._run_queue())),
            ],
        )
        self._show_dialog_safe(dlg)

    def _run_queue(self):
        """Start the queue: run pending jobs one at a time. Each site's run
        auto-advances to the next on completion (see IntegrationRunPage)."""
        q = self._get_queue()
        q.reset_running_to_pending()
        job = q.next_pending()
        if job is None:
            self._alert_simple("Run Queue", "No pending jobs.")
            return
        self.page.queue_active = True
        self._launch_queue_job(job)

    def _launch_queue_job(self, job) -> None:
        q = self._get_queue()
        q.set_status(job.id, "running")
        self.page.queue_current_job_id = job.id
        self.page.integration_form = dict(job.form)
        self.page.integration_selected_steps_per_node = {
            k: set(v) for k, v in (job.selected_per_node or {}).items()}
        self.page.integration_selected_steps = {
            s for v in (job.selected_per_node or {}).values() for s in v}
        self.page.go("/integration_run")


# ═════════════════════════════════════════════════════════════════
#  Phase 2 — Progress Page (runs the selected steps)
# ═════════════════════════════════════════════════════════════════
class IntegrationRunPage:
    """Multi-column progress view with high-level log output."""

    def __init__(self, page: ft.Page):
        self.page = page
        self.cancelled = False
        self._active_ssh = {}
        self._log_queue: queue.Queue = queue.Queue()
        self._run_finished = False
        # Set by _set_step (worker threads) when a step's status cell
        # changes but no log line was emitted — tells the UI-thread
        # flush loop to repaint so progress doesn't appear frozen.
        self._ui_dirty = False
        # Serializes operator dialogs so two parallel node workers never
        # stack/overlap modal dialogs (which made the topmost one
        # un-clickable). Only one dialog is shown at a time.
        self._dialog_lock = threading.Lock()
        # Node tags the operator chose to "Skip Node" on — the worker
        # loop checks this and aborts that node's remaining steps.
        self._aborted_nodes: set = set()
        # Serializes CPU-heavy post-processing (log parsing + openpyxl
        # Excel building). These hold Python's GIL for 1-3 s; if all 3
        # node workers run them at once the asyncio event loop is
        # starved and the Flet window paints BLACK (no render frames
        # sent). Serializing caps it to one CPU-bound section at a time
        # so the event loop gets GIL gaps to keep the UI alive.
        self._heavy_lock = threading.Lock()

        # ── Per-node live terminal tabs ──────────────────────────
        # A real mirror of the moshell/SSH session per node — every byte
        # the node sends back (the actual `crn ...`, `ERROR: MO already
        # exists`, prompts, etc.), not just a high-level summary.
        # CPU/memory-safe by design:
        #   * per node we keep a bounded TEXT scrollback (last N chars),
        #     so RAM is tiny and appends are cheap on the SSH hot path
        #   * ONLY the active tab is rendered → zero render cost while the
        #     operator stays on the default "Session" tab
        #   * ANSI stripping + line splitting happen at RENDER time on the
        #     ~300 visible lines only — never on the SSH read path
        self._RAW_MAXCHARS = 48000                # ~600 lines scrollback
        self._active_log_tab = "session"          # "session" | node tag
        # Scrollback stored as a deque of raw chunks + a running char
        # count, so appends are O(chunk) (append + a couple of poplefts)
        # instead of recopying the whole ~48 KB window every recv.
        self._node_raw: dict[str, collections.deque] = {}
        self._node_raw_len: dict[str, int] = {}
        self._node_live_dirty: dict[str, bool] = {}
        self._node_live_column: dict[str, ft.Column] = {}
        self._live_lock = threading.Lock()
        self._log_tab_buttons: dict[str, ft.Container] = {}

        # Relation remote-folder cleanup coordination. The relation zip
        # is unzipped into a SHARED ``/RELATION/<shortcode>/`` folder
        # with one subfolder per node. We must only delete it once ALL
        # relation-running nodes are done — otherwise the first node to
        # finish wipes a still-running sibling's scripts mid-run.
        self._relation_cleanup_lock = threading.Lock()
        self._relation_nodes_expected: Optional[set] = None
        self._relation_nodes_done: set = set()

        self.form = getattr(page, "integration_form", {})
        self.selected_steps: set = getattr(
            page, "integration_selected_steps", set())
        # Per-node selection from the multi-column checklist. Keys are
        # node tags ("lte", "lte2", "gsm"); values are sets of step
        # keys the operator ticked for that node.
        self.selected_steps_per_node: dict[str, set] = getattr(
            page, "integration_selected_steps_per_node", {})

        # Node identifiers
        self.lte_name = self.form.get("node_name", "LTE/NR Node")
        self.lte_ip = self.form.get("node_ip", "")
        self.lte_subnet = self.form.get("subnetwork", "")
        self.lte2_name = self.form.get("node2_name", "")
        self.lte2_ip = self.form.get("node2_ip", "")
        self.lte2_subnet = self.form.get("node2_subnetwork", "")
        self.gsm_name = self.form.get("gsm_node_name", "")
        self.gsm_ip = self.form.get("gsm_node_ip", "")
        self.gsm_subnet = self.form.get("gsm_subnetwork", "")
        self.has_lte2 = bool(self.lte2_name)
        self.has_gsm = bool(self.gsm_name)
        # See IntegrationPage.__init__ for rationale — co-located
        # GSM (BSC filled, GSM node name blank) routes GSM steps to
        # the primary LTE node instead of a separate column.
        self.gsm_on_primary = (
            not self.has_gsm
            and bool(str(self.form.get("bsc_name", "") or "").strip())
        )

        self.shortcode = self.form.get("shortcode", "UNKNOWN")

        from app_path import get_app_dir
        self.log_dir = os.path.join(get_app_dir(), "LOG", self.shortcode)
        os.makedirs(self.log_dir, exist_ok=True)

        # Per-node session log buffers — full detail for file saving
        # Keyed by node_tag ("lte", "lte2", "gsm") to avoid cross-contamination
        self._session_logs: dict[str, list[str]] = {"lte": [], "lte2": [], "gsm": []}
        self._session_log_lock = threading.Lock()

        # Per-node step result tracking: {node_tag: {step_key: state}}
        self._step_results: dict[str, dict[str, str]] = {}
        # Per-node step remark/detail: {node_tag: {step_key: detail}} — the same
        # short status string shown on the progress row (e.g. the SGw check's
        # "1/29 with packet loss"), carried into the summary's Remark column.
        self._step_details: dict[str, dict[str, str]] = {}

        # Per-node duration tracking: {node_tag: seconds}
        self._node_durations: dict[str, float] = {}

        # ── Per-node resume support ──────────────────────────────
        # When a node stops early (fail/skip) with steps still pending,
        # a "Resume" button on that node's column re-runs its workflow
        # from the failed step (skipping already-done steps) with a
        # fresh SSH session. Other nodes are unaffected.
        self._resume_buttons: dict[str, ft.Control] = {}
        self._node_params: dict[str, dict] = {}   # tag -> run kwargs
        self._resuming: set = set()               # tags currently re-running
        # All node workers currently running (initial + resumed). Finalize
        # (+ summary popup) only fires when this is empty, so the popup can
        # never appear while ANY node — including a resumed one — is busy.
        self._active_workers: set = set()
        # Nodes that stopped early (fail/abort/AMOS-fail) with steps left
        # unfinished — these get a Resume button. Explicit set, because
        # stopped steps are marked "skip" and can't be told apart from
        # legitimate skips by state alone.
        self._resumable_nodes: set = set()

    # ── Build ────────────────────────────────────────────────────
    def build(self) -> ft.View:
        # Returning from the CDD Audit page? Restore the last run's results and
        # re-show the summary instead of re-running the whole integration.
        self._summary_only = bool(
            getattr(self.page, "integration_show_summary_only", False))
        if self._summary_only:
            self.page.integration_show_summary_only = False
            self._step_results = (
                getattr(self.page, "integration_step_results", {}) or {})
            self._step_details = (
                getattr(self.page, "integration_step_details", {}) or {})
            self._node_durations = (
                getattr(self.page, "integration_node_durations", {}) or {})
            self._run_finished = True
        else:
            # Fresh run — no summary to return to until this one finishes, so a
            # stale "Back to Summary" from a previous run can't fire.
            self.page.integration_has_summary = False

        # Put the SHORTCODE in the OS window title so the taskbar
        # preview / Alt-Tab shows which site each window is running —
        # essential when the operator opens several NodeCraft windows.
        try:
            self.page.title = (
                f"NodeCraft — {self.shortcode} (Integration)"
            )
            self.page.update()
        except Exception:
            pass

        self.back_button = ft.ElevatedButton(
            "Back to Form",
            icon=ft.Icons.ARROW_BACK,
            style=secondary_button_style(),
            on_click=self._go_back,
            visible=False,
        )
        self.cancel_button = ft.ElevatedButton(
            "Cancel",
            icon=ft.Icons.CANCEL_OUTLINED,
            style=ft.ButtonStyle(
                bgcolor=DANGER, color=TEXT,
                padding=ft.Padding.symmetric(horizontal=18, vertical=14),
                shape=ft.RoundedRectangleBorder(radius=14),
            ),
            on_click=self._on_cancel,
        )
        self.status_text = ft.Text(
            "Integration in progress...",
            size=22, weight=ft.FontWeight.BOLD, color=TEXT,
        )
        self.elapsed_text = ft.Text("00:00", size=13, color=TEXT_MUTED)
        # ``auto_scroll=True`` makes Flet keep the view pinned to the
        # newest line automatically — no manual scroll_to (which is an
        # un-awaited coroutine in this Flet version and never actually
        # scrolled).
        self.log_column = ft.Column(
            spacing=2, scroll=ft.ScrollMode.AUTO, expand=True,
            auto_scroll=True,
        )

        # Build step rows for each node column
        self.lte_steps: dict[str, _StepRow] = {}
        self.lte2_steps: dict[str, _StepRow] = {}
        self.gsm_steps: dict[str, _StepRow] = {}

        # Per-node-tag label override: a few steps want a different
        # display label depending on whether the column is LTE/NR or
        # GSM. ``sgw_check`` is the canonical example — the underlying
        # check is the same, but the user-facing name differs.
        def _row_label(key: str, default: str, ntag: str) -> str:
            if key == "sgw_check":
                return (
                    "Verify BSC Broker IP Reachability"
                    if ntag == "gsm"
                    else "Verify SGw Reachability"
                )
            return default

        # Filter rule for progress columns (per-node):
        #   * Always skip SUMMARY_NA steps (they only live in the summary).
        #   * REMARK steps always show (bb_transport, enrollment_sync —
        #     their state is derived, not selected).
        #   * Otherwise: only show if the operator selected this step
        #     for THIS node in the multi-column checklist. Un-selected
        #     steps don't appear in the column.
        def _should_show_for_node(key: str, ntag: str) -> bool:
            if key in SUMMARY_NA_STEPS:
                return False
            if key in REMARK_STEPS:
                return True
            return self.is_step_selected(key, ntag)

        # Pre-verified REMARK rows that should always show as "done".
        # Distinct from ``enrollment_sync`` which derives its state
        # from the enrollment step at run time.
        PRE_VERIFIED = {"bb_transport", "sw_package_check", "bb_power_on"}

        def _apply_remark_state(row, key: str) -> None:
            if key in PRE_VERIFIED:
                row.set_state("done", "Pre-verified")
            elif key == "enrollment_sync":
                row.set_state("pending", "Waiting for enrollment")

        lte_rows = []
        lte2_rows = []
        gsm_rows = []
        gsm_on_primary = getattr(self, "gsm_on_primary", False)
        for key, label, applies_to, _ in INTEGRATION_STEPS:
            # Route to the correct column(s) based on the step's
            # applies_to scope. Uses the module-level helper so the
            # ``lte_primary`` scope (primary LTE only) is honoured —
            # and so ``gsm`` steps fold into the LTE column when
            # ``gsm_on_primary`` (BSC filled, no separate GSM node).
            if _step_applies_to_node(applies_to, "lte", gsm_on_primary):
                if _should_show_for_node(key, "lte"):
                    row = _StepRow(_row_label(key, label, "lte"))
                    self.lte_steps[key] = row
                    lte_rows.append(row.control)
                    if key in REMARK_STEPS:
                        _apply_remark_state(row, key)
            if self.has_lte2 and _step_applies_to_node(
                applies_to, "lte2", gsm_on_primary,
            ):
                if _should_show_for_node(key, "lte2"):
                    row2 = _StepRow(_row_label(key, label, "lte2"))
                    self.lte2_steps[key] = row2
                    lte2_rows.append(row2.control)
                    if key in REMARK_STEPS:
                        _apply_remark_state(row2, key)
            if _step_applies_to_node(applies_to, "gsm", gsm_on_primary):
                if _should_show_for_node(key, "gsm"):
                    row = _StepRow(_row_label(key, label, "gsm"))
                    self.gsm_steps[key] = row
                    gsm_rows.append(row.control)
                    if key in REMARK_STEPS:
                        _apply_remark_state(row, key)

        columns = []
        lte_column = self._node_card(
            title=self.lte_name or "LTE/NR Node",
            subtitle=self._node_subtitle(self.lte_ip, self.lte_subnet),
            accent=INFO,
            icon=ft.Icons.CELL_TOWER,
            step_rows=lte_rows,
            node_tag="lte",
        )
        columns.append(lte_column)

        if self.has_lte2:
            lte2_column = self._node_card(
                title=self.lte2_name or "LTE/NR #2 Node",
                subtitle=self._node_subtitle(self.lte2_ip, self.lte2_subnet),
                accent=ACCENT,
                icon=ft.Icons.CELL_TOWER,
                step_rows=lte2_rows,
                node_tag="lte2",
            )
            columns.append(lte2_column)

        if self.has_gsm:
            gsm_column = self._node_card(
                title=self.gsm_name or "GSM Node",
                subtitle=self._node_subtitle(self.gsm_ip, self.gsm_subnet),
                accent=ACCENT_WARM,
                icon=ft.Icons.SETTINGS_INPUT_ANTENNA,
                step_rows=gsm_rows,
                node_tag="gsm",
            )
            columns.append(gsm_column)

        header = ft.Row(
            [
                ft.Column(
                    [self.status_text, self.elapsed_text],
                    spacing=4, expand=True,
                ),
                self.cancel_button,
                self.back_button,
            ],
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        )

        columns_row = ft.Row(
            columns,
            spacing=16,
            alignment=ft.MainAxisAlignment.START,
            vertical_alignment=ft.CrossAxisAlignment.START,
            expand=True,
        )

        # ── Build the log tab bar: Session + one per connected node ──
        log_tabs: list[tuple[str, str]] = [("session", "Session")]
        log_tabs.append(("lte", self.lte_name or "LTE/NR"))
        if self.has_lte2:
            log_tabs.append(("lte2", self.lte2_name or "LTE/NR #2"))
        if self.has_gsm:
            log_tabs.append(("gsm", self.gsm_name or "GSM"))

        # Per-node scrollback + columns (only built for real nodes).
        for tab_key, _ in log_tabs:
            if tab_key == "session":
                continue
            self._node_raw[tab_key] = collections.deque()
            self._node_raw_len[tab_key] = 0
            self._node_live_dirty[tab_key] = False
            self._node_live_column[tab_key] = ft.Column(
                spacing=1, scroll=ft.ScrollMode.AUTO, expand=True,
                auto_scroll=True,
            )

        tab_btn_row = ft.Row(spacing=4, scroll=ft.ScrollMode.AUTO)
        for tab_key, tab_label in log_tabs:
            short = tab_label if len(tab_label) <= 22 else tab_label[:20] + "…"
            btn = ft.Container(
                content=ft.Text(short, size=12,
                                weight=ft.FontWeight.W_600),
                padding=ft.Padding(left=12, top=5, right=12, bottom=5),
                border_radius=8,
                on_click=lambda e, k=tab_key: self._switch_log_tab(k),
                ink=True,
            )
            self._log_tab_buttons[tab_key] = btn
            tab_btn_row.controls.append(btn)

        # Content host swaps between Session log and node live columns.
        self._log_content_host = ft.Container(
            content=self.log_column, expand=True,
        )
        self._style_log_tabs()  # set initial active highlight

        log_panel = panel(
            ft.Column(
                [
                    tab_btn_row,
                    self._log_content_host,
                ],
                spacing=8,
                expand=True,
            ),
            bgcolor=PANEL_RAISED,
            padding=8,
            expand=True,
        )
        log_panel.width = float("inf")

        body = ft.Container(
            expand=True,
            gradient=background_gradient(),
            padding=ft.Padding.symmetric(horizontal=4, vertical=8),
            content=ft.Column(
                [header, columns_row, log_panel],
                spacing=10,
                expand=True,
            ),
        )

        # Start timer and workflow
        self._start_time = datetime.now()
        self._timer_running = True
        if self._summary_only:
            # Re-entry from CDD Audit: don't re-run — just re-show the summary.
            self._timer_running = False
            self.page.run_task(self._reshow_summary_task)
        else:
            threading.Thread(target=self._tick_timer, daemon=True).start()
            threading.Thread(target=self._run_workflow, daemon=True).start()
        # Phone monitor pusher (no-op thread exits immediately if the
        # Telegram monitor isn't configured).
        self.page.run_task(self._flush_loop)

        return ft.View(
            route="/integration_run", padding=0, spacing=0,
            bgcolor=BG_TOP, controls=[body],
        )

    # ── UI helpers ───────────────────────────────────────────────
    def _node_subtitle(self, ip: str, subnet: str) -> str:
        parts = []
        if ip:
            parts.append(f"IP: {ip}")
        if subnet:
            parts.append(f"Subnet: {subnet}")
        return "  |  ".join(parts) if parts else "No IP / Subnetwork configured"

    def _node_card(
        self, title: str, subtitle: str, accent: str, icon: str,
        step_rows: list[ft.Control], dimmed: bool = False,
        node_tag: str = "",
    ) -> ft.Container:
        opacity = 0.4 if dimmed else 1.0
        # Per-node Resume button — hidden until the node stops early with
        # steps still pending (see _update_resume_button).
        resume_btn = ft.ElevatedButton(
            "Resume",
            icon=ft.Icons.PLAY_ARROW,
            visible=False,
            tooltip="Re-run this node from the failed step to the end",
            style=ft.ButtonStyle(
                bgcolor=ACCENT_WARM, color="#06242A",
                shape=ft.RoundedRectangleBorder(radius=10),
            ),
            on_click=lambda e, nt=node_tag: self._resume_node(nt),
        )
        if node_tag:
            self._resume_buttons[node_tag] = resume_btn
        return ft.Container(
            expand=1,
            opacity=opacity,
            content=panel(
                ft.Column(
                    [
                        ft.Row(
                            [
                                ft.Icon(icon, size=20, color=accent),
                                ft.Text(title, size=16,
                                        weight=ft.FontWeight.BOLD, color=TEXT),
                                ft.Container(expand=True),
                                resume_btn,
                            ],
                            spacing=8,
                            vertical_alignment=ft.CrossAxisAlignment.CENTER,
                        ),
                        ft.Text(subtitle, size=11, color=TEXT_MUTED),
                        ft.Divider(height=1, color=BORDER),
                        ft.Column(
                            step_rows,
                            spacing=10,
                            scroll=ft.ScrollMode.AUTO,
                            expand=True,
                        ),
                    ],
                    spacing=10,
                    expand=True,
                ),
                bgcolor=PANEL,
                padding=20,
            ),
        )

    # ── Logging ──────────────────────────────────────────────────
    def _ui_log(self, msg: str):
        """High-level log shown in the UI panel only (not saved to file)."""
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        self._log_queue.put_nowait(line)

    def _detail_log(self, msg: str, node_tag: str = "lte"):
        """Full-detail log saved to step log files only. (The per-node
        LIVE tab is fed separately from the raw SSH stream — see
        ``_feed_raw`` — so it mirrors the real moshell terminal.)"""
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        with self._session_log_lock:
            self._session_logs[node_tag].append(line)

    # ── Live per-node terminal feed (raw moshell stream) ─────────
    def _feed_raw(self, node_tag: str, chunk: str) -> None:
        """Append a raw SSH recv chunk to a node's scrollback. Runs on
        the SSH reader thread → kept minimal: just append + cap. ANSI
        cleanup and line splitting are deferred to render time (active
        tab only)."""
        if not chunk:
            return
        dq = self._node_raw.get(node_tag)
        if dq is None:
            return
        cap = self._RAW_MAXCHARS
        # A single recv can't exceed the whole window — trim it alone.
        if len(chunk) > cap:
            chunk = chunk[-cap:]
        with self._live_lock:
            dq.append(chunk)
            total = self._node_raw_len[node_tag] + len(chunk)
            # Drop oldest chunks until back under the cap (keep ≥1).
            while total > cap and len(dq) > 1:
                total -= len(dq.popleft())
            self._node_raw_len[node_tag] = total
            if self._active_log_tab == node_tag:
                self._node_live_dirty[node_tag] = True

    # ── Live log tab switching ─────────────────────────────────
    def _style_log_tabs(self) -> None:
        for key, btn in self._log_tab_buttons.items():
            active = (key == self._active_log_tab)
            btn.bgcolor = (
                ft.Colors.with_opacity(0.20, ACCENT) if active else None
            )
            try:
                btn.content.color = ACCENT if active else TEXT_MUTED
            except Exception:
                pass

    def _switch_log_tab(self, tab_key: str) -> None:
        self._active_log_tab = tab_key
        if tab_key == "session":
            self._log_content_host.content = self.log_column
        else:
            col = self._node_live_column.get(tab_key)
            if col is not None:
                # Full one-time render of the buffer on switch.
                self._render_node_live(tab_key, full=True)
                self._log_content_host.content = col
        self._style_log_tabs()
        try:
            self.page.update()
        except Exception:
            pass

    # Strips ANSI/VT100 escape sequences (colour codes, cursor moves)
    # that moshell emits so the terminal view shows clean text.
    _ANSI_RE = re.compile(r"\x1b\[[0-9;?]*[ -/]*[@-~]")

    def _render_node_live(self, tab_key: str, full: bool = False) -> None:
        """Rebuild the active node tab's terminal from its scrollback.
        Only ever called for the ACTIVE tab. ANSI strip + split run here
        (on ≤48 KB), never on the SSH read path."""
        col = self._node_live_column.get(tab_key)
        if col is None:
            return
        with self._live_lock:
            dq = self._node_raw.get(tab_key)
            raw = "".join(dq) if dq else ""
            self._node_live_dirty[tab_key] = False
        clean = self._ANSI_RE.sub("", raw).replace("\r", "")
        lines = clean.split("\n")[-300:]   # cap visible rows
        col.controls = [
            ft.Text(ln if ln else " ", size=11, color=TEXT_MUTED,
                    selectable=True, font_family="Consolas")
            for ln in lines
        ]

    def _save_step_log(self, step_number: int, node_name: str,
                       log_suffix: str, node_tag: str = "lte"):
        """Save per-node session log snapshot to LOG/<SHORTCODE>/SESSION/."""
        filename = f"{step_number:02d}_{node_name}_{log_suffix}.txt"
        session_dir = os.path.join(self.log_dir, "SESSION")
        os.makedirs(session_dir, exist_ok=True)
        filepath = os.path.join(session_dir, filename)
        # MEMORY FIX: snapshot THIS step's lines then CLEAR the buffer.
        # Previously we copied the entire accumulated list every step
        # (O(steps²) memory + CPU + disk, and each numbered file was
        # cumulative). With 3 nodes each logging multi-MB baseline /
        # relation output, that churn is what made 3-node runs slow.
        # Now each numbered file holds only that step's lines and the
        # in-memory buffer is freed after every step — peak memory is
        # bounded to roughly one step's output per node.
        with self._session_log_lock:
            snapshot = self._session_logs.get(node_tag, [])
            self._session_logs[node_tag] = []
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(f"Integration Log — {self.shortcode}\n")
                f.write(f"Node: {node_name} | Step: {log_suffix}\n")
                f.write(f"Saved at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 72 + "\n\n")
                f.write("\n".join(snapshot))
                f.write("\n")
            logger.info(f"[Integration] Log saved: {filepath}")
        except Exception as exc:
            logger.warning(f"[Integration] Failed to save log: {exc}")

    async def _flush_loop(self):
        # 0.4 s (not 0.25 s) so the event loop spends less time on
        # full-page updates and more handling client input — keeps the
        # modal dialog buttons responsive under 3-node load.
        while not self._run_finished or not self._log_queue.empty():
            self._flush_log()
            await asyncio.sleep(0.4)
        self._flush_log()

    def _flush_log(self):
        dirty = False
        had_new_log = False
        # Cap how many lines we ingest per tick so a sudden burst (3
        # nodes flooding) can't make one update() serialize a huge diff
        # and stall the event loop.
        ingested = 0
        while not self._log_queue.empty() and ingested < 200:
            try:
                msg = self._log_queue.get_nowait()
            except queue.Empty:
                break
            ingested += 1
            self.log_column.controls.append(
                ft.Text(msg, size=11, color=TEXT_MUTED, selectable=True,
                        font_family="Consolas")
            )
            # Smaller retained window → cheaper page.update() diffs.
            if len(self.log_column.controls) > 300:
                self.log_column.controls = self.log_column.controls[-250:]
            dirty = True
            had_new_log = True
        # Also repaint when a step status cell changed without a log
        # line (``_ui_dirty`` set by _set_step from a worker thread).
        if self._ui_dirty:
            dirty = True
            self._ui_dirty = False
        # Render the active node LIVE tab only when it has new lines.
        # Inactive node tabs are never rendered (their ring buffers
        # just hold the last N lines) → zero cost while on Session.
        atab = self._active_log_tab
        if atab != "session" and self._node_live_dirty.get(atab):
            self._render_node_live(atab)
            dirty = True
        if dirty:
            # No manual scroll_to — log_column has auto_scroll=True, so
            # Flet keeps the newest line in view on its own.
            try:
                self.page.update()
            except Exception:
                pass

    def _tick_timer(self):
        import time
        while self._timer_running:
            elapsed = datetime.now() - self._start_time
            mins, secs = divmod(int(elapsed.total_seconds()), 60)
            self.elapsed_text.value = f"{mins:02d}:{secs:02d}"
            # Do NOT call page.update() here — that was a second
            # cross-thread updater competing with the flush loop for
            # the event loop (and could stall click handling on the
            # modal dialogs). Just mark dirty; the flush loop (the
            # single updater, on the event loop) repaints the timer.
            self._ui_dirty = True
            time.sleep(1)

    # ── Step state updates ───────────────────────────────────────
    def _set_step(self, node: str, key: str, state: str, detail: str = ""):
        if node == "lte":
            steps = self.lte_steps
        elif node == "lte2":
            steps = self.lte2_steps
        else:
            steps = self.gsm_steps
        if key in steps:
            steps[key].set_state(state, detail)
        self._step_results.setdefault(node, {})[key] = state
        # Keep the latest non-empty remark for this step; don't let a later
        # empty detail wipe a meaningful one.
        if detail:
            self._step_details.setdefault(node, {})[key] = detail
        # Mark the UI dirty so the flush loop repaints the status cell
        # on its next tick (≤0.25 s) even if no log line follows. This
        # is what keeps the progress column in sync with the worker —
        # previously the cell only updated when a log line happened to
        # be emitted right after.
        self._ui_dirty = True

    def is_step_selected(self, step_key: str, node_tag: str) -> bool:
        """Whether the operator ticked ``step_key`` for ``node_tag`` in
        the multi-column checklist. Falls back to the flat union if a
        per-node selection wasn't provided (older page state)."""
        per_node = self.selected_steps_per_node.get(node_tag)
        if per_node is not None:
            return step_key in per_node
        return step_key in self.selected_steps

    def _retry_step(
        self,
        label: str,
        fn,
        node_tag: str,
        step_key: str,
        max_attempts: int = 3,
        backoff: float = 3.0,
    ) -> tuple[bool, str]:
        """Run ``fn`` up to ``max_attempts`` times; mark "running (retry N)"
        on each retry so the operator sees the attempt count live. Returns
        the last (success, output). Never raises — exceptions are treated
        as a failed attempt.

        Use only for **idempotent** steps. Side-effecting steps that
        submit jobs or apply MOs (enrollment, install_lkf, baseline,
        relation) are NOT wrapped with this — retrying them could create
        duplicate jobs or backups.
        """
        last_output = ""
        for attempt in range(1, max_attempts + 1):
            if attempt > 1:
                self._set_step(
                    node_tag, step_key, "running",
                    f"retry {attempt}/{max_attempts}",
                )
                self._ui_log(
                    f"[retry] {label}: attempt {attempt}/{max_attempts}…"
                )
            try:
                success, output = fn()
                last_output = output if isinstance(output, str) else str(output)
                if success:
                    if attempt > 1:
                        self._ui_log(
                            f"[retry] {label}: passed on attempt "
                            f"{attempt}/{max_attempts}."
                        )
                    return True, last_output
                self._ui_log(
                    f"[retry] {label}: attempt {attempt}/{max_attempts} "
                    "returned failure."
                )
            except Exception as exc:
                last_output = f"{type(exc).__name__}: {exc}"
                self._ui_log(
                    f"[retry] {label}: attempt {attempt}/{max_attempts} "
                    f"raised {last_output}."
                )
            if attempt < max_attempts:
                time.sleep(backoff)
        return False, last_output

    def _summary_cell(self, key: str, ntag: str, applies_to: str,
                      results: dict[str, str]) -> str:
        if key in SUMMARY_NA_STEPS:
            return "N/A"

        state = results.get(key, "skip")

        if key == "enrollment_sync":
            state = results.get("enrollment", state)

        # Pre-verified manual rows — always rendered as done.
        if key in ("bb_transport", "sw_package_check", "bb_power_on"):
            state = "done"

        return _RESULT_SYMBOLS.get(state, "—")

    # ── Dialog plumbing (Flet 0.84 correct API) ─────────────────
    def _show_dialog_safe(self, dlg) -> None:
        """Open a modal dialog using the Flet 0.84 dialog stack
        (``page.show_dialog``). Falls back to the legacy
        ``overlay.append + open=True`` only if show_dialog is absent.

        The OLD overlay approach is exactly what made dialogs
        un-closable: a dialog placed in ``page.overlay`` is NOT in
        Flet's managed ``_dialogs`` stack, so toggling ``dlg.open``
        never dismisses it and its action buttons don't resolve.
        """
        if hasattr(self.page, "show_dialog"):
            try:
                self.page.show_dialog(dlg)
                return
            except Exception as exc:
                logger.warning(f"show_dialog failed ({exc}); using overlay fallback")
        try:
            self.page.overlay.append(dlg)
            dlg.open = True
            self.page.update()
        except Exception as exc:
            logger.warning(f"overlay dialog fallback failed: {exc}")

    def _close_dialog_safe(self, dlg) -> None:
        """Dismiss a dialog opened via _show_dialog_safe. Tries the
        managed stack (``pop_dialog``) first, then direct ``open=False``,
        then overlay removal — whichever the current Flet supports."""
        closed = False
        if hasattr(self.page, "pop_dialog"):
            try:
                self.page.pop_dialog()
                closed = True
            except Exception:
                pass
        if not closed:
            try:
                dlg.open = False
                dlg.update()
                closed = True
            except Exception:
                pass
        # Belt-and-braces: remove from overlay if it ended up there.
        try:
            if dlg in (self.page.overlay or []):
                self.page.overlay.remove(dlg)
        except Exception:
            pass
        try:
            self.page.update()
        except Exception:
            pass

    # ── Blocking dialogs (called from worker threads) ────────────
    def _ask_user_retry(self, message: str, node_tag: Optional[str] = None) -> bool:
        """Modal Retry / Skip Node / Stop dialog.

        Returns True only for **Retry**. **Stop** and **Skip Node**
        both return False; **Skip Node** additionally marks
        ``node_tag`` so the worker loop aborts that node's remaining
        steps (the operator's "force-close everything on this node"
        escape hatch).

        Serialized via ``self._dialog_lock`` so parallel node workers
        never stack overlapping modal dialogs.
        """
        with self._dialog_lock:
            result = {"choice": "stop"}
            ev = threading.Event()

            def _resolve(choice):
                # Idempotent — ignore double fires (e.g. on_dismiss
                # after a button already resolved).
                if ev.is_set():
                    return
                result["choice"] = choice
                self._close_dialog_safe(dlg)
                ev.set()

            actions = [
                ft.TextButton(
                    "Stop",
                    style=ft.ButtonStyle(color=DANGER),
                    on_click=lambda e: _resolve("stop"),
                ),
            ]
            if node_tag:
                actions.append(
                    ft.TextButton(
                        "Skip Node",
                        icon=ft.Icons.SKIP_NEXT,
                        style=ft.ButtonStyle(color=ACCENT_WARM),
                        on_click=lambda e: _resolve("skip"),
                    )
                )
            actions.append(
                ft.ElevatedButton(
                    "Retry", icon=ft.Icons.REFRESH,
                    style=ft.ButtonStyle(
                        bgcolor=ACCENT, color="#06242A",
                        padding=ft.Padding.symmetric(horizontal=16,
                                                     vertical=12),
                        shape=ft.RoundedRectangleBorder(radius=12)),
                    on_click=lambda e: _resolve("retry"),
                )
            )

            hint = (
                "Retry re-runs this check. "
                + ("Skip Node aborts the rest of this node only. "
                   if node_tag else "")
                + "Stop ends this step."
            )
            dlg = ft.AlertDialog(
                modal=True,
                title=ft.Text("Verification Failed", color=DANGER,
                              weight=ft.FontWeight.BOLD),
                content=ft.Column(
                    [
                        ft.Text(message, size=13, color=TEXT,
                                selectable=True),
                        ft.Container(height=8),
                        ft.Text(hint, size=12, color=TEXT_MUTED,
                                italic=True),
                    ],
                    tight=True,
                ),
                actions=actions,
                actions_alignment=ft.MainAxisAlignment.END,
                bgcolor=PANEL,
                on_dismiss=lambda e: _resolve("stop"),
            )

            self._show_dialog_safe(dlg)
            ev.wait()

            if result["choice"] == "skip" and node_tag:
                self._aborted_nodes.add(node_tag)
                self._ui_log(
                    f"Operator chose Skip Node — aborting remaining "
                    f"steps for this node."
                )
            return result["choice"] == "retry"

    def _ask_user_confirm(self, title: str, message: str) -> bool:
        """Modal OK / Cancel confirmation. Serialized + Flet-0.84-safe."""
        with self._dialog_lock:
            result = {"ok": False}
            ev = threading.Event()

            def _resolve(ok: bool):
                if ev.is_set():
                    return
                result["ok"] = ok
                self._close_dialog_safe(dlg)
                ev.set()

            dlg = ft.AlertDialog(
                modal=True,
                title=ft.Text(title, color=INFO, weight=ft.FontWeight.BOLD),
                content=ft.Text(message, size=13, color=TEXT, selectable=True),
                actions=[
                    ft.TextButton(
                        "Cancel",
                        style=ft.ButtonStyle(color=DANGER),
                        on_click=lambda e: _resolve(False),
                    ),
                    ft.ElevatedButton(
                        "OK, Run It", icon=ft.Icons.CHECK,
                        style=ft.ButtonStyle(
                            bgcolor=ACCENT, color="#06242A",
                            padding=ft.Padding.symmetric(horizontal=16,
                                                         vertical=12),
                            shape=ft.RoundedRectangleBorder(radius=12)),
                        on_click=lambda e: _resolve(True),
                    ),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                bgcolor=PANEL,
                on_dismiss=lambda e: _resolve(False),
            )

            self._show_dialog_safe(dlg)
            ev.wait()
            return result["ok"]

    # ── Workflow ───────────────────────────────────────────────────
    def _run_workflow(self):
        from concurrent.futures import ThreadPoolExecutor

        self._ui_log("Starting integration workflow...")
        self._ui_log(f"Site ID: {self.shortcode}")
        self._ui_log(f"Log directory: {self.log_dir}")
        self._ui_log(
            f"LTE/NR #1: {self.lte_name}  IP: {self.lte_ip}  "
            f"Subnet: {self.lte_subnet}")

        if self.has_lte2:
            self._ui_log(
                f"LTE/NR #2: {self.lte2_name}  IP: {self.lte2_ip}  "
                f"Subnet: {self.lte2_subnet}")
        else:
            for key, label, applies_to, _ in INTEGRATION_STEPS:
                if applies_to in ("both", "lte_nr"):
                    self._set_step("lte2", key, "skip", "No 2nd LTE/NR node")

        # Log the selection per node so the operator can verify the
        # multi-column checklist actually got handed off correctly.
        node_label_map = {"lte": "LTE", "lte2": "LTE 2", "gsm": "GSM"}
        for ntag in ("lte", "lte2", "gsm"):
            per_node = self.selected_steps_per_node.get(ntag)
            if not per_node:
                continue
            names = [
                lbl for k, lbl, _, _ in INTEGRATION_STEPS
                if k in per_node
            ]
            self._ui_log(
                f"Selected steps for {node_label_map[ntag]}: "
                f"{', '.join(names)}"
            )

        if self.has_gsm:
            self._ui_log(
                f"GSM: {self.gsm_name}  IP: {self.gsm_ip}  "
                f"Subnet: {self.gsm_subnet}")
        else:
            for key, label, applies_to, _ in INTEGRATION_STEPS:
                if applies_to in ("both", "gsm"):
                    self._set_step("gsm", key, "skip", "No GSM node")
        self._ui_log("")

        # Remember each node's run parameters so the Resume button can
        # re-invoke _run_node_steps for a single node later.
        self._node_params["lte"] = dict(
            node_tag="lte", node_name=self.lte_name, node_ip=self.lte_ip,
            subnetwork=self.lte_subnet, node_type="lte_nr")
        if self.has_lte2:
            self._node_params["lte2"] = dict(
                node_tag="lte2", node_name=self.lte2_name,
                node_ip=self.lte2_ip, subnetwork=self.lte2_subnet,
                node_type="lte_nr")
        if self.has_gsm:
            self._node_params["gsm"] = dict(
                node_tag="gsm", node_name=self.gsm_name,
                node_ip=self.gsm_ip, subnetwork=self.gsm_subnet,
                node_type="gsm")

        num_workers = 1 + (1 if self.has_lte2 else 0) + (1 if self.has_gsm else 0)
        for ntag in self._node_params:
            self._active_workers.add(ntag)
        with ThreadPoolExecutor(max_workers=num_workers) as pool:
            futures = [pool.submit(self._node_worker, ntag, params)
                       for ntag, params in self._node_params.items()]
            for f in futures:
                try:
                    f.result()
                except Exception:
                    logger.exception("Node worker crashed")
        # NOTE: finalize is NOT called here. Each worker calls
        # _maybe_finalize() as it ends; the popup only fires once EVERY
        # worker (including any resumed node) is idle.

    def _node_worker(self, node_tag: str, params: dict):
        """Common wrapper for a node's run (initial OR resumed). Ensures
        the run is only finalized when no node worker is left running."""
        lbl = {"lte": "LTE/NR", "lte2": "LTE/NR #2"}.get(node_tag, "GSM")
        try:
            self._run_node_steps(**params)
        except Exception as exc:
            self._ui_log(f"[{lbl}] Node workflow failed: {exc}")
            logger.exception(f"Integration {lbl} failed")
        finally:
            self._resuming.discard(node_tag)
            self._active_workers.discard(node_tag)
            self._maybe_finalize()

    def _maybe_finalize(self):
        """Finalize (+ summary popup) only when NO node worker is busy.
        Guarded by ``_run_finished`` so two workers ending at once can't
        double-finalize; a resume resets the flag so it fires again."""
        if not self._active_workers and not self._run_finished:
            self._finalize_run(show_popup=True)

    def _finalize_run(self, show_popup: bool = True):
        """Called when the current pass (initial run or a resume) ends
        and no node is actively running."""
        self._timer_running = False
        self._run_finished = True
        # Reveal Resume on any node that stopped early.
        for ntag in self._node_params:
            self._update_resume_button(ntag)
        any_resumable = bool(self._resumable_nodes)
        self.status_text.value = (
            "Integration paused — Resume available on failed node(s)"
            if any_resumable else "Integration complete"
        )
        self.cancel_button.visible = False
        self.back_button.visible = True
        self._ui_log("")
        self._ui_log(f"All active steps finished. Logs saved to: {self.log_dir}")
        if show_popup:
            self._show_summary_popup()
        try:
            self.page.update()
        except Exception:
            pass
        # Multi-site queue: this site is done — advance to the next one.
        if getattr(self.page, "queue_active", False):
            try:
                self.page.run_task(self._advance_queue_task, any_resumable)
            except Exception:
                pass

    async def _advance_queue_task(self, failed: bool):
        """After a queued site finishes, mark it and launch the next pending
        one. A short pause lets the summary Excel save and the operator glimpse
        the result before the view swaps to the next site."""
        import asyncio
        await asyncio.sleep(2.5)
        try:
            from integration_queue import IntegrationQueue
            from app_path import get_app_dir
            q = IntegrationQueue(get_app_dir())
            jid = getattr(self.page, "queue_current_job_id", "")
            if jid:
                q.set_status(jid, "failed" if failed else "done")
            nxt = q.next_pending()
            if nxt is None:
                self.page.queue_active = False
                self.page.queue_current_job_id = ""
                self._ui_log("✓ Queue finished — all sites processed.")
                return
            q.set_status(nxt.id, "running")
            self.page.queue_current_job_id = nxt.id
            self.page.integration_form = dict(nxt.form)
            self.page.integration_selected_steps_per_node = {
                k: set(v) for k, v in (nxt.selected_per_node or {}).items()}
            self.page.integration_selected_steps = {
                s for v in (nxt.selected_per_node or {}).values() for s in v}
            self._ui_log(f"▶ Queue: starting next site {nxt.shortcode}…")
            # We're already on /integration_run. A same-route go() is a
            # no-op — both the Flutter client router and our server-side
            # duplicate-route guard drop it — so the next site would never
            # get a fresh view. Bounce through /form first (a real route
            # change the client acts on), wait for that view to actually
            # mount, then launch the next site's run. /form doesn't touch
            # page.integration_form, so the handoff above survives.
            self.page.go("/form")
            for _ in range(30):
                await asyncio.sleep(0.1)
                top = self.page.views[-1] if self.page.views else None
                if top is not None and getattr(top, "route", None) == "/form":
                    break
            self.page.go("/integration_run")
        except Exception as exc:
            print(f"[queue] advance failed: {exc}")

    # ── Per-node resume ──────────────────────────────────────────
    def _update_resume_button(self, node_tag: str):
        btn = self._resume_buttons.get(node_tag)
        if btn is None:
            return
        btn.visible = (node_tag not in self._resuming
                       and node_tag in self._resumable_nodes)
        self._ui_dirty = True

    def _resume_node(self, node_tag: str):
        """Re-run a single node from its failed step. Runs concurrently;
        other nodes are untouched."""
        if not node_tag or node_tag in self._resuming:
            return
        params = self._node_params.get(node_tag)
        if not params:
            return
        self._resuming.add(node_tag)
        self._active_workers.add(node_tag)
        self._resumable_nodes.discard(node_tag)   # being re-run now
        self._update_resume_button(node_tag)      # hides it
        # Bring the run back to life for the resumed node.
        self.cancelled = False                    # resume un-cancels
        self._run_finished = False
        self.cancel_button.visible = False
        self.back_button.visible = False
        self.status_text.value = "Resuming node..."
        if not self._timer_running:
            self._timer_running = True
            threading.Thread(target=self._tick_timer, daemon=True).start()
        try:
            self.page.run_task(self._flush_loop)   # restart UI flusher
        except Exception:
            pass
        self._ui_log(f"[Resume] Re-running {params['node_name']} "
                     "from the failed step...")
        # Same wrapper as the initial run → finalize only when all idle.
        threading.Thread(target=self._node_worker,
                         args=(node_tag, params), daemon=True).start()

    # ── Spreadsheet-styled summary table ───────────────────────
    def _build_summary_table_flet(
        self, node_order: list, node_labels: dict,
    ) -> ft.Control:
        """Build a spreadsheet-style summary table matching the
        operator's ``PRE INTEGRATION CHECK LIST`` template:

          - Left column: step labels on bright-yellow background, bold
          - One column per node, white background, value text
          - Cell values: ``Yes`` (green) / ``No`` (red) / ``N/A`` (black)
            / ``pending`` (black)
          - Thin black borders between every cell

        Visual mirrors the screenshot the operator pasted; we reuse
        the existing INTEGRATION_STEPS list and the existing filter
        rules (no new rows, no new step keys)."""
        # Colour palette — matches the operator's reference template
        # (light blue row labels, conditional-format style cell fills).
        HEADER_BG = "#D9D9D9"        # light gray for top-left + node headers
        LABEL_BG = "#B6DDE8"         # light blue for row-label column
        YES_BG = "#C6EFCE"           # light green fill for "Yes"
        YES_FG = "#006100"           # dark green text for "Yes"
        NO_BG = "#FFC7CE"            # light pink fill for "No"
        NO_FG = "#9C0006"            # dark red text for "No"
        NA_BG = "#FFFF00"            # bright yellow fill for "N/A"
        NA_FG = "#000000"            # black text on yellow
        PENDING_BG = "#FFEB9C"       # muted yellow for "pending"
        PENDING_FG = "#9C5700"       # dark amber text for "pending"
        NEUTRAL_FG = "#000000"
        BORDER_C = "#000000"

        LABEL_COL_W = 340
        CELL_COL_W = 210
        ROW_H = 30

        def make_cell(
            text: str, bg: str, fg: str,
            weight=ft.FontWeight.NORMAL,
            width: int = CELL_COL_W,
            align_left: bool = False,
        ) -> ft.Container:
            return ft.Container(
                content=ft.Text(
                    text, color=fg, weight=weight, size=13,
                    text_align=ft.TextAlign.LEFT if align_left
                              else ft.TextAlign.CENTER,
                ),
                bgcolor=bg,
                width=width,
                height=ROW_H,
                alignment=ft.Alignment(-1.0 if align_left else 0, 0),
                border=ft.Border.all(0.5, BORDER_C),
                padding=ft.Padding(
                    left=10 if align_left else 6, top=0,
                    right=6, bottom=0,
                ),
            )

        rows: list = []

        # ── Header row ──────────────────────────────────────────
        header_cells = [
            make_cell(
                "", HEADER_BG, NEUTRAL_FG,
                weight=ft.FontWeight.BOLD,
                width=LABEL_COL_W, align_left=True,
            ),
        ]
        for ntag in node_order:
            header_cells.append(
                make_cell(
                    node_labels[ntag], HEADER_BG, NEUTRAL_FG,
                    weight=ft.FontWeight.BOLD,
                )
            )
        rows.append(ft.Row(header_cells, spacing=0, tight=True))

        # ── Step rows ───────────────────────────────────────────
        for key, label, applies_to, _ in INTEGRATION_STEPS:
            # Same filter as the prettytable path: REMARK + SUMMARY_NA
            # always show; everything else only if selected for at
            # least one node.
            if key not in REMARK_STEPS and key not in SUMMARY_NA_STEPS:
                if not any(
                    self.is_step_selected(key, nt)
                    for nt in node_order
                ):
                    continue

            # Label: same as the prettytable path (sgw_check varies
            # by node type composition).
            if key == "sgw_check":
                if len(node_order) == 1 and node_order[0] == "gsm":
                    row_label = "Verify BSC Broker IP Reachability"
                else:
                    row_label = "Verify SGw Reachability"
                if len(node_order) > 1:
                    row_label = "Verify SGw / BSC Broker Reachability"
            else:
                row_label = SUMMARY_LABELS.get(key, label)

            row_cells = [
                make_cell(
                    row_label, LABEL_BG, NEUTRAL_FG,
                    weight=ft.FontWeight.BOLD,
                    width=LABEL_COL_W, align_left=False,  # centered like screenshot
                ),
            ]

            for ntag in node_order:
                # Step doesn't apply to this node type → N/A (yellow).
                # Pass gsm_on_primary so GSM-scope steps render in the
                # LTE column when there is no separate GSM node.
                if not _step_applies_to_node(
                    applies_to, ntag,
                    gsm_on_primary=getattr(self, "gsm_on_primary", False),
                ):
                    row_cells.append(
                        make_cell("N/A", NA_BG, NA_FG)
                    )
                    continue
                # Operator didn't select for this node → pending (amber).
                if (key not in REMARK_STEPS
                        and key not in SUMMARY_NA_STEPS
                        and not self.is_step_selected(key, ntag)):
                    row_cells.append(
                        make_cell("pending", PENDING_BG, PENDING_FG)
                    )
                    continue
                # SUMMARY_NA rows always render N/A.
                if key in SUMMARY_NA_STEPS:
                    row_cells.append(
                        make_cell("N/A", NA_BG, NA_FG)
                    )
                    continue

                # Resolve run state — same derivations the existing
                # ``_summary_cell`` uses, but emit text not emoji.
                results = self._step_results.get(ntag, {})
                state = results.get(key, "skip")
                if key == "enrollment_sync":
                    state = results.get("enrollment", state)
                # Pre-verified manual rows: always show Yes
                if key in ("bb_transport", "sw_package_check", "bb_power_on"):
                    state = "done"

                if state == "done":
                    txt, bg, fg, w = "Yes", YES_BG, YES_FG, ft.FontWeight.BOLD
                elif state == "error":
                    txt, bg, fg, w = "No", NO_BG, NO_FG, ft.FontWeight.BOLD
                elif state == "skip":
                    txt, bg, fg, w = "N/A", NA_BG, NA_FG, ft.FontWeight.NORMAL
                else:
                    # "running" or anything unknown → pending
                    txt, bg, fg, w = "pending", PENDING_BG, PENDING_FG, ft.FontWeight.NORMAL

                row_cells.append(make_cell(txt, bg, fg, weight=w))

            rows.append(ft.Row(row_cells, spacing=0, tight=True))

        return ft.Column(rows, spacing=0, tight=True)

    def _build_summary_grid_data(
        self, node_order: list, node_labels: dict, full: bool = False,
    ) -> list[list[str]]:
        """Return the summary table as a list of rows (each row = list
        of plain strings). Reused by both the clipboard-copy helper
        (which joins with tabs for Excel paste) and the XLSX export.

        Layout matches the on-screen ``_build_summary_table_flet``
        exactly, so what the operator sees == what gets copied / saved.
        """
        rows: list[list[str]] = []

        # Header — trailing "Remark" column carries each step's short status
        # detail (e.g. the SGw check's "1/29 with packet loss").
        header = ["PRE INTEGRATION CHECK LIST"] + [
            node_labels[nt] for nt in node_order
        ] + ["Remark"]
        rows.append(header)

        for key, label, applies_to, _ in INTEGRATION_STEPS:
            # Visibility filter — same rule as the on-screen table. In ``full``
            # mode (the Excel export) every step is kept even if it wasn't run
            # this time, so the saved summary is always complete; the un-run
            # steps just read "pending".
            if not full and key not in REMARK_STEPS and key not in SUMMARY_NA_STEPS:
                if not any(
                    self.is_step_selected(key, nt) for nt in node_order
                ):
                    continue

            # Row label (sgw_check varies by node composition)
            if key == "sgw_check":
                if len(node_order) == 1 and node_order[0] == "gsm":
                    row_label = "Verify BSC Broker IP Reachability"
                else:
                    row_label = "Verify SGw Reachability"
                if len(node_order) > 1:
                    row_label = "Verify SGw / BSC Broker Reachability"
            else:
                row_label = SUMMARY_LABELS.get(key, label)

            row = [row_label]
            for ntag in node_order:
                if not _step_applies_to_node(
                    applies_to, ntag,
                    gsm_on_primary=getattr(self, "gsm_on_primary", False),
                ):
                    row.append("N/A")
                    continue
                if (key not in REMARK_STEPS
                        and key not in SUMMARY_NA_STEPS
                        and not self.is_step_selected(key, ntag)):
                    row.append("pending")
                    continue
                if key in SUMMARY_NA_STEPS:
                    row.append("N/A")
                    continue

                results = self._step_results.get(ntag, {})
                state = results.get(key, "skip")
                if key == "enrollment_sync":
                    state = results.get("enrollment", state)
                if key in ("bb_transport", "sw_package_check", "bb_power_on"):
                    state = "done"

                if state == "done":
                    cell = "Yes"
                elif state == "error":
                    cell = "No"
                elif state == "skip":
                    cell = "N/A"
                else:
                    cell = "pending"
                row.append(cell)

            # Remark column: the step's detail string, per node. Prefix with the
            # node label when more than one node so it's clear which is which.
            remarks = []
            for ntag in node_order:
                details = self._step_details.get(ntag, {})
                d = details.get(key, "")
                if not d and key == "enrollment_sync":
                    d = details.get("enrollment", "")
                if d:
                    remarks.append(
                        f"{node_labels[ntag]}: {d}" if len(node_order) > 1 else d)
            row.append("; ".join(remarks))
            rows.append(row)
        return rows

    # ── Copy summary to clipboard ──────────────────────────────
    # Flet's clipboard service is async (``await page.clipboard.set(s)``),
    # so we dispatch via ``page.run_task`` from the sync button handler.
    # The async core lives in ``_copy_summary_to_clipboard_async``.
    def _copy_summary_to_clipboard(
        self, node_order: list, node_labels: dict,
    ) -> None:
        try:
            self.page.run_task(
                self._copy_summary_to_clipboard_async,
                node_order, node_labels,
            )
        except Exception as exc:
            self._ui_log(f"✗ Copy dispatch failed: {exc}")

    async def _copy_summary_to_clipboard_async(
        self, node_order: list, node_labels: dict,
    ) -> None:
        """Build TSV (tab-separated values) and push to the OS
        clipboard via Flet's async ``page.clipboard.set()``. TSV is
        the format Excel/Sheets paste natively — tab = next column,
        newline = next row."""
        rows = self._build_summary_grid_data(node_order, node_labels)
        tsv = "\n".join("\t".join(r) for r in rows)

        # 1. Flet native clipboard — preferred path
        try:
            clipboard = getattr(self.page, "clipboard", None)
            if clipboard is not None and hasattr(clipboard, "set"):
                await clipboard.set(tsv)
                self._ui_log(
                    "✓ Summary copied to clipboard "
                    "(paste into Excel: Ctrl+V on cell A1)."
                )
                return
        except Exception as exc:
            self._ui_log(
                f"(Flet clipboard failed: {exc}; trying tkinter…)"
            )

        # 2. tkinter fallback — sync, off the UI thread to avoid hang.
        # On some Windows setups, creating a Tk root from inside the
        # main UI thread can lock up; run in a worker thread.
        def _tk_copy():
            try:
                import tkinter as _tk
                root = _tk.Tk()
                root.withdraw()
                root.clipboard_clear()
                root.clipboard_append(tsv)
                root.update()
                root.destroy()
                self._ui_log(
                    "✓ Summary copied to clipboard (tkinter fallback)."
                )
            except Exception as exc:
                self._ui_log(f"✗ Copy fallback failed: {exc}")
        try:
            self.page.run_thread(_tk_copy)
        except Exception as exc:
            self._ui_log(f"✗ Copy dispatch failed: {exc}")

    # ── Save summary to xlsx ───────────────────────────────────
    # Flet's modern ``FilePicker.save_file()`` is an async coroutine
    # that RETURNS the chosen path directly (no ``on_result`` callback
    # — that was the old API). Dispatch via ``page.run_task``.
    def _save_summary_xlsx(
        self, node_order: list, node_labels: dict,
    ) -> None:
        try:
            self.page.run_task(
                self._save_summary_xlsx_async,
                node_order, node_labels,
            )
        except Exception as exc:
            self._ui_log(f"✗ Save dispatch failed: {exc}")

    async def _save_summary_xlsx_async(
        self, node_order: list, node_labels: dict,
    ) -> None:
        """Pop the OS save dialog, await the chosen path, then write
        the styled xlsx via openpyxl.

        Flet 0.84+ note: ``FilePicker`` is a Service, not a Control.
        It must be attached to ``page.services``, not ``page.overlay``
        — otherwise the renderer complains "Unknown control: FilePicker"
        and the dialog never opens. The old overlay placement worked
        in pre-0.80 Flet but was deprecated in the same release that
        introduced the async ``save_file()`` return-value API.
        """
        picker = ft.FilePicker()
        attached_to = None
        try:
            services = getattr(self.page, "services", None)
            if services is not None:
                services.append(picker)
                attached_to = "services"
            else:
                # Fallback for older Flet where FilePicker WAS a Control
                self.page.overlay.append(picker)
                attached_to = "overlay"
            self.page.update()
        except Exception as exc:
            self._ui_log(f"✗ Could not mount file picker: {exc}")
            return

        default_name = (
            f"INTEGRATION_SUMMARY_"
            f"{re.sub(r'[^A-Za-z0-9._-]', '_', self.shortcode)}_"
            f"{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        path = None
        try:
            path = await picker.save_file(
                dialog_title="Save integration summary",
                file_name=default_name,
                file_type=ft.FilePickerFileType.CUSTOM,
                allowed_extensions=["xlsx"],
            )
        except Exception as exc:
            self._ui_log(f"✗ Save dialog failed: {exc}")

        # Remove the transient picker from wherever we put it.
        try:
            if attached_to == "services":
                services = getattr(self.page, "services", None)
                if services is not None and picker in services:
                    services.remove(picker)
            elif attached_to == "overlay":
                if picker in (self.page.overlay or []):
                    self.page.overlay.remove(picker)
            self.page.update()
        except Exception:
            pass

        if not path:
            # Operator cancelled — silent.
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            self._write_summary_xlsx(path, node_order, node_labels)
            self._ui_log(f"✓ Saved summary → {path}")
        except Exception as exc:
            self._ui_log(f"✗ Save failed: {exc}")

    def _style_summary_sheet(self, ws, rows: list) -> None:
        """Paint ``rows`` onto ``ws`` with the exact on-screen palette:
        gray header, light-blue bold labels, green ``Yes`` / red ``No`` /
        bright-yellow ``N/A`` / muted-yellow ``pending``, thin black borders."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        thin = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="D9D9D9")   # light gray
        label_fill = PatternFill("solid", fgColor="B6DDE8")    # light blue
        yes_fill = PatternFill("solid", fgColor="C6EFCE")      # light green
        no_fill = PatternFill("solid", fgColor="FFC7CE")       # light pink
        na_fill = PatternFill("solid", fgColor="FFFF00")       # bright yellow
        pending_fill = PatternFill("solid", fgColor="FFEB9C")  # muted yellow
        center = Alignment(horizontal="center", vertical="center")
        left_wrap = Alignment(horizontal="left", vertical="center",
                              wrap_text=True)
        remark_col = len(rows[0])   # last column is the free-text Remark

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = cell_border
                if r_idx == 1:
                    cell.font = Font(bold=True, color="000000")
                    cell.fill = header_fill
                    cell.alignment = center
                elif c_idx == 1:
                    cell.font = Font(bold=True, color="000000")
                    cell.fill = label_fill
                    cell.alignment = center
                elif c_idx == remark_col:
                    # Free-text remark — left-aligned, wrapped, no status fill.
                    cell.font = Font(color="000000")
                    cell.alignment = left_wrap
                else:
                    if value == "Yes":
                        cell.font = Font(bold=True, color="006100")
                        cell.fill = yes_fill
                    elif value == "No":
                        cell.font = Font(bold=True, color="9C0006")
                        cell.fill = no_fill
                    elif value == "N/A":
                        cell.font = Font(color="000000")
                        cell.fill = na_fill
                    elif value == "pending":
                        cell.font = Font(color="9C5700")
                        cell.fill = pending_fill
                    else:
                        cell.font = Font(color="000000")
                    cell.alignment = center

        ws.column_dimensions["A"].width = 40
        for c in range(2, len(rows[0]) + 1):
            ws.column_dimensions[get_column_letter(c)].width = (
                48 if c == remark_col else 26)
        for r in range(1, len(rows) + 1):
            ws.row_dimensions[r].height = 20

    def _write_summary_xlsx(
        self, path: str, node_order: list, node_labels: dict,
    ) -> None:
        """Manual 'Save as Excel' — one styled sheet (the on-screen view)."""
        from openpyxl import Workbook
        rows = self._build_summary_grid_data(node_order, node_labels)
        wb = Workbook()
        ws = wb.active
        ws.title = "Pre Integration Check List"
        self._style_summary_sheet(ws, rows)
        wb.save(path)

    def _autosave_summary_xlsx(
        self, node_order: list, node_labels: dict,
    ) -> None:
        """Auto-save every run's summary into one workbook per site,
        ``LOG/<siteid>/<siteid>_Integration_Summary.xlsx`` — each run appends a
        new timestamped sheet. The saved table is always the FULL check list
        (un-run steps read 'pending'), styled exactly like the on-screen view."""
        try:
            from openpyxl import Workbook, load_workbook
            from app_path import get_app_dir

            safe = re.sub(r"[^A-Za-z0-9._-]", "_", self.shortcode or "UNKNOWN")
            out_dir = os.path.join(get_app_dir(), "LOG", safe)
            os.makedirs(out_dir, exist_ok=True)
            path = os.path.join(out_dir, f"{safe}_Integration_Summary.xlsx")

            rows = self._build_summary_grid_data(
                node_order, node_labels, full=True)

            if os.path.isfile(path):
                wb = load_workbook(path)
            else:
                wb = Workbook()
                wb.remove(wb.active)   # start clean; every run is a named sheet

            base = time.strftime("Run_%Y%m%d_%H%M%S")
            name, i = base, 1
            while name in wb.sheetnames:      # avoid a same-second collision
                name = f"{base}_{i}"[:31]
                i += 1
            self._style_summary_sheet(wb.create_sheet(title=name), rows)
            wb.save(path)
            self._ui_log(f"✓ Summary auto-saved → {path} (sheet {name})")
        except PermissionError:
            self._ui_log("⚠ Summary auto-save skipped — the Excel file is open. "
                         "Close it to let the next run append.")
        except Exception as exc:
            self._ui_log(f"⚠ Summary auto-save failed: {exc}")

    async def _reshow_summary_task(self):
        """Re-open the Integration Summary after returning from CDD Audit —
        the view has just rebuilt, so wait a beat for it to mount, then show
        the popup from the restored (not re-run) results. No auto-save here —
        re-showing an existing summary must not append a duplicate sheet."""
        import asyncio
        await asyncio.sleep(0.4)
        try:
            self._show_summary_popup(save_excel=False)
        except Exception as exc:
            print(f"[summary] re-show failed: {exc}")

    def _show_summary_popup(self, save_excel: bool = True):
        # Persist just enough state on the page object so this summary can be
        # re-shown after a trip to the CDD Audit page (which rebuilds this view
        # from scratch) WITHOUT re-running the integration.
        self.page.integration_step_results = {
            k: dict(v) for k, v in self._step_results.items()}
        self.page.integration_step_details = {
            k: dict(v) for k, v in self._step_details.items()}
        self.page.integration_node_durations = dict(self._node_durations)
        self.page.integration_has_summary = True

        node_order = []
        node_labels = {}
        node_order.append("lte")
        node_labels["lte"] = self.lte_name or "LTE/NR"
        if self.has_lte2:
            node_order.append("lte2")
            node_labels["lte2"] = self.lte2_name or "LTE/NR #2"
        if self.has_gsm:
            node_order.append("gsm")
            node_labels["gsm"] = self.gsm_name or "GSM"

        # Auto-save this run's summary into the per-site workbook (a new sheet
        # each run). Skipped on a re-show so it never appends a duplicate.
        if save_excel:
            self._autosave_summary_xlsx(node_order, node_labels)

        # Build a prettytable with columns: Check List | Node1 | Node2 | ...
        try:
            from prettytable import PrettyTable
            use_prettytable = True
        except ImportError:
            use_prettytable = False

        if use_prettytable:
            col_names = ["Pre Integration Check List"] + [
                node_labels[ntag] for ntag in node_order
            ]
            table = PrettyTable(col_names)
            table.align = "l"
            table.align["Pre Integration Check List"] = "l"
            for ntag in node_order:
                table.align[node_labels[ntag]] = "c"

            for key, label, applies_to, _ in INTEGRATION_STEPS:
                # Filter rule for the summary table:
                #   * REMARK steps (bb_transport, enrollment_sync) always
                #     show — their value is derived from other steps.
                #   * SUMMARY_NA steps (bsc_neighbours, network_audit)
                #     always show as N/A — they're documentation rows.
                #   * Everything else only shows if the operator
                #     selected this step for AT LEAST ONE node in the
                #     multi-column checklist. Skipping un-selected
                #     rows keeps the summary aligned with what was
                #     actually attempted.
                if key not in REMARK_STEPS and key not in SUMMARY_NA_STEPS:
                    if not any(
                        self.is_step_selected(key, nt)
                        for nt in node_order
                    ):
                        continue

                row = []
                # Build label that varies by node type for sgw_check
                if key == "sgw_check":
                    # Show "Verify SGw Reachability" for LTE/NR,
                    # "Verify BSC Broker IP Reachability" for GSM
                    row_label = "Verify BSC Broker IP Reachability" if len(node_order) == 1 and node_order[0] == "gsm" else "Verify SGw Reachability"
                else:
                    row_label = SUMMARY_LABELS.get(key, label)

                # For multi-node tables, infer the row label from the
                # first applicable node type
                if key == "sgw_check" and len(node_order) > 1:
                    row_label = "Verify SGw / BSC Broker Reachability"

                row = [row_label]
                for ntag in node_order:
                    # Steps that don't apply to this node type
                    if applies_to not in ("both", "lte_nr" if ntag != "gsm" else "gsm"):
                        row.append("N/A")
                        continue
                    # Operator opted out of this step for THIS node →
                    # show "—" instead of forcing a misleading status.
                    if (key not in REMARK_STEPS
                            and key not in SUMMARY_NA_STEPS
                            and not self.is_step_selected(key, ntag)):
                        row.append("—")
                        continue
                    results = self._step_results.get(ntag, {})
                    row.append(self._summary_cell(key, ntag, applies_to, results))
                table.add_row(row)

            # Add summary footer: total time and overall status
            for ntag in node_order:
                nname = node_labels[ntag]
                results = self._step_results.get(ntag, {})
                done_count = sum(1 for v in results.values() if v == "done")
                total = len(self.selected_steps)
                dur = self._node_durations.get(ntag, 0)
                mins, secs = divmod(int(dur), 60)

            summary_text = table.get_string()

            # Add per-node status lines below the table
            for ntag in node_order:
                nname = node_labels[ntag]
                results = self._step_results.get(ntag, {})
                done_count = sum(1 for v in results.values() if v == "done")
                # Total = steps selected for THIS node specifically
                total_steps = len([
                    k for k, _, a, _ in INTEGRATION_STEPS
                    if self.is_step_selected(k, ntag)
                    and a in ("both", "lte_nr" if ntag != "gsm" else "gsm")
                ])
                dur = self._node_durations.get(ntag, 0)
                mins, secs = divmod(int(dur), 60)

                any_error = any(v == "error" for v in results.values())
                any_skip = any(v == "skip" for v in results.values())

                if done_count == total_steps and not any_error:
                    status = "Done"
                elif any_error:
                    status = "Completed with errors"
                else:
                    status = f"{done_count}/{total_steps} steps"
                summary_text += f"\n{nname} {status}"
                summary_text += f"\ntime {mins:02d}:{secs:02d}"
        else:
            lines = []
            for ntag in node_order:
                nname = node_labels[ntag]
                results = self._step_results.get(ntag, {})
                if len(node_order) > 1:
                    lines.append(f"── {nname} ──")
                for key, label, applies_to, _ in INTEGRATION_STEPS:
                    if applies_to not in ("both", "lte_nr" if ntag != "gsm" else "gsm"):
                        continue
                    # Per-node filter: REMARK + SUMMARY_NA always show;
                    # everything else only if selected for THIS node.
                    if (key not in REMARK_STEPS
                            and key not in SUMMARY_NA_STEPS
                            and not self.is_step_selected(key, ntag)):
                        continue
                    if key == "sgw_check":
                        summary_name = (
                            "Verify BSC Broker IP Reachability"
                            if ntag == "gsm"
                            else "Verify SGw Reachability"
                        )
                    else:
                        summary_name = SUMMARY_LABELS.get(key, label)
                    cell = self._summary_cell(key, ntag, applies_to, results)
                    lines.append(f"{summary_name}\t{cell}")
                dur = self._node_durations.get(ntag, 0)
                mins, secs = divmod(int(dur), 60)
                lines.append(f"⏱ Total Time\t{mins:02d}:{secs:02d}")
                lines.append("")
            summary_text = "\n".join(lines).strip()

        close_event = threading.Event()

        def _on_close(e):
            self._close_dialog_safe(dlg)
            close_event.set()

        def _on_next_audit(e):
            # Go to the CDD Audit page; the audit page shows a "Back to
            # Summary" button that returns here (see integration_show_summary
            # _only handling in build()).
            self._close_dialog_safe(dlg)
            close_event.set()
            self.page.go("/audit")

        window = getattr(self.page, "window", None)
        window_width = getattr(window, "width", None) or 1380
        window_height = getattr(window, "height", None) or 920
        dialog_width = max(920, min(int(window_width * 0.92), 1320))
        dialog_height = max(520, min(int(window_height * 0.82), 720))

        # Spreadsheet-style table (per the operator's reference image).
        # Replaces the previous monospace text-blob rendering.
        summary_table = self._build_summary_table_flet(node_order, node_labels)

        # Per-node footer: total time + completion ratio. Plain text
        # under the table, no styling beyond muted colour.
        footer_lines: list[str] = []
        for ntag in node_order:
            nname = node_labels[ntag]
            results = self._step_results.get(ntag, {})
            done_count = sum(1 for v in results.values() if v == "done")
            total_steps = len([
                k for k, _, a, _ in INTEGRATION_STEPS
                if self.is_step_selected(k, ntag)
                and a in ("both", "lte_nr" if ntag != "gsm" else "gsm")
            ])
            dur = self._node_durations.get(ntag, 0)
            mins, secs = divmod(int(dur), 60)
            any_error = any(v == "error" for v in results.values())
            if done_count == total_steps and not any_error:
                status = "Done"
            elif any_error:
                status = "Completed with errors"
            else:
                status = f"{done_count}/{total_steps} steps"
            footer_lines.append(
                f"{nname} — {status}    ⏱ {mins:02d}:{secs:02d}"
            )

        footer_text = ft.Text(
            "\n".join(footer_lines),
            size=12,
            color=TEXT_MUTED,
            selectable=True,
            font_family="Consolas",
        )

        summary_panel = ft.Container(
            content=ft.Column(
                [
                    # Horizontal scroll wrapper for the table — keeps
                    # wide multi-node layouts usable on small windows.
                    ft.Row(
                        [summary_table],
                        scroll=ft.ScrollMode.AUTO,
                        vertical_alignment=ft.CrossAxisAlignment.START,
                    ),
                    ft.Container(height=12),
                    footer_text,
                ],
                spacing=4,
                scroll=ft.ScrollMode.AUTO,
                expand=True,
            ),
            bgcolor=ft.Colors.with_opacity(0.15, PANEL),
            border=ft.Border.all(1, BORDER),
            border_radius=14,
            padding=12,
            width=dialog_width,
            height=dialog_height,
        )

        # Capture node_order / labels in closures so the action
        # buttons don't need to recompute them.
        _node_order_snap = list(node_order)
        _node_labels_snap = dict(node_labels)

        def _on_copy(e):
            self._copy_summary_to_clipboard(
                _node_order_snap, _node_labels_snap,
            )

        def _on_save_xlsx(e):
            self._save_summary_xlsx(
                _node_order_snap, _node_labels_snap,
            )

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("Integration Summary", color=SUCCESS,
                          weight=ft.FontWeight.BOLD),
            content=summary_panel,
            actions=[
                # Copy to clipboard (TSV) — paste into Excel/Sheets
                # directly with Ctrl+V on cell A1. Each tab becomes a
                # column, each newline a new row.
                ft.TextButton(
                    "Copy to Clipboard",
                    icon=ft.Icons.CONTENT_COPY,
                    tooltip=(
                        "Copy table as tab-separated values. "
                        "Paste into Excel/Google Sheets with Ctrl+V."
                    ),
                    style=ft.ButtonStyle(color=ACCENT),
                    on_click=_on_copy,
                ),
                # Save as .xlsx with the same styling as the on-screen
                # table (yellow row labels, green Yes / red No)
                ft.TextButton(
                    "Save as Excel",
                    icon=ft.Icons.SAVE_OUTLINED,
                    tooltip=(
                        "Save as a styled .xlsx file (same layout "
                        "and colours as the on-screen table)."
                    ),
                    style=ft.ButtonStyle(color=ACCENT),
                    on_click=_on_save_xlsx,
                ),
                # Continue to the CDD Audit page (returns to this summary).
                ft.TextButton(
                    "Next → CDD Audit",
                    icon=ft.Icons.FACT_CHECK,
                    tooltip="Open the CDD Audit page. You can return to this "
                            "summary with the Back button there.",
                    style=ft.ButtonStyle(color=SUCCESS),
                    on_click=_on_next_audit,
                ),
                ft.TextButton("Close",
                              style=ft.ButtonStyle(color=TEXT_MUTED),
                              on_click=_on_close),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            bgcolor=PANEL,
        )

        self._show_dialog_safe(dlg)

    def _run_node_steps(
        self,
        node_tag: str,
        node_name: str,
        node_ip: str,
        subnetwork: str,
        node_type: str,
    ):
        """Run selected integration steps for one node."""
        import time
        node_start = time.monotonic()
        from integration_runner import (
            IntegrationSSH, run_create_arne, run_enrollment,
            run_install_lkf, run_baseline, run_relation, run_verify_mme,
            run_backup_cv, run_take_dump, run_gsm_cell_define,
            run_take_cm_dump,
            run_uri_setting, run_pm_measurement, run_sgw_check,
            run_external_alarm, run_bsc_neighbours, run_sync_check,
            run_sw_check,
        )

        label = "LTE/NR" if node_tag == "lte" else ("LTE/NR #2" if node_tag == "lte2" else "GSM")
        host = self.form.get("host", "")
        port = self.form.get("port", 5023)
        username = self.form.get("username", "")
        password = self.form.get("password", "")

        def detail_cb(msg: str):
            """Full detail — saved to log file only."""
            self._detail_log(f"[{label}] {msg}", node_tag=node_tag)

        def ui_cb(msg: str):
            """High-level — shown in UI log panel."""
            self._ui_log(f"[{label}] {msg}")

        # Node-bound retry dialog: carries this node's tag so the
        # "Skip Node" button can abort only this node's remaining steps.
        def node_wait_for_user(msg: str) -> bool:
            return self._ask_user_retry(msg, node_tag=node_tag)

        # Connect SSH
        try:
            ssh = IntegrationSSH(
                host=host, port=port,
                username=username, password=password,
                log_callback=detail_cb,
            )
            # Mirror the raw moshell/SSH stream into this node's live
            # terminal tab (every byte the node sends back).
            ssh.set_live_sink(
                lambda chunk, _nt=node_tag: self._feed_raw(_nt, chunk))
            ssh.connect(timeout=30)
            self._active_ssh[node_tag] = ssh
            ui_cb("SSH connected.")
        except Exception as exc:
            ui_cb(f"SSH connection failed: {exc}")
            for key, _, applies_to, _ in INTEGRATION_STEPS:
                if applies_to in ("both", node_type) and \
                        self.is_step_selected(key, node_tag):
                    self._set_step(node_tag, key, "error", "SSH failed")
            return

        # Steps that need an active AMOS session
        amos_steps = {
            "enrollment", "install_lkf", "baseline", "ret_scripts",
            "relation", "uri_setting", "verify_mme", "sgw_check",
            "backup_cv", "take_dump", "take_cm_dump", "gsm_cell_define",
            "pm_measurement", "external_alarm", "bsc_neighbours",
        }
        custom_moshell_log_steps = {"relation", "sgw_check"}
        in_amos = False

        # A resume re-enters this method for the same node — clear any
        # prior abort mark so it isn't skipped, and hide the button.
        self._aborted_nodes.discard(node_tag)

        try:
            step_num = 1
            for key, step_label, applies_to, log_suffix in INTEGRATION_STEPS:
                if self.cancelled:
                    break
                # Scope filter: the step's applies_to must include
                # this node. The module-level helper handles "both",
                # "lte_nr", "lte_primary" (lte only) and "gsm". With
                # ``gsm_on_primary=True`` (single LTE node hosting
                # co-located GSM via BSC), gsm-scope steps also run
                # on the lte node — that's where they belong.
                if not _step_applies_to_node(
                    applies_to, node_tag,
                    gsm_on_primary=self.gsm_on_primary,
                ):
                    continue

                # Skip if the operator didn't select this step for THIS
                # node in the multi-column checklist.
                if not self.is_step_selected(key, node_tag):
                    continue

                # Skip summary-only steps: they have no runner logic.
                if key in REMARK_STEPS or key in SUMMARY_NA_STEPS:
                    continue

                # On a resume, don't re-run steps that already succeeded —
                # pick up from the failed/pending ones.
                if self._step_results.get(node_tag, {}).get(key) == "done":
                    continue

                # Enter AMOS lazily before the first step that needs it
                if not in_amos and key in amos_steps:
                    ui_cb(f"Entering AMOS for {node_name}...")
                    try:
                        ssh.enter_amos(node_name, timeout=90)
                        in_amos = True
                        ui_cb("AMOS session ready.")
                    except Exception as exc:
                        ui_cb(f"Failed to enter AMOS: {exc}")
                        self._resumable_nodes.add(node_tag)
                        # Mark remaining selected steps (for THIS node)
                        # as error
                        remaining = False
                        for rk, _, ra, _ in INTEGRATION_STEPS:
                            if ra not in ("both", node_type):
                                continue
                            if not self.is_step_selected(rk, node_tag):
                                continue
                            if rk == key:
                                remaining = True
                            if remaining:
                                self._set_step(node_tag, rk, "error",
                                               "AMOS failed")
                        break

                self._set_step(node_tag, key, "running")
                ui_cb(f"Running {step_label}...")

                # Start live session log for this step. Every byte read
                # from the SSH shell (prompts, [Proxy ID] lines, crn blocks,
                # curl output, everything) is teed to this file.
                session_dir = os.path.join(self.log_dir, "SESSION")
                os.makedirs(session_dir, exist_ok=True)
                session_log_path = os.path.join(
                    session_dir,
                    f"SESSION_{key.upper()}_{node_name}.log",
                )
                try:
                    ssh.start_step_log(session_log_path)
                except Exception as _exc:
                    ui_cb(f"(could not start session log: {_exc})")

                remote_step_log_path = None
                if key in amos_steps and key not in custom_moshell_log_steps:
                    safe_node = (
                        node_name.replace("/", "_")
                        .replace("\\", "_")
                        .replace(" ", "_")
                    )
                    remote_step_log_path = (
                        f"/home/shared/{ssh.username}/{key.upper()}_{safe_node}.log"
                    )
                    try:
                        ssh.run_amos_command_safe(
                            f"!rm -f {remote_step_log_path}",
                            node_name,
                            timeout=15,
                        )
                        ssh.run_amos_command_safe(
                            f"l+ {remote_step_log_path}",
                            node_name,
                            timeout=15,
                        )
                    except Exception as _exc:
                        remote_step_log_path = None
                        ui_cb(f"(could not start MOSHELL capture: {_exc})")

                stopped = False
                try:
                    if key == "create_arne":
                        # Spec: "Add Node in ENM" is a hard gate — if it
                        # fails, STOP the node (don't run enrollment etc.
                        # against a node that isn't in ENM). The node
                        # becomes resumable (Resume button). Retry via the
                        # runner's internal ``wait_for_user`` re-checks
                        # ``cmedit get`` after the user clicks Retry.
                        #
                        # ``bsc_name`` triggers the BSC link step inside
                        # ``run_create_arne``. Pass it when:
                        #   * node is the separate GSM node, OR
                        #   * node is the primary LTE AND we're in
                        #     co-located mode (no separate GSM, just a
                        #     BSC name in the form) — single-radio
                        #     multi-RAT site, controllingBsc lives on
                        #     the primary node's NetworkElement.
                        bsc_for_create = None
                        if node_tag == "gsm":
                            bsc_for_create = self.form.get("bsc_name", "")
                        elif node_tag == "lte" and self.gsm_on_primary:
                            bsc_for_create = self.form.get("bsc_name", "")
                        # Diagnostic so the field can see WHY the BSC
                        # link step ran or was skipped.
                        ui_cb(
                            f"create_arne: node_tag={node_tag}, "
                            f"gsm_on_primary={self.gsm_on_primary}, "
                            f"bsc_name={bsc_for_create!r}"
                        )
                        success, output = run_create_arne(
                            ssh, node_name, node_ip, subnetwork, detail_cb,
                            wait_for_user=node_wait_for_user,
                            bsc_name=bsc_for_create,
                            log_dir=self.log_dir,
                        )
                        if success:
                            self._set_step(node_tag, key, "done", "Verified")
                            ui_cb(f"{step_label} — verified.")
                        else:
                            self._set_step(node_tag, key, "error",
                                           "Failed — node stopped")
                            ui_cb(f"{step_label} — failed; stopping node "
                                  "(Resume available).")
                            stopped = True

                    elif key == "enrollment":
                        # Spec: enrollment STOPS the workflow on failure
                        # (no point running URI/LKF/baseline against an
                        # un-enrolled node). Internal retries on
                        # credential-status and sync checks are handled
                        # inside ``run_enrollment`` (2 attempts, ~2 min
                        # apart).
                        success, output = run_enrollment(
                            ssh, node_name, detail_cb,
                            wait_for_user=node_wait_for_user,
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "SYNCHRONIZED")
                            self._set_step(node_tag, "enrollment_sync",
                                           "done", "Auto-synced")
                            ui_cb(f"{step_label} — SYNCHRONIZED.")
                        else:
                            self._set_step(node_tag, key, "error",
                                           "Stopped by user")
                            self._set_step(node_tag, "enrollment_sync",
                                           "error", "Enrollment failed")
                            ui_cb(f"{step_label} — stopped by user.")
                            stopped = True

                    elif key == "install_lkf":
                        # LKF always runs — with or without a zip file.
                        #
                        # WITH zip:
                        #   auto-retry 3x, then one operator-driven retry,
                        #   then continue regardless. ``lkfinstall.py``
                        #   submits an ENM job; the previous job either
                        #   COMPLETED (retry = no-op verify) or FAILED
                        #   (re-submit is what we want).
                        #
                        # WITHOUT zip:
                        #   run lkfinstall.py <node> directly (the LKF may
                        #   already be imported in ENM from a batch). If
                        #   it fails, DON'T prompt — just mark
                        #   "LKF not available" and continue to the next
                        #   step. Single attempt, no retry.
                        lkf_file = self.form.get("lkf_file", "")
                        if not lkf_file:
                            # No zip → direct install attempt, non-blocking.
                            ui_cb(
                                f"{step_label} — no zip file; running "
                                "lkfinstall.py directly."
                            )
                            success, output = run_install_lkf(
                                ssh, node_name, "", detail_cb,
                                wait_for_user=None,
                            )
                            if success:
                                self._set_step(node_tag, key, "done",
                                               "COMPLETED (no zip)")
                                ui_cb(f"{step_label} — completed (no zip).")
                            else:
                                self._set_step(node_tag, key, "error",
                                               "LKF not available")
                                ui_cb(
                                    f"{step_label} — LKF not available "
                                    "(install/status failed); continuing "
                                    "to next step."
                                )
                        else:
                            success, output = self._retry_step(
                                step_label,
                                lambda: run_install_lkf(
                                    ssh, node_name, lkf_file, detail_cb,
                                    wait_for_user=None,
                                ),
                                node_tag, key,
                                max_attempts=3,
                            )
                            # If all 3 auto-attempts failed, give the
                            # operator one user-driven retry before
                            # giving up — but still continue to the next
                            # step regardless of their choice.
                            if not success:
                                ui_cb(
                                    f"{step_label} — 3 auto-attempts "
                                    "failed; asking operator."
                                )
                                self._set_step(
                                    node_tag, key, "running",
                                    "manual retry?",
                                )
                                success, output = run_install_lkf(
                                    ssh, node_name, lkf_file, detail_cb,
                                    wait_for_user=node_wait_for_user,
                                )
                            if success:
                                self._set_step(node_tag, key, "done",
                                               "COMPLETED")
                                ui_cb(f"{step_label} — completed.")
                            else:
                                self._set_step(node_tag, key, "error",
                                               "Failed (continued)")
                                ui_cb(
                                    f"{step_label} — failed, continuing "
                                    "to next step."
                                )

                    elif key == "baseline":
                        success, output = run_baseline(
                            ssh, node_name, detail_cb,
                            wait_for_user=None,
                            confirm_baseline=self._ask_user_confirm,
                        )
                        if success:
                            self._set_step(node_tag, key, "done", "Verified")
                            ui_cb(f"{step_label} — verified.")
                        else:
                            self._set_step(node_tag, key, "error", "Failed (continued)")
                            ui_cb(f"{step_label} — failed, continuing to next step.")

                    elif key == "relation":
                        # Spec: continue on fail, no auto-retry — but if
                        # any per-file error is detected, prompt the user
                        # to check the log first, then continue (the
                        # runner exposes this via wait_for_user).
                        relation_file = self.form.get("relation_file", "")
                        if not relation_file:
                            self._set_step(node_tag, key, "skip",
                                           "No relation file")
                            ui_cb(f"{step_label} — skipped (no file).")
                        else:
                            success, output = run_relation(
                                ssh, node_name, self.shortcode,
                                relation_file, self.log_dir, detail_cb,
                                wait_for_user=node_wait_for_user,
                                ui_cb=ui_cb,  # live per-script progress
                            )
                            if success:
                                self._set_step(node_tag, key, "done",
                                               "Completed")
                                ui_cb(f"{step_label} — completed.")
                            else:
                                self._set_step(node_tag, key, "error",
                                               "Failed (continued)")
                                ui_cb(f"{step_label} — failed, continuing to next step.")

                    elif key == "ret_scripts":
                        self._set_step(node_tag, key, "skip", "Coming soon")
                        ui_cb(f"{step_label} — skipped (coming soon).")

                    elif key == "uri_setting":
                        success, output = self._retry_step(
                            step_label,
                            lambda: run_uri_setting(
                                ssh, node_name, username, password,
                                detail_cb, wait_for_user=None,
                            ),
                            node_tag, key,
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "SUCCESS")
                            ui_cb(f"{step_label} — SUCCESS.")
                        else:
                            self._set_step(node_tag, key, "error",
                                           "Failed (continued)")
                            ui_cb(f"{step_label} — failed, continuing to next step.")

                    elif key == "sw_level_check":
                        # Compare active UpgradePackage vs config.json
                        # (uri_setting.upgrade_package_id). Report-only.
                        success, output, detail = run_sw_check(
                            ssh, node_name, detail_cb,
                            wait_for_user=None,
                        )
                        if success:
                            self._set_step(node_tag, key, "done", detail)
                            ui_cb(f"{step_label} — {detail}.")
                        else:
                            self._set_step(node_tag, key, "error", detail)
                            ui_cb(
                                f"{step_label} — {detail}; continuing "
                                "to next step."
                            )

                    elif key == "verify_mme":
                        # Spec: no internal retry — single attempt.
                        success, output = run_verify_mme(
                            ssh, node_name, detail_cb,
                            wait_for_user=None,
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "All ENABLED")
                            ui_cb(f"{step_label} — all ENABLED.")
                        else:
                            self._set_step(node_tag, key, "error",
                                           "DISABLED (continued)")
                            ui_cb(f"{step_label} — DISABLED found, continuing.")

                    elif key == "sgw_check":
                        # Spec: no internal retry — single attempt.
                        # ``run_sgw_check`` returns a 3-tuple with a
                        # short detail string so the progress / summary
                        # cell can show the 4-level status:
                        #   all_ok       → "All N pings OK"           (done)
                        #   partial_loss → "X/N with packet loss"      (done w/ warning)
                        #   some_failed  → "X/N failed"                (error)
                        #   all_failed   → "All N pings failed"        (error)
                        #
                        # ``gsm_on_primary`` makes the runner also
                        # execute the GSM ping script when a single
                        # LTE/NR node hosts co-located GSM (no
                        # separate GSM DN).
                        # BSC broker validation only applies when the GSM
                        # ping script runs (gsm node, or co-located gsm on
                        # the primary LTE node).
                        sgw_bsc = None
                        if node_tag == "gsm" or (
                                node_tag == "lte" and self.gsm_on_primary):
                            sgw_bsc = self.form.get("bsc_name", "")
                        result = run_sgw_check(
                            ssh, node_name, detail_cb,
                            wait_for_user=None,
                            node_type=("gsm" if node_tag == "gsm" else "lte_nr"),
                            gsm_on_primary=(
                                self.gsm_on_primary
                                and node_tag == "lte"
                            ),
                            bsc_name=sgw_bsc,
                        )
                        # Backward-compat: runner has returned 2-, 3- and
                        # now 4-tuples across versions.
                        broker_wrong = False
                        if len(result) == 4:
                            success, output, detail, broker_wrong = result
                        elif len(result) == 3:
                            success, output, detail = result
                        else:
                            success, output = result
                            detail = (
                                "All pings OK" if success
                                else "Ping failed"
                            )
                        if broker_wrong and success:
                            # Ping fine but the node points at the WRONG
                            # BSC broker → yellow warning.
                            self._set_step(node_tag, key, "warn", detail)
                            ui_cb(f"{step_label} — ⚠ {detail}.")
                        elif success:
                            self._set_step(node_tag, key, "done", detail)
                            ui_cb(f"{step_label} — {detail}.")
                        else:
                            self._set_step(node_tag, key, "error", detail)
                            ui_cb(
                                f"{step_label} — {detail}; continuing "
                                "to next step."
                            )

                    elif key == "gsm_cell_define":
                        # This is the GSM check — per spec, every GSM
                        # check first ensures controllingBsc is set.
                        # Pass the BSC name (always available for GSM /
                        # co-located nodes) + log_dir for the dedicated
                        # controllingBsc trace file.
                        gsm_bsc = self.form.get("bsc_name", "")
                        ui_cb(
                            f"gsm_cell_define: node_tag={node_tag}, "
                            f"bsc_name={gsm_bsc!r}"
                        )
                        success, output = run_gsm_cell_define(
                            ssh, node_name, self.shortcode, detail_cb,
                            wait_for_user=None,  # no prompt — just report
                            bsc_name=gsm_bsc,
                            log_dir=self.log_dir,
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "Cell & MO OK")
                            ui_cb(f"{step_label} — Cell & MO OK.")
                        else:
                            self._set_step(node_tag, key, "error",
                                           "0 instances (continued)")
                            ui_cb(f"{step_label} — 0 instances, continuing.")

                    elif key == "bsc_neighbours":
                        # Verify GSM neighbour relations exist in BSC:
                        #   gerancellrelation + externalgerancellrelation
                        # keyed on the modified shortcode (M<digits>).
                        success, output = run_bsc_neighbours(
                            ssh, node_name, self.shortcode, detail_cb,
                            wait_for_user=None,  # no prompt — just report
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "Relations OK")
                            ui_cb(f"{step_label} — relations OK.")
                        else:
                            self._set_step(node_tag, key, "error",
                                           "0 instances (continued)")
                            ui_cb(f"{step_label} — 0 instances, continuing.")

                    elif key == "sync_check":
                        # Synchronization (GPS/PTP) via moshell `sts`.
                        # OK when radioClockState contains "LOCKED".
                        # Detail e.g. "GPS - OK (RNT_TIME_LOCKED)".
                        success, output, detail = run_sync_check(
                            ssh, node_name, detail_cb,
                            wait_for_user=None,  # report-only, no prompt
                        )
                        if success:
                            self._set_step(node_tag, key, "done", detail)
                            ui_cb(f"{step_label} — {detail}.")
                        else:
                            self._set_step(node_tag, key, "error", detail)
                            ui_cb(
                                f"{step_label} — {detail}; continuing "
                                "to next step."
                            )

                    elif key == "backup_cv":
                        # Workaround: if URI Reconfig FAILED earlier, give
                        # it one more shot right before the backup — the
                        # config upload to ENM depends on the URI being
                        # set. Failure here doesn't block the backup.
                        if (self._step_results.get(node_tag, {})
                                .get("uri_setting") == "error"):
                            ui_cb(
                                "URI Reconfig failed earlier — retrying "
                                "once before Configuration Backup..."
                            )
                            self._set_step(node_tag, "uri_setting",
                                           "running", "retry before backup")
                            try:
                                uri_ok, _uri_out = run_uri_setting(
                                    ssh, node_name, username, password,
                                    detail_cb, wait_for_user=None,
                                )
                            except Exception as _uri_exc:
                                uri_ok = False
                                detail_cb(f"URI retry crashed: {_uri_exc}")
                            if uri_ok:
                                self._set_step(node_tag, "uri_setting",
                                               "done", "SUCCESS (retried)")
                                ui_cb("URI Reconfig retry — SUCCESS.")
                            else:
                                self._set_step(node_tag, "uri_setting",
                                               "error", "Failed (retried)")
                                ui_cb(
                                    "URI Reconfig retry failed — "
                                    "continuing to backup anyway."
                                )
                        # Spec: 2 attempts.
                        success, output = self._retry_step(
                            step_label,
                            lambda: run_backup_cv(
                                ssh, node_name, detail_cb,
                                wait_for_user=None,
                            ),
                            node_tag, key,
                            max_attempts=2,
                        )
                        if success:
                            self._set_step(node_tag, key, "done", "SUCCESS")
                            ui_cb(f"{step_label} â€” SUCCESS.")
                        else:
                            self._set_step(node_tag, key, "error", "Failed (continued)")
                            ui_cb(f"{step_label} â€” failed, continuing to next step.")

                    elif key == "take_dump":
                        # Spec: 2 attempts.
                        success, output = self._retry_step(
                            step_label,
                            lambda: run_take_dump(
                                ssh, node_name, self.shortcode,
                                self.log_dir, detail_cb,
                                wait_for_user=None,
                            ),
                            node_tag, key,
                            max_attempts=2,
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "Downloaded")
                            ui_cb(f"{step_label} — downloaded.")
                        else:
                            self._set_step(node_tag, key, "error", "Failed (continued)")
                            ui_cb(f"{step_label} — failed, continuing to next step.")

                    elif key == "take_cm_dump":
                        # Spec: 2 attempts.
                        success, output = self._retry_step(
                            step_label,
                            lambda: run_take_cm_dump(
                                ssh, node_name, self.shortcode,
                                self.log_dir, detail_cb,
                                wait_for_user=None,
                            ),
                            node_tag, key,
                            max_attempts=2,
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "Downloaded")
                            ui_cb(f"{step_label} — downloaded.")
                        else:
                            self._set_step(node_tag, key, "error", "Failed (continued)")
                            ui_cb(f"{step_label} — failed, continuing to next step.")

                    elif key == "pm_measurement":
                        # Spec: 3 attempts, 2-minute pause between retries.
                        pm_type = "gsm" if node_tag == "gsm" else "lte_nr"
                        success, output = self._retry_step(
                            step_label,
                            lambda: run_pm_measurement(
                                ssh, node_name, pm_type, detail_cb,
                                wait_for_user=None,
                            ),
                            node_tag, key,
                            max_attempts=3, backoff=120.0,
                        )
                        if success:
                            self._set_step(node_tag, key, "done", "Active")
                            ui_cb(f"{step_label} — PM active.")
                        else:
                            self._set_step(node_tag, key, "error", "Failed (continued)")
                            ui_cb(f"{step_label} — failed, continuing to next step.")

                    elif key == "external_alarm":
                        # applies_to="lte_primary" — the run-loop scope
                        # filter above already prevents this from
                        # firing on lte2/gsm, so we don't double-check
                        # here. Verify-only retry via user prompt.
                        success, output = run_external_alarm(
                            ssh, node_name, detail_cb,
                            wait_for_user=node_wait_for_user,
                        )
                        if success:
                            self._set_step(node_tag, key, "done",
                                           "8 AlarmPort MOs")
                            ui_cb(f"{step_label} — verified (8 MOs).")
                        else:
                            self._set_step(node_tag, key, "error",
                                           "Failed (continued)")
                            ui_cb(
                                f"{step_label} — failed, continuing to "
                                "next step."
                            )

                    else:
                        time.sleep(0.3)
                        self._set_step(node_tag, key, "skip",
                                       "Not implemented yet")
                        ui_cb(f"{step_label} — not yet implemented.")

                except Exception as exc:
                    self._set_step(node_tag, key, "error",
                                   str(exc)[:40])
                    ui_cb(f"{step_label} — FAILED: {exc}")
                    # Per user spec: a failed step never stops the
                    # workflow. Every step continues to the next so the
                    # operator gets a full status matrix at the end
                    # rather than a half-finished run.
                    ui_cb(f"{step_label} — continuing to next step despite error.")
                finally:
                    if remote_step_log_path:
                        try:
                            ssh.run_amos_command_safe("l-", node_name, timeout=15)
                            ssh.register_remote_log(remote_step_log_path)
                        except Exception as _exc:
                            ui_cb(f"(could not finalize MOSHELL capture: {_exc})")
                    # Close the session log for this step (always, even on error)
                    try:
                        ssh.stop_step_log()
                    except Exception:
                        pass
                    # Download any server-side moshell logs registered
                    # during this step into LOG/{SHORTCODE}/MOSHELL/
                    try:
                        moshell_dir = os.path.join(self.log_dir, "MOSHELL")
                        downloaded = ssh.drain_remote_logs(moshell_dir)
                        if downloaded:
                            ui_cb(
                                f"Downloaded {len(downloaded)} moshell log "
                                f"file(s) to MOSHELL/"
                            )
                            if key == "relation":
                                try:
                                    # Serialize CPU-heavy Excel build so
                                    # parallel nodes don't all peg the GIL
                                    # at once (black-screen prevention).
                                    with self._heavy_lock:
                                        parsed_path = build_relation_log_excel(
                                            downloaded,
                                            self.log_dir,
                                            node_name,
                                            log_cb=ui_cb,
                                        )
                                    ui_cb(
                                        "Relation log Excel created: "
                                        f"{os.path.basename(parsed_path)}"
                                    )
                                except Exception as parse_exc:
                                    ui_cb(
                                        f"(relation log parsing skipped: {parse_exc})"
                                    )
                                # Clean up the SHARED remote RELATION
                                # folder only when EVERY relation-running
                                # node is done — otherwise this node
                                # would delete a still-running sibling's
                                # scripts. Ref-counted via
                                # _relation_nodes_expected/_done.
                                try:
                                    remote_rel = (
                                        f"/home/shared/{ssh.username}"
                                        f"/RELATION/{self.shortcode}"
                                    )
                                    do_delete = False
                                    with self._relation_cleanup_lock:
                                        if self._relation_nodes_expected is None:
                                            # Compute once: which nodes will
                                            # run relation (lte_nr + selected
                                            # + a relation file is present).
                                            rel_file = self.form.get(
                                                "relation_file", "")
                                            exp = set()
                                            if rel_file:
                                                for nt, exists in (
                                                    ("lte", True),
                                                    ("lte2", self.has_lte2),
                                                ):
                                                    if exists and \
                                                       self.is_step_selected(
                                                           "relation", nt):
                                                        exp.add(nt)
                                            self._relation_nodes_expected = exp
                                        self._relation_nodes_done.add(node_tag)
                                        remaining = (
                                            self._relation_nodes_expected
                                            - self._relation_nodes_done
                                        )
                                        do_delete = not remaining
                                    if do_delete:
                                        ssh.run_amos_command_safe(
                                            f'!rm -rf "{remote_rel}"/*',
                                            node_name, timeout=30,
                                        )
                                        ui_cb(
                                            "All relation nodes done — "
                                            f"cleaned up remote {remote_rel}/"
                                        )
                                    else:
                                        ui_cb(
                                            "Relation done for this node; "
                                            "keeping remote folder (still "
                                            f"waiting on: {', '.join(sorted(remaining))})"
                                        )
                                except Exception as _clean_exc:
                                    ui_cb(
                                        f"(remote cleanup skipped: {_clean_exc})"
                                    )
                            elif key == "baseline":
                                # Multi-source fallback chain for baseline
                                # log parsing:
                                #
                                #   1. Wrapped l+/l- log from server
                                #      (already downloaded). Best case —
                                #      clean isolated log of just the
                                #      baseline run.
                                #   2. SESSION log file at
                                #      ``LOG/<node>/SESSION/SESSION_BASELINE_<node>.log``.
                                #      Captured locally by
                                #      ``start_step_log`` (tees every byte
                                #      from the SSH channel), so it always
                                #      contains the full live output —
                                #      even when moshell's l+/l- inside
                                #      the baseline script rebinds logging
                                #      away from our wrapper.
                                #
                                # Detection rule: we want a log that
                                # CONTAINS actual MO operations (the
                                # parser looks for "Total: N MOs
                                # attempted" / "MOs set" lines). Byte
                                # size alone isn't enough — sometimes
                                # the wrapped log captures just the
                                # script source being echoed back (lots
                                # of bytes, zero useful content) when
                                # the operator's baseline.mos does its
                                # own ``l-``/``l+`` right at the top.
                                # If the wrapped log has ZERO MO
                                # markers, we fall back to the SESSION
                                # log unconditionally.
                                MO_MARKERS = (
                                    "MOs attempted",
                                    "MOs set",
                                    "MOs unchanged",
                                )

                                def _has_mo_content(path: str) -> bool:
                                    try:
                                        if (not os.path.exists(path)
                                                or os.path.getsize(path) < 200):
                                            return False
                                        with open(path, "r",
                                                  encoding="utf-8",
                                                  errors="replace") as _fh:
                                            content = _fh.read()
                                        return any(
                                            m in content for m in MO_MARKERS
                                        )
                                    except Exception:
                                        return False

                                parse_sources = list(downloaded)
                                use_fallback = (
                                    not parse_sources
                                    or not any(
                                        _has_mo_content(p)
                                        for p in parse_sources
                                    )
                                )

                                if use_fallback:
                                    # Diagnostic — surface why we fell back
                                    for p in parse_sources:
                                        try:
                                            sz = (
                                                os.path.getsize(p)
                                                if os.path.exists(p) else 0
                                            )
                                            ui_cb(
                                                f"Wrapped log {os.path.basename(p)} "
                                                f"({sz} bytes) has no MO-set "
                                                "markers — likely captured "
                                                "only the script source echo."
                                            )
                                        except Exception:
                                            pass
                                    ui_cb(
                                        "Falling back to local SESSION log "
                                        "for baseline parsing."
                                    )
                                    if os.path.exists(session_log_path):
                                        parse_sources = [session_log_path]
                                        sz = os.path.getsize(session_log_path)
                                        ui_cb(
                                            f"Using session log "
                                            f"({sz} bytes): "
                                            f"{session_log_path}"
                                        )
                                    else:
                                        ui_cb(
                                            f"(session log also missing: "
                                            f"{session_log_path} — parser "
                                            "will get nothing)"
                                        )

                                try:
                                    # Serialize CPU-heavy parse + Excel
                                    # build (baseline output can be MB-
                                    # large) so parallel nodes don't all
                                    # peg the GIL and black-screen the UI.
                                    with self._heavy_lock:
                                        parsed_path = build_baseline_log_excel(
                                            parse_sources,
                                            self.log_dir,
                                            node_name,
                                            log_cb=ui_cb,
                                        )
                                        bsummary = parse_baseline_summary(
                                            parse_sources,
                                        )
                                    ui_cb(
                                        "Baseline log Excel created: "
                                        f"{os.path.basename(parsed_path)}"
                                    )
                                    try:
                                        if bsummary:
                                            ui_cb(
                                                f"Baseline summary: "
                                                f"{bsummary['total_commands']} commands — "
                                                f"{bsummary['success']} OK, "
                                                f"{bsummary['zero_mo']} zero MOs, "
                                                f"{bsummary['error']} errors"
                                            )
                                            self._set_step(
                                                node_tag, key,
                                                self._step_results.get(node_tag, {}).get(key, "done"),
                                                (
                                                    f"{bsummary['success']} OK / "
                                                    f"{bsummary['zero_mo']} zero / "
                                                    f"{bsummary['error']} err"
                                                ),
                                            )
                                    except Exception:
                                        pass
                                except Exception as parse_exc:
                                    ui_cb(
                                        f"(baseline log parsing skipped: {parse_exc})"
                                    )
                    except Exception as _exc:
                        ui_cb(f"(could not download moshell logs: {_exc})")

                self._save_step_log(step_num, node_name, log_suffix,
                                    node_tag=node_tag)
                step_num += 1

                # Operator pressed "Skip Node" on a dialog for this node
                # → abort all remaining steps for THIS node (other nodes
                # keep running).
                if node_tag in self._aborted_nodes:
                    ui_cb("Skip Node requested — aborting remaining steps "
                          "for this node.")
                    self._resumable_nodes.add(node_tag)
                    remaining = False
                    for rkey, rlabel, rapplies, _ in INTEGRATION_STEPS:
                        if rapplies not in ("both", node_type):
                            continue
                        if not self.is_step_selected(rkey, node_tag):
                            continue
                        if remaining:
                            self._set_step(
                                node_tag, rkey, "skip",
                                "Skipped (node aborted by operator)")
                        if rkey == key:
                            remaining = True
                    break

                if stopped:
                    ui_cb("Stopping remaining steps due to failure.")
                    self._resumable_nodes.add(node_tag)
                    remaining = False
                    for rkey, rlabel, rapplies, _ in INTEGRATION_STEPS:
                        if rapplies not in ("both", node_type):
                            continue
                        if not self.is_step_selected(rkey, node_tag):
                            continue
                        if remaining:
                            self._set_step(
                                node_tag, rkey, "skip",
                                "Skipped (previous step failed)")
                        if rkey == key:
                            remaining = True
                    break
        finally:
            elapsed = time.monotonic() - node_start
            self._node_durations[node_tag] = elapsed
            if in_amos:
                try:
                    ssh.exit_amos()
                except Exception:
                    pass
            try:
                ssh.disconnect()
            except Exception:
                pass
            self._active_ssh.pop(node_tag, None)
            ui_cb("SSH disconnected.")
            # Show this node's Resume button the moment IT stops — don't
            # wait for the other nodes to finish (they run independently).
            self._update_resume_button(node_tag)

    # ── Navigation ───────────────────────────────────────────────
    def _force_disconnect(self) -> None:
        """Forcibly tear down any active SSH so blocked recv() unwinds."""
        for node_tag, ssh in list(self._active_ssh.items()):
            try:
                if getattr(ssh, "shell", None) is not None:
                    try:
                        ssh.shell.close()
                    except Exception:
                        pass
                if getattr(ssh, "client", None) is not None:
                    try:
                        t = ssh.client.get_transport()
                        if t is not None:
                            t.close()
                    except Exception:
                        pass
                    try:
                        ssh.client.close()
                    except Exception:
                        pass
                ssh._connected = False
            except Exception:
                pass
        self._active_ssh.clear()

    def _on_cancel(self, e):
        self.cancelled = True
        self._timer_running = False
        self._run_finished = True
        self._force_disconnect()
        self.status_text.value = "Cancelled"
        self.cancel_button.visible = False
        self.back_button.visible = True
        self.page.update()

    def _go_back(self, e):
        self.cancelled = True
        self._timer_running = False
        self._run_finished = True
        self._force_disconnect()
        self.page.go("/form")


# ── Step row widget ──────────────────────────────────────────────
class _StepRow:
    """Single checklist row: icon + label + status detail."""

    _ICONS = {
        "pending":  (ft.Icons.RADIO_BUTTON_UNCHECKED, TEXT_MUTED),
        "running":  (ft.Icons.HOURGLASS_TOP,          ACCENT),
        "done":     (ft.Icons.CHECK_CIRCLE,            SUCCESS),
        "warn":     (ft.Icons.WARNING_AMBER_ROUNDED,   ACCENT_WARM),
        "error":    (ft.Icons.ERROR,                   DANGER),
        "skip":     (ft.Icons.REMOVE_CIRCLE_OUTLINE,   TEXT_MUTED),
    }

    def __init__(self, label: str):
        # NOTE: we use a STATIC icon for the "running" state (an
        # hourglass), NOT an animated ft.ProgressRing. An indeterminate
        # ProgressRing repaints the Flutter renderer every frame
        # (~60 fps) for as long as it's visible — with several running
        # steps across 3 node columns that pegged a CPU core at 60%+.
        # A static icon lets the renderer go idle between actual
        # updates, dropping CPU to near-zero when nothing changes.
        self._icon = ft.Icon(ft.Icons.RADIO_BUTTON_UNCHECKED, size=18,
                             color=TEXT_MUTED)
        self._label = ft.Text(label, size=13, color=TEXT,
                              weight=ft.FontWeight.W_500)
        self._detail = ft.Text("", size=11, color=TEXT_MUTED)
        self.control = ft.Container(
            padding=ft.Padding.symmetric(horizontal=8, vertical=6),
            border_radius=10,
            content=ft.Row(
                [
                    self._icon,
                    self._label,
                    ft.Container(expand=True),
                    self._detail,
                ],
                spacing=8,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        )

    def set_state(self, state: str, detail: str = ""):
        icon_name, color = self._ICONS.get(state, self._ICONS["pending"])
        self._icon.name = icon_name
        self._icon.color = color
        self._icon.visible = True
        self._label.color = color if state == "error" else TEXT
        self._detail.value = detail
        self._detail.color = SUCCESS if state == "done" else (
            DANGER if state == "error" else (
                ACCENT_WARM if state == "warn" else TEXT_MUTED
            )
        )
        self.control.bgcolor = (
            ft.Colors.with_opacity(0.08, ACCENT)
            if state == "running" else None
        )
