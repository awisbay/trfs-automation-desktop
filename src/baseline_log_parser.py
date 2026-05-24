"""Baseline moshell log parser.

Builds a compact Excel workbook with a per-command breakdown and a summary
count of baseline command outcomes.
"""

from __future__ import annotations

import os
import re
from collections import Counter
from typing import Callable, Iterable, Optional

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LogCallback = Optional[Callable[[str], None]]

_PROMPT_PATTERN = re.compile(r"^([A-Z0-9_\-]+)>\s*(.*)$", re.IGNORECASE)
_TOTAL_PATTERN = re.compile(
    r"^Total:\s*(\d+)\s+MOs attempted,\s*(\d+)\s+MOs\s+(set|actioned)\s*$",
    re.IGNORECASE,
)
_ERROR_PATTERN = re.compile(r"^ERROR:\s*(.*)$", re.IGNORECASE)

_STATUS_COLORS = {
    "SUCCESS": "42FF00",
    "NO_CHANGE": "FFD966",
    "ALREADY_EXISTS": "F4B183",
    "ERROR": "FF7575",
    "UNKNOWN": "D9E2F3",
}
_EXCEL_CELL_LIMIT = 32767


def _log(log_cb: LogCallback, message: str) -> None:
    if log_cb:
        log_cb(message)


def _safe_excel_text(value: str) -> str:
    text = value or ""
    if len(text) <= _EXCEL_CELL_LIMIT:
        return text
    return text[: _EXCEL_CELL_LIMIT - 24] + "\n... [truncated for Excel]"


def _classify_block(command: str, lines: list[str]) -> dict:
    node = ""
    cmd_text = command.strip()
    attempted = None
    completed = None
    action = ""
    total_line = ""
    errors: list[str] = []

    prompt_match = _PROMPT_PATTERN.match(command)
    if prompt_match:
        node = prompt_match.group(1).strip()
        cmd_text = prompt_match.group(2).strip()

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        total_match = _TOTAL_PATTERN.match(line)
        if total_match:
            attempted = int(total_match.group(1))
            completed = int(total_match.group(2))
            action = total_match.group(3).lower()
            total_line = line
            continue
        error_match = _ERROR_PATTERN.match(line)
        if error_match:
            errors.append(error_match.group(1).strip())

    if attempted is not None and completed is not None:
        if attempted == 0 or completed == 0:
            report = f"0 MOs {action}"
            status = "NO_CHANGE"
        else:
            report = f"{completed} MOs {action}"
            status = "SUCCESS" if completed == attempted else "ERROR"
        return {
            "node": node,
            "command": cmd_text,
            "report": report,
            "status": status,
            "attempted": attempted,
            "completed": completed,
            "total_line": total_line,
            "error": "",
            "full_log": _safe_excel_text("\n".join(lines).strip()),
        }

    if errors:
        error_text = errors[0]
        if "already exists" in error_text.lower():
            status = "ALREADY_EXISTS"
            report = "MO already exists"
        else:
            status = "ERROR"
            report = error_text
        return {
            "node": node,
            "command": cmd_text,
            "report": report,
            "status": status,
            "attempted": 0,
            "completed": 0,
            "total_line": "",
            "error": error_text,
            "full_log": _safe_excel_text("\n".join(lines).strip()),
        }

    return {
        "node": node,
        "command": cmd_text,
        "report": "No result summary",
        "status": "UNKNOWN",
        "attempted": 0,
        "completed": 0,
        "total_line": "",
        "error": "",
        "full_log": _safe_excel_text("\n".join(lines).strip()),
    }


def _parse_single_log(log_path: str) -> list[dict]:
    rows: list[dict] = []
    current_command = ""
    current_lines: list[str] = []

    with open(log_path, "r", encoding="utf-8", errors="ignore") as handle:
        for raw_line in handle:
            prompt_match = _PROMPT_PATTERN.match(raw_line.strip())
            if prompt_match:
                command_text = prompt_match.group(2).strip()
                if current_command:
                    rows.append(_classify_block(current_command, current_lines))
                    current_lines = []
                current_command = raw_line.strip() if command_text else ""
                continue

            if current_command:
                current_lines.append(raw_line.rstrip("\n"))

    if current_command:
        rows.append(_classify_block(current_command, current_lines))

    return [
        row
        for row in rows
        if row.get("command") and row.get("status") != "UNKNOWN"
    ]


def _write_detail_sheet(workbook: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = workbook.active
    ws.title = "Baseline Log"

    headers = [
        "Node",
        "Command",
        "Report",
        "Status",
        "Attempted",
        "Completed",
        "Error",
        "Total Line",
        "Full Log",
    ]
    widths = [28, 72, 28, 18, 12, 12, 42, 36, 90]

    for idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row_idx, row in enumerate(rows, start=2):
        values = [
            row["node"],
            row["command"],
            row["report"],
            row["status"],
            row["attempted"],
            row["completed"],
            row["error"],
            row["total_line"],
            row["full_log"],
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value).font = Font(size=9)
        ws.cell(row=row_idx, column=3).fill = PatternFill(
            start_color=_STATUS_COLORS.get(row["status"], _STATUS_COLORS["UNKNOWN"]),
            end_color=_STATUS_COLORS.get(row["status"], _STATUS_COLORS["UNKNOWN"]),
            fill_type="solid",
        )


def _write_summary_sheet(workbook: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = workbook.create_sheet(title="Summary")
    headers = ["Report", "Status", "Count"]
    widths = [32, 18, 12]

    for idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ws.column_dimensions[get_column_letter(idx)].width = width

    counts = Counter((row["report"], row["status"]) for row in rows)
    for row_idx, ((report, status), count) in enumerate(sorted(counts.items()), start=2):
        ws.cell(row=row_idx, column=1, value=report)
        ws.cell(row=row_idx, column=2, value=status)
        ws.cell(row=row_idx, column=3, value=count)
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        ws.cell(row=row_idx, column=1).fill = PatternFill(
            start_color=_STATUS_COLORS.get(status, _STATUS_COLORS["UNKNOWN"]),
            end_color=_STATUS_COLORS.get(status, _STATUS_COLORS["UNKNOWN"]),
            fill_type="solid",
        )


def parse_baseline_summary(log_files: Iterable[str]) -> dict:
    """Parse baseline log files and return summary counters.

    Returns a dict with keys:
        total_commands   — total number of parsed command blocks
        success          — commands where all MOs were set/actioned
        zero_mo          — commands where 0 MOs were set (no change)
        already_exists   — commands with "MO already exists" soft warning
        error            — commands with real errors
        unknown          — commands with no clear result
    """
    existing_logs = [
        path
        for path in log_files
        if os.path.exists(path) and os.path.splitext(path)[1].lower() in {".log", ".txt"}
    ]
    if not existing_logs:
        return {}

    rows: list[dict] = []
    for log_path in sorted(existing_logs):
        rows.extend(_parse_single_log(log_path))

    if not rows:
        return {}

    counts = Counter(row["status"] for row in rows)
    return {
        "total_commands": len(rows),
        "success": counts.get("SUCCESS", 0),
        "zero_mo": counts.get("NO_CHANGE", 0),
        "already_exists": counts.get("ALREADY_EXISTS", 0),
        "error": counts.get("ERROR", 0),
        "unknown": counts.get("UNKNOWN", 0),
    }


def build_baseline_log_excel(
    log_files: Iterable[str],
    output_dir: str,
    node_name: str,
    log_cb: LogCallback = None,
) -> str:
    """Parse downloaded baseline moshell logs into an Excel workbook."""
    existing_logs = [
        path
        for path in log_files
        if os.path.exists(path) and os.path.splitext(path)[1].lower() in {".log", ".txt"}
    ]
    if not existing_logs:
        raise ValueError("No downloaded .log files were available to parse.")

    os.makedirs(output_dir, exist_ok=True)
    rows: list[dict] = []

    for idx, log_path in enumerate(sorted(existing_logs), start=1):
        _log(log_cb, f"Parsing baseline log {idx}/{len(existing_logs)}: {os.path.basename(log_path)}")
        rows.extend(_parse_single_log(log_path))

    if not rows:
        raise ValueError("Baseline log parser did not find any executable command blocks.")

    workbook = openpyxl.Workbook()
    _write_detail_sheet(workbook, rows)
    _write_summary_sheet(workbook, rows)

    safe_node = re.sub(r"[^A-Za-z0-9_\-]+", "_", node_name).strip("_") or "BASELINE"
    output_path = os.path.join(output_dir, f"BASELINE_{safe_node}_PARSED.xlsx")
    workbook.save(output_path)
    return output_path
