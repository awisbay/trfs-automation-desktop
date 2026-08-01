"""
NodeCraft License Generator — Local Web App (ADMIN ONLY)

Run:
    python app.py

Then open:  http://localhost:5555

Features:
  - Single license generation (form)
  - Bulk generation from Excel/CSV upload
  - Download all generated keys as Excel
"""
import base64
import io
import json
import os
import sys
from datetime import datetime

from flask import Flask, render_template, request, send_file, jsonify, flash, redirect, url_for
from openpyxl import Workbook, load_workbook
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey

# ── Import the private key from keygen (keep SECRET) ─────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from nodecraft_keygen import generate_license, _PRIVATE_KEY_B64  # noqa

app = Flask(__name__)
app.secret_key = "nodecraft-license-gen-local-only"

# ── Per-feature licensing ────────────────────────────────────────
# Feature keys MUST match src/license_manager.py (ALL_FEATURES).
ALL_FEATURES = ["integration", "terminal", "audit", "trfs", "cutover"]
FEATURE_LABELS = {
    "integration": "Integration",
    "terminal": "Terminal",
    "audit": "CDD Audit",
    "trfs": "TRFS",
    "cutover": "Cut Over",
}


def _normalize_features(raw) -> list:
    """Accept a list, or a comma/semicolon/space/pipe separated string,
    or the wildcards '*'/'all', and return a clean ordered list of the
    known feature keys. Empty / unknown input → all features (so a blank
    column in a bulk sheet still yields a fully-featured key, matching
    the previous behaviour)."""
    if raw is None:
        return list(ALL_FEATURES)
    if isinstance(raw, (list, tuple, set)):
        items = [str(x) for x in raw]
    else:
        s = str(raw).strip()
        if not s or s.lower() in ("*", "all"):
            return list(ALL_FEATURES)
        for sep in (";", "|", " "):
            s = s.replace(sep, ",")
        items = s.split(",")
    picked = []
    for it in items:
        k = it.strip().lower()
        if k in ALL_FEATURES and k not in picked:
            picked.append(k)
    return picked or list(ALL_FEATURES)


def build_license(user: str, expires: str, note: str, hostname: str,
                  features: list) -> str:
    """Sign a license payload directly (Ed25519), embedding the enabled
    feature list. Produces the same wire format the desktop app verifies:
    base64( payload_json_bytes + b"|SIG|" + signature )."""
    from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey

    payload = {
        "product": "NodeCraft",
        "user": user,
        "expires": expires,
        "issued": datetime.now().strftime("%Y-%m-%d"),
        "note": note or "",
        "features": list(features),
    }
    if hostname:
        payload["hostname"] = hostname

    payload_bytes = json.dumps(payload, separators=(",", ":"),
                               sort_keys=True).encode("utf-8")
    priv = Ed25519PrivateKey.from_private_bytes(base64.b64decode(_PRIVATE_KEY_B64))
    signature = priv.sign(payload_bytes)
    return base64.b64encode(payload_bytes + b"|SIG|" + signature).decode("ascii")


def parse_date(raw: str) -> str:
    """Accept YYYYMMDD or YYYY-MM-DD, return YYYY-MM-DD."""
    raw = str(raw).strip().replace("/", "-").replace(".", "-")
    # Try YYYYMMDD
    if len(raw) == 8 and raw.isdigit():
        return f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
    # Try YYYY-MM-DD
    try:
        d = datetime.strptime(raw, "%Y-%m-%d")
        return d.strftime("%Y-%m-%d")
    except ValueError:
        pass
    # Try DD-MM-YYYY
    try:
        d = datetime.strptime(raw, "%d-%m-%Y")
        return d.strftime("%Y-%m-%d")
    except ValueError:
        pass
    raise ValueError(f"Invalid date format: {raw!r}. Use YYYYMMDD or YYYY-MM-DD.")


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/generate_single", methods=["POST"])
def generate_single():
    name = request.form.get("name", "").strip()
    hostname = request.form.get("hostname", "").strip()
    expires = request.form.get("expires", "").strip()
    note = request.form.get("note", "").strip()
    # Checkboxes: features=integration&features=terminal&...
    features = _normalize_features(request.form.getlist("features"))

    if not name or not expires:
        return jsonify({"ok": False, "error": "Name and Expires are required."})

    try:
        exp_date = parse_date(expires)
    except ValueError as ex:
        return jsonify({"ok": False, "error": str(ex)})

    try:
        key = build_license(user=name, expires=exp_date, note=note,
                            hostname=hostname, features=features)
    except Exception as ex:
        return jsonify({"ok": False, "error": f"Generation failed: {ex}"})

    return jsonify({
        "ok": True,
        "name": name,
        "hostname": hostname or "(any PC)",
        "expires": exp_date,
        "note": note,
        "features": ", ".join(FEATURE_LABELS.get(f, f) for f in features),
        "key": key,
    })


@app.route("/generate_bulk", methods=["POST"])
def generate_bulk():
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "No file uploaded."})

    fname = file.filename.lower()
    rows = []

    try:
        if fname.endswith(".csv"):
            import csv
            text = file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append(r)
        elif fname.endswith((".xlsx", ".xlsm")):
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value is not None else ""
                       for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})
        else:
            return jsonify({"ok": False, "error": "Please upload .xlsx or .csv"})
    except Exception as ex:
        return jsonify({"ok": False, "error": f"Failed to read file: {ex}"})

    results = []
    for i, r in enumerate(rows, 1):
        # Accept flexible column names
        name = str(r.get("name") or r.get("Name") or r.get("user") or "").strip()
        hostname = str(r.get("hostname") or r.get("Hostname") or r.get("host") or "").strip()
        expires_raw = r.get("expires") or r.get("Expires") or r.get("expired") \
            or r.get("Expired") or r.get("expired date") or r.get("Expired Date") or ""
        note = str(r.get("note") or r.get("Note") or "").strip()
        # Optional "features" column: comma/space separated keys, or
        # blank / "all" / "*" for every feature.
        features_raw = r.get("features") or r.get("Features") or r.get("feature") \
            or r.get("Feature") or ""
        features = _normalize_features(features_raw)

        if not name or not expires_raw:
            results.append({
                "row": i, "name": name, "hostname": hostname,
                "expires": "", "key": "", "error": "Missing name or expires",
            })
            continue

        try:
            exp_date = parse_date(str(expires_raw))
            key = build_license(user=name, expires=exp_date, note=note,
                                hostname=hostname, features=features)
            results.append({
                "row": i, "name": name, "hostname": hostname or "(any PC)",
                "expires": exp_date, "note": note,
                "features": ", ".join(FEATURE_LABELS.get(f, f) for f in features),
                "key": key, "error": "",
            })
        except Exception as ex:
            results.append({
                "row": i, "name": name, "hostname": hostname,
                "expires": str(expires_raw), "key": "", "error": str(ex),
            })

    return jsonify({"ok": True, "results": results})


@app.route("/download_excel", methods=["POST"])
def download_excel():
    """Receive JSON of results and return an Excel file."""
    data = request.get_json(silent=True) or {}
    results = data.get("results", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Licenses"

    headers = ["Name", "Hostname", "Expires", "Note", "Features", "License Key", "Error"]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="39C0BA")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        ws.append([
            r.get("name", ""),
            r.get("hostname", ""),
            r.get("expires", ""),
            r.get("note", ""),
            r.get("features", ""),
            r.get("key", ""),
            r.get("error", ""),
        ])

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 90
    ws.column_dimensions["G"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"NodeCraft_Licenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/template.xlsx")
def template():
    """Download a blank bulk-upload template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Licenses"
    ws.append(["Name", "Hostname", "Expires", "Note", "Features"])
    ws.append(["John Doe", "DESKTOP-ABC123", "20270101", "Q1 batch", "all"])
    ws.append(["Jane Smith", "", "20261231", "Portable — any PC", "integration, terminal"])
    ws.append(["Bob Audit", "", "20261231", "Audit only", "audit"])

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="39C0BA")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDE", [20, 20, 14, 30, 28]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="nodecraft_bulk_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    print()
    print("=" * 60)
    print("  NodeCraft License Generator — Local Web")
    print("=" * 60)
    print(f"  Open in browser:  http://localhost:5555")
    print(f"  Press Ctrl+C to stop.")
    print("=" * 60)
    print()
    app.run(host="127.0.0.1", port=5555, debug=False)
